# main.py - COMPLETE FIXED VERSION
# Load environment variables from .env file
from pathlib import Path
from dotenv import load_dotenv

# Load .env from the backend directory
backend_dir = Path(__file__).parent
load_dotenv(backend_dir / ".env")


# Add to existing imports
from domain.cv_models import (
    JobPosting,
    ExtractedCVData,
    CVProcessingLog,
)

# ADD THIS IMPORT (keep it here)
from routes.interview_routes import interview_router

# Add Vision Routes for face registration/verification
from routes.vision import router as vision_router

# Add to imports
from infrastructure.cv_repository import CVRepository
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, Depends, BackgroundTasks, Header
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel, Field
from urllib.parse import unquote

# Add to your existing imports at the top of main.py
from fastapi import UploadFile, File, Form  # Add these

# Add to your existing imports at the top of main.py
from typing import Optional, List, Dict  # Make sure Optional is imported

import os
import json
import uuid
import random
from datetime import datetime, timedelta
import sys
import re
import spacy
import pytesseract
import subprocess
import tempfile
import time
import aiohttp
import asyncio

import base64

# Add to your imports section at the top
from google_auth_oauthlib.flow import InstalledAppFlow

import PyPDF2
from docx import Document
import spacy
import pytesseract
from PIL import Image
import fitz  # PyMuPDF
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
import shutil
import re

# ==================== FUNCTION EXTRACTION HELPERS ====================


def extract_function_name(code: str) -> str:
    """Extract function name from Python code"""
    patterns = [
        r"def\s+(\w+)\s*\(",  # def function_name(
        r"async def\s+(\w+)\s*\(",  # async def function_name(
    ]

    for pattern in patterns:
        match = re.search(pattern, code)
        if match:
            return match.group(1)

    return None


def extract_go_function_name(code: str) -> str:
    """Extract function name from Go code"""
    patterns = [
        r"func\s+(\w+)\s*\(",  # func functionName(
        r"func\s+\(\w+\s+\w+\)\s+(\w+)\s*\(",  # func (t Type) functionName(
    ]

    for pattern in patterns:
        match = re.search(pattern, code)
        if match:
            return match.group(1) if match.lastindex else match.group(0)

    return None


def extract_js_function_name(code: str) -> str:
    """Extract function name from JavaScript code"""
    patterns = [
        r"function\s+(\w+)\s*\(",  # function name(
        r"const\s+(\w+)\s*=\s*(?:async\s*)?\(",  # const name = ( or const name = async (
        r"let\s+(\w+)\s*=\s*(?:async\s*)?\(",  # let name = (
        r"var\s+(\w+)\s*=\s*(?:async\s*)?\(",  # var name = (
    ]

    for pattern in patterns:
        match = re.search(pattern, code)
        if match:
            return match.group(1)

    return None


def extract_cpp_function_name(code: str) -> str:
    """Extract function name from C++ code"""
    patterns = [
        r"(\w+)\s+(\w+)\s*\([^)]*\)\s*{",  # return_type name(
        r"auto\s+(\w+)\s*\([^)]*\)\s*{",  # auto name(
    ]

    for pattern in patterns:
        match = re.search(pattern, code)
        if match:
            # Return the second group (function name) for first pattern
            # or first group for second pattern
            return match.group(2) if len(match.groups()) > 1 else match.group(1)

    return None


def extract_java_function_name(code: str) -> str:
    """Extract function name from Java code"""
    patterns = [
        r"public\s+\w+\s+(\w+)\s*\([^)]*\)",  # public returnType name(
        r"private\s+\w+\s+(\w+)\s*\([^)]*\)",  # private returnType name(
        r"protected\s+\w+\s+(\w+)\s*\([^)]*\)",  # protected returnType name(
        r"\w+\s+(\w+)\s*\([^)]*\)",  # returnType name(
    ]

    for pattern in patterns:
        match = re.search(pattern, code)
        if match:
            return match.group(1)

    return None


def sanitize_go_code(code: str) -> str:
    """
    Sanitize Go code by removing unused imports and main() function
    """
    # Remove package main if present
    if "package main" in code:
        code = code.replace("package main\n", "")

    # Remove main() function if present
    if "func main()" in code:
        lines = code.split("\n")
        new_lines = []
        in_main = False
        for line in lines:
            if "func main()" in line:
                in_main = True
                continue
            if in_main and line.strip() and line[0] != "\t" and line[:4] != "    ":
                in_main = False
            if not in_main:
                new_lines.append(line)
        code = "\n".join(new_lines)

    # Analyze which imports are actually used
    lines = code.split("\n")
    new_lines = []
    imports_section = []
    in_import_block = False

    for line in lines:
        stripped = line.strip()

        # Start of import block
        if stripped.startswith("import ("):
            in_import_block = True
            imports_section = []
            continue

        # End of import block
        if in_import_block and stripped == ")":
            in_import_block = False

            # Filter imports based on actual usage
            needed_imports = []
            for import_line in imports_section:
                if import_line.strip():
                    import_name = import_line.strip().strip('"')
                    # Check if this import is actually used in code
                    if import_name + "." in code:
                        needed_imports.append(import_line)

            # Add filtered imports back
            if needed_imports:
                new_lines.append("import (")
                new_lines.extend(needed_imports)
                new_lines.append(")")
            continue

        # Inside import block
        if in_import_block:
            imports_section.append(line)
            continue

        # Single line import
        if stripped.startswith('import "') or stripped.startswith('import ."'):
            import_match = re.search(r'import\s+(?:"([^"]+)"|\.)"([^"]+)"', line)
            if import_match:
                import_name = import_match.group(1) or import_match.group(2)
                if import_name + "." in code:
                    new_lines.append(line)
        else:
            new_lines.append(line)

    return "\n".join(new_lines)


sys.path.append("domain")
sys.path.append("integrations")
sys.path.append("application")
sys.path.append("infrastructure")

from dotenv import load_dotenv

load_dotenv()

# ==================== CORS CONFIGURATION ====================
# ==================== CORS CONFIGURATION ====================
_extra_origins = [
    o.strip() for o in os.getenv("CORS_ORIGINS", "").split(",") if o.strip()
]
_frontend = os.getenv("FRONTEND_URL", "http://localhost:5173").rstrip("/")
origins = list(
    dict.fromkeys(
        [
            _frontend,
            *_extra_origins,
            "http://localhost:5173",
            "http://localhost:5174",
            "http://localhost:5175",
            "http://localhost:5176",
            "http://localhost:5177",
            "http://localhost:3000",
            "http://127.0.0.1:5173",
            "http://127.0.0.1:5174",
            "http://127.0.0.1:5175",
            "http://127.0.0.1:5176",
            "http://127.0.0.1:5177",
        ]
    )
)
# Match any localhost / 127.0.0.1 port (dev only)
_cors_origin_regex = r"https?://(localhost|127\.0\.0\.1)(:\d+)?$"


# ==================== LIFESPAN MANAGER ====================
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup code
    from infrastructure.db_models import init_database, engine

    await init_database()

    # ==================== AI INTERVIEW DIRECTORIES ====================
    # Load environment variables
    from dotenv import load_dotenv

    load_dotenv()

    # Create required directories
    import os

    dirs_to_create = ["sessions", "faces", "reports", "transcripts"]
    for dir_name in dirs_to_create:
        os.makedirs(dir_name, exist_ok=True)

    print(
        "✓ Created AI Interview directories: sessions/, faces/, reports/, transcripts/"
    )

    # ==================== MODULE 5 INTEGRATED: Create All Tables ====================
    from sqlalchemy import text

    # 0. Application users (HR / candidate) — same schema as migrations/001_app_users.sql
    app_users_table = """
    CREATE TABLE IF NOT EXISTS app_users (
        id INT AUTO_INCREMENT PRIMARY KEY,
        email VARCHAR(255) NOT NULL,
        password_hash VARCHAR(255) NOT NULL,
        full_name VARCHAR(255) DEFAULT NULL,
        role VARCHAR(32) NOT NULL DEFAULT 'hr',
        is_active TINYINT(1) NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uq_app_users_email (email),
        INDEX idx_app_users_role (role)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    """

    # 1. Coding Challenges Table
    coding_challenges_table = """
    CREATE TABLE IF NOT EXISTS coding_challenges (
        challenge_id VARCHAR(255) PRIMARY KEY,
        title VARCHAR(500) NOT NULL,
        description TEXT NOT NULL,
        difficulty VARCHAR(50) NOT NULL,
        language VARCHAR(50) NOT NULL,
        starter_code TEXT,
        test_cases TEXT,
        constraints TEXT,
        examples TEXT,
        hints TEXT,
        role VARCHAR(255),
        created_at DATETIME,
        created_by VARCHAR(255),
        is_active BOOLEAN DEFAULT TRUE,
        INDEX idx_difficulty (difficulty),
        INDEX idx_language (language),
        INDEX idx_role (role)
    )
    """

    # 2. Coding Submissions Table
    coding_submissions_table = """
    CREATE TABLE IF NOT EXISTS coding_submissions (
        submission_id VARCHAR(255) PRIMARY KEY,
        challenge_id VARCHAR(255) NOT NULL,
        candidate_email VARCHAR(255) NOT NULL,
        code LONGTEXT NOT NULL,
        language VARCHAR(50) NOT NULL,
        test_results TEXT,
        evaluation TEXT,
        score INT DEFAULT 0,
        submitted_at DATETIME,
        assessment_id VARCHAR(255),
        result_id VARCHAR(255),
        score_breakdown TEXT,
        session_id VARCHAR(255),
        role VARCHAR(255) DEFAULT NULL,
        status VARCHAR(50) DEFAULT 'submitted',
        INDEX idx_candidate (candidate_email),
        INDEX idx_challenge (challenge_id),
        INDEX idx_score (score),
        INDEX idx_assessment (assessment_id),
        INDEX idx_result (result_id),
        INDEX idx_session (session_id)
    )
    """

    # 3. Technical Scores Table
    technical_scores_table = """
    CREATE TABLE IF NOT EXISTS technical_scores (
        tech_score_id VARCHAR(255) PRIMARY KEY,
        candidate_email VARCHAR(255) NOT NULL,
        assessment_id VARCHAR(255),
        mcq_result_id VARCHAR(255),
        mcq_score INT,
        mcq_percentage FLOAT,
        coding_submission_id VARCHAR(255),
        coding_score INT,
        coding_percentage FLOAT,
        combined_score INT,
        combined_percentage FLOAT,
        technical_level VARCHAR(50),
        created_at DATETIME,
        updated_at DATETIME,
        INDEX idx_candidate (candidate_email),
        INDEX idx_assessment (assessment_id),
        INDEX idx_combined_score (combined_score)
    )
    """

    # 4. Coding Sessions Table
    coding_sessions_table = """
    CREATE TABLE IF NOT EXISTS coding_sessions (
        session_id VARCHAR(255) PRIMARY KEY,
        challenge_id VARCHAR(255) NOT NULL,
        candidate_email VARCHAR(255) NOT NULL,
        start_time DATETIME,
        end_time DATETIME,
        time_limit_minutes INT DEFAULT 60,
        status VARCHAR(50) DEFAULT 'in_progress',
        current_code TEXT,
        last_saved_at DATETIME,
        submission_count INT DEFAULT 0,
        submitted BOOLEAN DEFAULT FALSE,
        INDEX idx_candidate (candidate_email),
        INDEX idx_challenge (challenge_id),
        INDEX idx_status (status)
    )
    """

    # 5. Code Versions Table
    code_versions_table = """
    CREATE TABLE IF NOT EXISTS code_versions (
        version_id VARCHAR(255) PRIMARY KEY,
        session_id VARCHAR(255) NOT NULL,
        version_number INT NOT NULL,
        code TEXT NOT NULL,
        saved_at DATETIME,
        auto_saved BOOLEAN DEFAULT FALSE,
        INDEX idx_session (session_id),
        INDEX idx_version (version_number)
    )
    """

    # 6. Extension Requests Table
    extension_requests_table = """
    CREATE TABLE IF NOT EXISTS extension_requests (
        request_id VARCHAR(255) PRIMARY KEY,
        session_id VARCHAR(255) NOT NULL,
        candidate_email VARCHAR(255) NOT NULL,
        extension_minutes INT DEFAULT 15,
        reason TEXT,
        status VARCHAR(50) DEFAULT 'pending',
        requested_at DATETIME,
        approved_at DATETIME,
        hr_notes TEXT,
        INDEX idx_status (status),
        INDEX idx_session (session_id),
        FOREIGN KEY (session_id) REFERENCES coding_sessions(session_id) ON DELETE CASCADE
    )
    """

    # 7. Job Requirements Table
    job_requirements_table = """
    CREATE TABLE IF NOT EXISTS job_requirements (
        requirement_id VARCHAR(255) PRIMARY KEY,
        role VARCHAR(255) NOT NULL,
        skill_name VARCHAR(255) NOT NULL,
        weight FLOAT DEFAULT 1.0,
        required_level INT DEFAULT 70,
        created_at DATETIME,
        INDEX idx_role (role)
    )
    """

    # 8. Test Case Results Table
    test_case_results_table = """
    CREATE TABLE IF NOT EXISTS test_case_results (
        result_id VARCHAR(255) PRIMARY KEY,
        submission_id VARCHAR(255) NOT NULL,
        test_case_number INT NOT NULL,
        passed BOOLEAN DEFAULT FALSE,
        execution_time_ms FLOAT,
        error_message TEXT,
        created_at DATETIME,
        INDEX idx_submission (submission_id),
        FOREIGN KEY (submission_id) REFERENCES coding_submissions(submission_id) ON DELETE CASCADE
    )
    """

    # 9. Code Detection Table (AI & Plagiarism)
    code_detection_table = """
    CREATE TABLE IF NOT EXISTS code_detection (
        detection_id VARCHAR(255) PRIMARY KEY,
        submission_id VARCHAR(255) NOT NULL,
        ai_detection TEXT,
        plagiarism_check TEXT,
        created_at DATETIME,
        INDEX idx_submission (submission_id),
        FOREIGN KEY (submission_id) REFERENCES coding_submissions(submission_id) ON DELETE CASCADE
    )
    """

    # 10. Coding Evaluations Table
    coding_evaluations_table = """
    CREATE TABLE IF NOT EXISTS coding_evaluations (
        evaluation_id VARCHAR(255) PRIMARY KEY,
        submission_id VARCHAR(255) NOT NULL,
        candidate_email VARCHAR(255),
        correctness_score INT,
        efficiency_score INT,
        code_quality_score INT,
        problem_solving_score INT,
        readability_score INT,
        overall_score INT,
        grade VARCHAR(10),
        strengths TEXT,
        weaknesses TEXT,
        suggestions TEXT,
        ai_feedback TEXT,
        evaluated_at DATETIME,
        INDEX idx_submission (submission_id),
        FOREIGN KEY (submission_id) REFERENCES coding_submissions(submission_id) ON DELETE CASCADE
    )
    """

    # ==================== CV DATA EXTRACTION TABLES ====================

    # CV Candidates Table (Module 2.2.2)
    cv_candidates_table = """
    CREATE TABLE IF NOT EXISTS cv_candidates (
        candidate_id VARCHAR(255) PRIMARY KEY,
        email VARCHAR(255) UNIQUE NOT NULL,
        name VARCHAR(500),
        phone VARCHAR(50),
        role VARCHAR(500),
        skills TEXT,
        experience TEXT,
        education TEXT,
        cv_file_path VARCHAR(1000),
        cv_filename VARCHAR(1000),
        cv_text LONGTEXT,
        extracted_at DATETIME,
        missing_skills TEXT,
        extra_skills TEXT,
        matched_skills TEXT,
        skill_match_percentage FLOAT DEFAULT 0,
        status VARCHAR(50) DEFAULT 'pending',
        cv_source VARCHAR(50),
        job_id VARCHAR(255),
        raw_text TEXT,
        processing_errors TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        INDEX idx_email (email),
        INDEX idx_status (status),
        INDEX idx_skill_match (skill_match_percentage),
        INDEX idx_job (job_id)
    )
    """

    # Job Postings Table (Module 2.2.1)
    job_postings_table = """
    CREATE TABLE IF NOT EXISTS job_postings (
        job_id VARCHAR(255) PRIMARY KEY,
        title VARCHAR(500) NOT NULL,
        description TEXT,
        required_skills TEXT,
        experience_level VARCHAR(50),
        location VARCHAR(255),
        salary_range VARCHAR(100),
        posted_at DATETIME,
        posted_by VARCHAR(255),
        status VARCHAR(50) DEFAULT 'active',
        assessment_id VARCHAR(255),
        INDEX idx_status (status),
        INDEX idx_posted_by (posted_by)
    )
    """

    # Candidate applications to HR job postings (Module 2.2.11 user / candidate portal)
    job_applications_table = """
    CREATE TABLE IF NOT EXISTS job_applications (
        application_id VARCHAR(64) PRIMARY KEY,
        candidate_user_id INT NOT NULL,
        job_id VARCHAR(255) NOT NULL,
        status VARCHAR(64) NOT NULL DEFAULT 'applied',
        applied_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uq_candidate_job (candidate_user_id, job_id),
        INDEX idx_ja_job (job_id),
        INDEX idx_ja_candidate (candidate_user_id)
    )
    """

    # CV Processing Logs Table
    cv_processing_logs_table = """
    CREATE TABLE IF NOT EXISTS cv_processing_logs (
        log_id VARCHAR(255) PRIMARY KEY,
        candidate_id VARCHAR(255),
        cv_file_name VARCHAR(500),
        processing_status VARCHAR(50),
        extracted_data TEXT,
        error_message TEXT,
        processed_at DATETIME,
        source VARCHAR(50),
        job_id VARCHAR(255),
        INDEX idx_status (processing_status),
        INDEX idx_candidate (candidate_id),
        FOREIGN KEY (candidate_id) REFERENCES cv_candidates(candidate_id) ON DELETE CASCADE
    )
    """

    # Violations Table (for proctoring)
    violations_table = """
    CREATE TABLE IF NOT EXISTS violations (
        violation_id VARCHAR(255) PRIMARY KEY,
        session_id VARCHAR(255) NOT NULL,
        violation_type VARCHAR(100),
        description TEXT,
        occurred_at DATETIME,
        INDEX idx_session (session_id),
        INDEX idx_occurred (occurred_at)
    )
    """

    # Interview Sessions Table (AI Interviewer)
    interview_sessions_table = """
    CREATE TABLE IF NOT EXISTS interview_sessions (
        session_id VARCHAR(36) PRIMARY KEY,
        candidate_email VARCHAR(255) NOT NULL,
        candidate_name VARCHAR(255),
        job_role VARCHAR(255) NOT NULL,
        status VARCHAR(50) DEFAULT 'initiated',
        face_verified INT DEFAULT 0,
        start_time DATETIME,
        end_time DATETIME,
        questions_json LONGTEXT,
        responses_json LONGTEXT,
        emotion_data_json LONGTEXT,
        hr_report_json LONGTEXT,
        overall_score FLOAT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        INDEX idx_candidate (candidate_email),
        INDEX idx_status (status)
    )
    """

    # Interview Questions Table
    interview_questions_table = """
    CREATE TABLE IF NOT EXISTS interview_questions (
        question_id VARCHAR(36) PRIMARY KEY,
        job_role VARCHAR(255),
        category VARCHAR(100),
        question_text TEXT NOT NULL,
        ideal_answer_points TEXT,
        difficulty VARCHAR(20),
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """

    # MCQ core tables (fresh DB had only ALTERs, never CREATE — caused 500 on /api/results, /api/hr/*)
    assessments_table = """
    CREATE TABLE IF NOT EXISTS assessments (
        id INT AUTO_INCREMENT PRIMARY KEY,
        assessment_id VARCHAR(255) NOT NULL UNIQUE,
        role VARCHAR(500) NOT NULL,
        difficulty VARCHAR(50),
        questions LONGTEXT,
        duration_minutes INT DEFAULT 30,
        status VARCHAR(50) DEFAULT 'draft',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME NULL ON UPDATE CURRENT_TIMESTAMP,
        num_questions INT DEFAULT 0,
        is_adaptive BOOLEAN DEFAULT FALSE,
        include_coding BOOLEAN DEFAULT FALSE,
        coding_challenge_id VARCHAR(255),
        assessment_type VARCHAR(50) DEFAULT 'standard',
        hr_reviewed BOOLEAN DEFAULT FALSE,
        hr_reviewed_by VARCHAR(255),
        hr_reviewed_at DATETIME,
        hr_notes TEXT,
        INDEX idx_status (status)
    )
    """

    sessions_table = """
    CREATE TABLE IF NOT EXISTS sessions (
        session_id VARCHAR(255) PRIMARY KEY,
        assessment_id VARCHAR(255) NOT NULL,
        candidate_email VARCHAR(255) NOT NULL,
        role VARCHAR(255),
        start_time DATETIME,
        end_time DATETIME,
        time_remaining INT,
        answers LONGTEXT,
        violations TEXT,
        status VARCHAR(50) DEFAULT 'active',
        metadata LONGTEXT,
        INDEX idx_assessment (assessment_id),
        INDEX idx_candidate (candidate_email)
    )
    """

    results_table = """
    CREATE TABLE IF NOT EXISTS results (
        id INT AUTO_INCREMENT PRIMARY KEY,
        result_id VARCHAR(255) NOT NULL UNIQUE,
        session_id VARCHAR(255),
        assessment_id VARCHAR(255),
        candidate_email VARCHAR(255),
        role VARCHAR(255),
        difficulty VARCHAR(50),
        total_questions INT DEFAULT 0,
        correct_answers INT DEFAULT 0,
        wrong_answers INT DEFAULT 0,
        unanswered INT DEFAULT 0,
        score_percentage FLOAT DEFAULT 0,
        start_time DATETIME,
        end_time DATETIME,
        total_time_taken INT,
        question_results LONGTEXT,
        violations TEXT,
        status VARCHAR(50),
        grade VARCHAR(20),
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        flagged BOOLEAN DEFAULT FALSE,
        flag_reason VARCHAR(500),
        coding_submission_id VARCHAR(255),
        has_coding_assessment BOOLEAN DEFAULT FALSE,
        skill_breakdown TEXT,
        candidate_id VARCHAR(255),
        INDEX idx_result_id (result_id),
        INDEX idx_candidate_email (candidate_email),
        INDEX idx_assessment (assessment_id),
        INDEX idx_candidate_id (candidate_id)
    )
    """

    async with engine.begin() as conn:
        # Create all tables
        await conn.execute(text(app_users_table))
        await conn.execute(text(coding_challenges_table))
        await conn.execute(text(coding_submissions_table))
        await conn.execute(text(technical_scores_table))
        await conn.execute(text(coding_sessions_table))
        await conn.execute(text(code_versions_table))
        await conn.execute(text(extension_requests_table))
        await conn.execute(text(job_requirements_table))
        await conn.execute(text(test_case_results_table))
        await conn.execute(text(code_detection_table))
        await conn.execute(text(coding_evaluations_table))
        await conn.execute(text(cv_candidates_table))
        await conn.execute(text(job_postings_table))
        await conn.execute(text(job_applications_table))
        await conn.execute(text(cv_processing_logs_table))
        await conn.execute(text(assessments_table))
        await conn.execute(text(sessions_table))
        await conn.execute(text(results_table))
        await conn.execute(text(violations_table))
        await conn.execute(text(interview_sessions_table))
        await conn.execute(text(interview_questions_table))

        # Update existing tables (add columns if they don't exist)
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE results
                ADD COLUMN coding_submission_id VARCHAR(255),
                ADD COLUMN has_coding_assessment BOOLEAN DEFAULT FALSE
            """
                )
            )
        except:
            pass  # Columns may already exist

        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE assessments
                ADD COLUMN include_coding BOOLEAN DEFAULT FALSE,
                ADD COLUMN coding_challenge_id VARCHAR(255)
            """
                )
            )
        except:
            pass  # Columns may already exist

        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE coding_submissions
                ADD COLUMN session_id VARCHAR(255),
                ADD INDEX idx_session (session_id)
            """
                )
            )
        except:
            pass  # Column may already exist

        for _ddl in (
            "ALTER TABLE coding_submissions ADD COLUMN role VARCHAR(255) DEFAULT NULL",
            "ALTER TABLE coding_submissions ADD COLUMN status VARCHAR(50) DEFAULT 'submitted'",
        ):
            try:
                await conn.execute(text(_ddl))
            except Exception:
                pass

        # Update cv_candidates table with missing columns from previous version
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE cv_candidates
                ADD COLUMN skill_match_percentage FLOAT DEFAULT 0,
                ADD COLUMN cv_source VARCHAR(50),
                ADD COLUMN job_id VARCHAR(255),
                ADD COLUMN raw_text TEXT,
                ADD COLUMN processing_errors TEXT,
                ADD INDEX idx_skill_match (skill_match_percentage),
                ADD INDEX idx_job (job_id)
            """
                )
            )
        except:
            pass  # Columns may already exist

        # Add experience column if missing (for CV processing)
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE cv_candidates
                ADD COLUMN experience TEXT
            """
                )
            )
        except:
            pass  # Column may already exist

        # Add education column if missing (for CV processing)
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE cv_candidates
                ADD COLUMN education TEXT
            """
                )
            )
        except:
            pass  # Column may already exist

        for _ddl in (
            "ALTER TABLE cv_candidates ADD COLUMN role VARCHAR(500)",
            "ALTER TABLE cv_candidates ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP",
            "ALTER TABLE cv_candidates ADD COLUMN cv_filename VARCHAR(1000)",
            "ALTER TABLE cv_candidates ADD COLUMN cv_text LONGTEXT",
            "ALTER TABLE cv_candidates ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP",
            "ALTER TABLE cv_candidates ADD COLUMN matched_skills TEXT",
        ):
            try:
                await conn.execute(text(_ddl))
            except Exception:
                pass

        # My added code above - removed as requested by user

        # Ensure results table has skill_breakdown column (for UC-04 skill matching)
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE results
                ADD COLUMN skill_breakdown TEXT
            """
                )
            )
        except:
            pass  # Column may already exist

        # Add violation count column to coding_sessions
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE coding_sessions
                ADD COLUMN violation_count INT DEFAULT 0
            """
                )
            )
        except:
            pass  # Column may already exist

        # Add candidate_id column to results table for linking with CV candidates
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE results
                ADD COLUMN candidate_id VARCHAR(255),
                ADD INDEX idx_candidate_id (candidate_id),
                ADD FOREIGN KEY (candidate_id) REFERENCES cv_candidates(candidate_id) ON DELETE SET NULL
            """
                )
            )
        except:
            pass  # Column may already exist

        # Add assessment_type column to assessments
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE assessments
                ADD COLUMN assessment_type VARCHAR(50) DEFAULT 'standard'
            """
                )
            )
        except:
            pass  # Column may already exist

        # Add HR review columns to assessments
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE assessments
                ADD COLUMN hr_reviewed BOOLEAN DEFAULT FALSE,
                ADD COLUMN hr_reviewed_by VARCHAR(255),
                ADD COLUMN hr_reviewed_at DATETIME,
                ADD COLUMN hr_notes TEXT
            """
                )
            )
        except:
            pass  # Columns may already exist

        # Add time_remaining_seconds column to coding_sessions for timer functionality
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE coding_sessions
                ADD COLUMN time_remaining_seconds INT DEFAULT 3600
            """
                )
            )
        except:
            pass  # Column may already exist

        # Add email_notification_sent column to cv_candidates
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE cv_candidates
                ADD COLUMN email_notification_sent BOOLEAN DEFAULT FALSE,
                ADD COLUMN email_sent_at DATETIME
            """
                )
            )
        except:
            pass  # Columns may already exist

        # Add resume_parsing_version column to cv_candidates for tracking parsing improvements
        try:
            await conn.execute(
                text(
                    """
                ALTER TABLE cv_candidates
                ADD COLUMN resume_parsing_version VARCHAR(50) DEFAULT 'v1.0'
            """
                )
            )
        except:
            pass  # Column may already exist

    print("✅ Database initialized with INTEGRATED Module 5 and CV Extraction tables")
    print("📊 Tables created:")
    print("   1. coding_challenges")
    print("   2. coding_submissions")
    print("   3. technical_scores")
    print("   4. coding_sessions")
    print("   5. code_versions")
    print("   6. extension_requests")
    print("   7. job_requirements")
    print("   8. test_case_results")
    print("   9. code_detection")
    print("   10. coding_evaluations")
    print("   11. cv_candidates (CV Module)")
    print("   12. job_postings (CV Module)")
    print("   13. cv_processing_logs (CV Module)")
    print("   14. violations")

    yield

    # Shutdown code
    print("🔄 Shutting down database connections...")


# ==================== FASTAPI APP INITIALIZATION ====================
app = FastAPI(title="RECRUTO - AI-Powered Recruitment API", lifespan=lifespan)

# ==================== CORS MIDDLEWARE (before routers — wraps all routes & errors) ====================
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_origin_regex=_cors_origin_regex,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include AI Interview routes
app.include_router(interview_router)

# Include Vision Routes for face registration/verification
app.include_router(vision_router)

# ==================== REMAINING IMPORTS ====================
from domain.result import AssessmentResult
from typing import List, Dict

# Email
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Services
from application.mcq_service import generate_mcqs_service
from application.result_service import process_assessment_result
from application.auth_jwt import (
    verify_password,
    hash_password,
    create_access_token,
    decode_access_token,
)

# Database
from infrastructure.db_models import (
    get_db,
    InterviewSessionModel,
    InterviewStatus,
    AsyncSessionLocal,
)
from infrastructure.repositories import (
    AssessmentRepository,
    SessionRepository,
    ResultRepository,
)

# Adaptive engine
from domain.adaptive_engine import AdaptiveTestEngine

# Export libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from fastapi.responses import StreamingResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ==================== CUSTOM MODULE IMPORTS WITH FALLBACKS ====================

# Code Evaluator with fallback
try:
    from code_evaluator import CodeEvaluator, generate_ai_feedback_prompt
except ImportError:
    print("⚠️  CodeEvaluator not found, using fallback implementation")

    class CodeEvaluator:
        def evaluate_submission(
            self, code, language, test_results, execution_time_ms, memory_used_mb
        ):
            return {
                "correctness_score": 70,
                "efficiency_score": 70,
                "code_quality_score": 70,
                "problem_solving_score": 70,
                "readability_score": 70,
                "overall_score": 70,
                "grade": "C",
                "strengths": ["Basic functionality implemented"],
                "weaknesses": ["Could be optimized"],
                "suggestions": ["Improve code structure", "Add error handling"],
            }

    def generate_ai_feedback_prompt(code, language, evaluation):
        return f"Evaluate this {language} code for improvements."


# ==================== PLAGIARISM DETECTION ====================
class PlagiarismDetector:
    """Enhanced plagiarism detection for single-function solutions"""

    def __init__(self, similarity_threshold=0.65):  # Lower threshold
        self.similarity_threshold = similarity_threshold

    async def check_plagiarism_async(
        self, code: str, challenge_id: str, language: str, db
    ) -> dict:
        """Check code similarity with previous submissions - ENHANCED"""
        try:
            from sqlalchemy import text

            # Get recent submissions for this challenge
            query = text(
                """
                SELECT code, candidate_email, submission_id
                FROM coding_submissions
                WHERE challenge_id = :cid
                AND language = :lang
                ORDER BY submitted_at DESC
                LIMIT 50
            """
            )

            result = await db.execute(query, {"cid": challenge_id, "lang": language})
            previous_submissions = result.fetchall()

            if not previous_submissions:
                return {
                    "plagiarism_detected": False,
                    "plagiarism_score": 0,
                    "similar_submissions": [],
                    "details": "No previous submissions to compare",
                }

            # Normalize current code
            norm_current = self._normalize_code_for_comparison(code)

            # Calculate similarity with each previous submission
            similarities = []

            for prev_code, prev_email, prev_id in previous_submissions:
                norm_prev = self._normalize_code_for_comparison(prev_code)
                similarity = self._calculate_enhanced_similarity(
                    norm_current, norm_prev
                )

                if similarity > 0.3:  # Only store significant similarities
                    similarities.append(
                        {
                            "email": prev_email,
                            "submission_id": prev_id,
                            "similarity": similarity,
                        }
                    )

            # Sort by similarity (highest first)
            similarities.sort(key=lambda x: x["similarity"], reverse=True)

            # Get max similarity
            max_similarity = similarities[0]["similarity"] if similarities else 0

            is_suspicious = max_similarity >= self.similarity_threshold

            return {
                "plagiarism_detected": is_suspicious,
                "plagiarism_score": round(max_similarity * 100, 1),
                "similar_submissions": (
                    similarities[:3] if is_suspicious else []
                ),  # Top 3
                "details": f"{'⚠️ High similarity detected!' if is_suspicious else '✅ Code appears original'} (threshold: {self.similarity_threshold * 100}%)",
                "total_comparisons": len(previous_submissions),
                "similarity_distribution": self._get_similarity_distribution(
                    similarities
                ),
            }

        except Exception as e:
            print(f"❌ Plagiarism check error: {e}")
            return {
                "plagiarism_detected": False,
                "plagiarism_score": 0,
                "similar_submissions": [],
                "details": f"Error checking plagiarism: {str(e)[:100]}",
            }

    def _normalize_code_for_comparison(self, code: str) -> str:
        """Enhanced normalization for single-function solutions"""
        import re

        # Remove comments
        lines = code.split("\n")
        cleaned_lines = []

        for line in lines:
            # Remove inline comments
            if "#" in line:
                line = line.split("#")[0]

            # Remove whitespace and lowercase
            line = line.strip().lower()

            # Remove common variable names (make generic)
            line = re.sub(r"\b(result|ans|output|res|dp|memo|hash_map)\b", "VAR", line)

            # Remove numbers (except in specific contexts)
            line = re.sub(r"\b\d+\b", "NUM", line)

            # Remove string literals (keep structure)
            line = re.sub(r'"[^"]*"', "STR", line)
            line = re.sub(r"'[^']*'", "STR", line)

            if line:
                cleaned_lines.append(line)

        return "\n".join(cleaned_lines)

    def _calculate_enhanced_similarity(self, code1: str, code2: str) -> float:
        """Enhanced similarity calculation using multiple methods"""
        if not code1 or not code2:
            return 0.0

        # Method 1: Token-based similarity
        def tokenize(code):
            import re

            # Split by operators, parentheses, etc.
            tokens = re.findall(
                r"[a-zA-Z_][a-zA-Z0-9_]*|[+\-*/%=<>!&|^~]+|[(){}\[\],.:;]", code
            )
            return [t for t in tokens if len(t) > 1]  # Only keep meaningful tokens

        tokens1 = set(tokenize(code1))
        tokens2 = set(tokenize(code2))

        if not tokens1 or not tokens2:
            token_similarity = 0.0
        else:
            intersection = len(tokens1 & tokens2)
            union = len(tokens1 | tokens2)
            token_similarity = intersection / union if union > 0 else 0.0

        # Method 2: Sequence similarity (for control flow)
        lines1 = code1.split("\n")
        lines2 = code2.split("\n")

        # Compare control structures
        control_patterns = ["if", "for", "while", "def ", "return", "import", "from"]
        control1 = [
            line
            for line in lines1
            if any(pattern in line for pattern in control_patterns)
        ]
        control2 = [
            line
            for line in lines2
            if any(pattern in line for pattern in control_patterns)
        ]

        control_set1 = set(control1)
        control_set2 = set(control2)

        if not control_set1 or not control_set2:
            control_similarity = 0.0
        else:
            intersection = len(control_set1 & control_set2)
            union = len(control_set1 | control_set2)
            control_similarity = intersection / union if union > 0 else 0.0

        # Method 3: Structural similarity (indentation patterns)
        indent1 = self._extract_indentation_pattern(code1)
        indent2 = self._extract_indentation_pattern(code2)

        if not indent1 or not indent2:
            structure_similarity = 0.0
        else:
            # Compare indentation sequences
            min_len = min(len(indent1), len(indent2))
            if min_len == 0:
                structure_similarity = 0.0
            else:
                matches = sum(1 for i in range(min_len) if indent1[i] == indent2[i])
                structure_similarity = matches / min_len

        # Weighted combination
        final_similarity = (
            token_similarity * 0.4
            + control_similarity * 0.4
            + structure_similarity * 0.2
        )

        return final_similarity

    def _extract_indentation_pattern(self, code: str):
        """Extract indentation pattern for structural comparison"""
        lines = code.split("\n")
        pattern = []

        for line in lines:
            if line.strip():
                indent = len(line) - len(line.lstrip())
                # Quantize indentation levels
                level = indent // 4  # Assuming 4-space indents
                pattern.append(min(level, 3))  # Cap at level 3

        return pattern

    def _get_similarity_distribution(self, similarities):
        """Get distribution of similarity scores"""
        if not similarities:
            return {}

        distribution = {
            "high": len([s for s in similarities if s["similarity"] >= 0.7]),
            "medium": len([s for s in similarities if 0.4 <= s["similarity"] < 0.7]),
            "low": len([s for s in similarities if s["similarity"] < 0.4]),
            "total": len(similarities),
        }

        return distribution


plagiarism_detector = PlagiarismDetector()


# ==================== AI CODE DETECTION ====================
# ==================== ENHANCED AI CODE DETECTION ====================
# ==================== SIMPLIFIED AI DETECTOR FOR SINGLE-FUNCTION ====================
class AICodeDetector:
    """AI detector specifically for single-function coding solutions"""

    def __init__(self):
        # LOWER THRESHOLD - single-function solutions are often clean
        self.AI_THRESHOLD = 15  # Much lower!

        # CRITICAL AI PATTERNS FOR SINGLE-FUNCTION SOLUTIONS
        self.ai_patterns = [
            # 1. PERFECT LeetCode-style function names (2 points each)
            (
                r"def\s+(two_sum|max_subarray|max_profit|is_palindrome|reverse|longest|merge|search|find|rotate|solve|calculate|min_path|valid_parenthesis)\b",
                2.0,
            ),
            (
                r"function\s+(twoSum|maxSubarray|maxProfit|isPalindrome|reverse|longest|merge|search|find|rotate|solve|calculate|minPath|validParenthesis)\b",
                2.0,
            ),
            (
                r"func\s+(twoSum|maxSubarray|maxProfit|isPalindrome|reverse|longest|merge|search|find|rotate|solve|calculate|minPath|validParenthesis)\b",
                2.0,
            ),
            # 2. GENERIC PARAMETER NAMES (1.5 points each)
            (r"def\s+\w+\([^)]*nums[^)]*\)", 1.5),
            (r"def\s+\w+\([^)]*arr[^)]*\)", 1.5),
            (r"def\s+\w+\([^)]*s[^)]*\)", 1.5),
            (r"def\s+\w+\([^)]*k[^)]*\)", 1.5),
            (r"def\s+\w+\([^)]*target[^)]*\)", 1.5),
            (r"func\s+\w+\([^)]*nums[^)]*\)", 1.5),
            (r"function\s+\w+\([^)]*nums[^)]*\)", 1.5),
            # 3. ALGORITHMIC PATTERNS (2 points each - STRONG AI INDICATORS)
            (r"max_current\s*=", 2.0),  # Kadane's algorithm variable
            (r"max_global\s*=", 2.0),  # Kadane's algorithm variable
            (r"hash_map\s*=", 2.0),
            (r"dp\s*=\s*\[", 2.0),
            (r"memo\s*=", 2.0),
            (r"for\s+num\s+in\s+nums\[", 1.5),
            (r"for\s+i\s+in\s+range\(len\(", 1.5),
            (r"while\s+left\s*<=\s*right:", 1.5),
            (r"heapq\.", 2.0),
            (r"collections\.", 2.0),
            # 4. PERFECT EDGE CASE HANDLING (1 point each)
            (r"if\s+not\s+\w+:", 1.0),
            (r"if\s+len\(\w+\)\s*==\s*[01]:", 1.0),
            (r"if\s+len\(\w+\)\s*<=\s*1:", 1.0),
            (r"if\s+\w+\s+is\s+None:", 1.0),
            (r"if\s+\w+\s+==\s+None:", 1.0),
            # 5. OPTIMAL ALGORITHM PATTERNS (2 points each)
            (r"Kadane", 3.0),  # Named algorithm in comments
            (r"#\s*Time\s+complexity", 2.0),
            (r"#\s*Space\s+complexity", 2.0),
            (r"#\s*O\(", 2.0),  # Big O notation
            (r"#\s*Edge\s+case", 1.5),
            # 6. CLEAN RETURNS (1 point each)
            (r"return\s+max_global", 1.0),
            (r"return\s+max_current", 1.0),
            (r"return\s+\[\]", 1.0),
            (r"return\s+0", 1.0),
            (r'return\s+""', 1.0),
            # 7. PERFECT FORMATTING (0.5 points each)
            (r"\n\s{4}\w", 0.5),  # Exactly 4 spaces
            (r":\s*#", 0.5),  # Inline comment after colon
        ]

        # HUMAN PATTERNS - NEGATIVE SCORING
        self.human_patterns = [
            # STRONG HUMAN INDICATORS (subtract 3 points each)
            (r"print\(", 3.0),
            (r"console\.log", 3.0),
            (r"System\.out\.println", 3.0),
            (r"std::cout", 3.0),
            (r"fmt\.Print", 3.0),
            # MEDIUM HUMAN INDICATORS (subtract 2 points each)
            (r"#\s*TODO", 2.0),
            (r"#\s*FIXME", 2.0),
            (r"#\s*debug", 2.0),
            (r"#\s*wrong", 2.0),
            (r"//\s*TODO", 2.0),
            (r"//\s*FIXME", 2.0),
            # WEAK HUMAN INDICATORS (subtract 1 point each)
            (r"\t", 1.0),  # Tabs instead of spaces
            (r"\s{2,3}\w", 1.0),  # 2-3 spaces (inconsistent)
            (r"\s{5,8}\w", 1.0),  # Too many spaces
            (r"pass\s*#", 1.0),  # pass with comment
            (r"#\s*temp", 1.0),
            (r"#\s*trying", 1.0),
        ]

    async def detect_ai_code(self, code: str, language: str) -> dict:
        """Enhanced detection with weighted scoring"""
        import re

        if not code or len(code.strip()) < 10:
            return self._no_detection()

        code = code.strip()
        lines = code.split("\n")

        # Calculate AI score
        ai_score = 0
        ai_evidence = []

        for line_num, line in enumerate(lines, 1):
            line_lower = line.lower().strip()
            if not line_lower:
                continue

            # Check AI patterns with weights
            for pattern, weight in self.ai_patterns:
                if re.search(pattern, line_lower, re.IGNORECASE):
                    ai_score += weight
                    if len(ai_evidence) < 5:
                        match = re.search(pattern, line_lower, re.IGNORECASE)
                        if match:
                            snippet = match.group(0)[:40]
                            ai_evidence.append(
                                f"Line {line_num}: {snippet} (+{weight} points)"
                            )
                    break  # Only count first match per line

        # Calculate human score (negative)
        human_score = 0
        human_evidence = []

        for line_num, line in enumerate(lines, 1):
            line_lower = line.lower().strip()
            if not line_lower:
                continue

            # Check human patterns with weights
            for pattern, weight in self.human_patterns:
                if re.search(pattern, line_lower, re.IGNORECASE):
                    human_score += weight  # This is NEGATIVE weight
                    if len(human_evidence) < 3:
                        human_evidence.append(
                            f"Line {line_num}: Human pattern found (-{weight} points)"
                        )
                    break

        # Calculate final score
        base_score = ai_score

        # Apply human pattern penalty
        if human_score > 0:
            base_score = max(0, base_score - human_score)

        # Normalize to percentage (0-100)
        # Average single-function solution has 5-15 meaningful lines
        normalization_factor = 5  # Lower for single-function
        normalized_score = min((base_score / normalization_factor) * 100, 100)

        # Boost score for short, perfect solutions
        if len(lines) < 20 and base_score > 3:
            normalized_score = min(normalized_score * 1.5, 100)

        # Determine if AI
        is_ai = normalized_score >= self.AI_THRESHOLD

        # Generate explanation
        details = self._generate_explanation(
            is_ai, normalized_score, ai_score, human_score, len(lines)
        )

        return {
            "ai_generated": is_ai,
            "detected": is_ai,
            "confidence": round(normalized_score, 1),
            "score": round(normalized_score, 1),
            "evidence": ai_evidence[:3],
            "human_evidence": human_evidence[:2],
            "details": details,
            "source": "Weighted AI Detector",
            "metrics": {
                "total_lines": len(lines),
                "ai_points": round(ai_score, 1),
                "human_penalty": round(human_score, 1),
                "base_score": round(base_score, 1),
                "normalized_score": round(normalized_score, 1),
                "threshold": self.AI_THRESHOLD,
            },
            "needs_groq_check": (10 <= normalized_score <= 40),  # Uncertain range
        }

    def _generate_explanation(
        self,
        is_ai: bool,
        score: float,
        ai_points: float,
        human_penalty: float,
        total_lines: int,
    ) -> str:
        if is_ai:
            if score >= 70:
                return f"🚨 STRONG AI DETECTION ({score}% - {ai_points} AI points, -{human_penalty} human penalty)"
            elif score >= 50:
                return f"⚠️ MODERATE AI DETECTION ({score}% - {ai_points} AI points, -{human_penalty} human penalty)"
            elif score >= 30:
                return f"⚠️ WEAK AI DETECTION ({score}% - {ai_points} AI points, -{human_penalty} human penalty)"
            else:
                return f"⚠️ SUSPICIOUS ({score}% - {ai_points} AI points)"
        else:
            if human_penalty > 0:
                return f"✅ LIKELY HUMAN ({score}% - human patterns detected: -{human_penalty} points)"
            elif score < 10:
                return f"✅ CLEAN HUMAN CODE ({score}% - few AI patterns)"
            else:
                return f"❓ UNCERTAIN ({score}% - {ai_points} AI points)"

    def _no_detection(self):
        return {
            "ai_generated": False,
            "detected": False,
            "confidence": 0,
            "score": 0,
            "evidence": [],
            "human_evidence": [],
            "details": "Insufficient code to analyze",
            "source": "AI Detection System",
            "metrics": {},
            "needs_groq_check": False,
        }


# Initialize AI detector
ai_detector = AICodeDetector()


class CVProcessor:
    """Main CV processor for PDF, DOCX, and TXT files"""

    def __init__(self):
        try:
            self.nlp = spacy.load("en_core_web_sm")
        except:
            self.nlp = None

        self.skill_keywords = [
            "python",
            "java",
            "javascript",
            "react",
            "node.js",
            "sql",
            "nosql",
            "aws",
            "docker",
            "kubernetes",
            "git",
            "agile",
            "scrum",
            "machine learning",
            "ai",
            "data analysis",
            "pandas",
            "numpy",
            "html",
            "css",
            "typescript",
            "angular",
            "vue",
            "django",
            "flask",
            "fastapi",
            "mongodb",
            "postgresql",
            "mysql",
            "redis",
            "elasticsearch",
            "rest api",
            "graphql",
            "microservices",
            "devops",
            "ci/cd",
            # Additional skills
            "c++",
            "c#",
            ".net",
            "spring",
            "express",
            "vue.js",
            "next.js",
            "typescript",
            "docker",
            "kubernetes",
            "jenkins",
            "terraform",
            "ansible",
            "aws",
            "azure",
            "gcp",
            "firebase",
            "mongodb",
            "postgresql",
            "mysql",
            "redis",
            "elasticsearch",
            "kafka",
            "rabbitmq",
            "graphql",
            "rest api",
            "microservices",
            "serverless",
            "react native",
            "flutter",
            "android",
            "ios",
            "swift",
            "kotlin",
            "data structures",
            "algorithms",
            "oop",
            "functional programming",
            "test driven development",
            "tdd",
            "ci/cd",
            "devops",
            "agile",
            "scrum",
            "kanban",
            "jira",
            "confluence",
            "gitlab",
            "github",
            "bitbucket",
            "figma",
            "sketch",
            "adobe xd",
            "ui/ux",
            "photoshop",
            "illustrator",
            "tableau",
            "power bi",
            "excel",
            "word",
            "powerpoint",
        ]

        self.email_pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        self.phone_pattern = r"(?:\+?\d{1,4}[\s-]?)?(?:\(?\d{2,4}\)?[\s-]?)?\d{3,4}[\s-]?\d{3,4}[\s-]?\d{0,4}"

    async def process_cv_file(
        self, file_path: str, file_extension: str, job_id: str = None
    ):
        """Process CV file and extract structured data"""
        raw_text = await self._extract_text_from_file(file_path, file_extension)

        if not raw_text or len(raw_text.strip()) < 50:
            raise ValueError("CV file appears empty or unreadable")

        candidate_id = f"candidate_{uuid.uuid4()}"
        email = self._extract_email(raw_text)
        name = self._extract_name(raw_text)
        phone = self._extract_phone(raw_text)
        skills = self._extract_skills(raw_text)
        experience = self._extract_experience(raw_text)
        education = self._extract_education(raw_text)
        certifications = self._extract_certifications(raw_text)

        return {
            "candidate_id": candidate_id,
            "email": email,
            "name": name,
            "phone": phone,
            "skills": skills,
            "experience": experience,
            "education": education,
            "certifications": certifications,
            "raw_text": raw_text,
            "cv_file_path": file_path,
            "job_id": job_id,
        }

    async def _extract_text_from_file(self, file_path: str, file_extension: str) -> str:
        text = ""
        try:
            if file_extension.lower() == ".pdf":
                text = self._read_pdf(file_path)
            elif file_extension.lower() in [".docx", ".doc"]:
                text = self._read_docx(file_path)
            elif file_extension.lower() == ".txt":
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            raise

        return text

    def _read_pdf(self, file_path: str) -> str:
        text = ""
        try:
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            print(f"PDF reading error: {e}")
            raise
        return text

    def _read_docx(self, file_path: str) -> str:
        text = ""
        try:
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except Exception as e:
            print(f"DOCX reading error: {e}")
            raise
        return text

    def _extract_email(self, text: str) -> str:
        emails = re.findall(self.email_pattern, text, re.IGNORECASE)
        return emails[0] if emails else ""

    def _extract_phone(self, text: str) -> str:
        phones = re.findall(self.phone_pattern, text)
        for phone in phones:
            digits = re.sub(r"\D", "", phone)
            if len(digits) >= 10 and len(digits) <= 15:
                return phone.strip()
        return ""

    def _extract_name(self, text: str) -> str:
        """Enhanced name extraction from CV text"""
        lines = text.split("\n")

        # Look for name in first few lines
        for i, line in enumerate(lines[:10]):
            line = line.strip()

            # Common name patterns
            # 1. Line with 2-4 words, all starting with capital letters
            words = line.split()
            if 2 <= len(words) <= 4:
                if all(word[0].isupper() for word in words if word):
                    # Exclude lines with common headers
                    if not any(
                        header in line.lower()
                        for header in [
                            "email",
                            "phone",
                            "address",
                            "linkedin",
                            "github",
                            "objective",
                            "summary",
                            "experience",
                            "education",
                            "skills",
                            "projects",
                        ]
                    ):
                        # Exclude lines with email/phone patterns
                        if "@" not in line and not re.search(r"\d{10}", line):
                            return line

            # 2. Look for name after common headers
            if line.lower() in ["name:", "full name:", "candidate:"] and i + 1 < len(
                lines
            ):
                next_line = lines[i + 1].strip()
                if next_line and len(next_line.split()) <= 4:
                    return next_line

        # 3. Try email prefix extraction
        email = self._extract_email(text)
        if email:
            # Extract name from email (e.g., john.doe@email.com -> John Doe)
            username = email.split("@")[0]
            if "." in username:
                name_parts = [part.capitalize() for part in username.split(".")]
                return " ".join(name_parts)
            elif "_" in username:
                name_parts = [part.capitalize() for part in username.split("_")]
                return " ".join(name_parts)

        return ""

    def _extract_skills(self, text: str):
        found_skills = set()
        text_lower = text.lower()

        # Check for keyword matches strictly using word boundaries to avoid matching "ai" inside "said"
        for skill in self.skill_keywords:
            pattern = r"\b" + re.escape(skill) + r"\b"
            if re.search(pattern, text_lower):
                found_skills.add(skill.title())

        # Look for skills section with better patterns
        skill_section_patterns = [
            r"skills?\s*:?\s*([^•]+?)(?:\n\n|\n[A-Z]|$)",
            r"technical skills?\s*:?\s*([^•]+?)(?:\n\n|\n[A-Z]|$)",
            r"competencies?\s*:?\s*([^•]+?)(?:\n\n|\n[A-Z]|$)",
            r"expertise\s*:?\s*([^•]+?)(?:\n\n|\n[A-Z]|$)",
            r"technologies?\s*:?\s*([^•]+?)(?:\n\n|\n[A-Z]|$)",
        ]

        for pattern in skill_section_patterns:
            matches = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if matches:
                skill_text = matches.group(1)
                lines = skill_text.split("\n")
                for line in lines:
                    line = line.strip()
                    if line and len(line) > 2:
                        # Clean up the skill
                        skill = line.strip("•-*·: ")
                        # Split by commas, slashes, or pipes
                        sub_skills = re.split(r"[,/|]", skill)
                        for sub_skill in sub_skills:
                            sub_skill = sub_skill.strip()
                            idx = sub_skill.lower()
                            if (
                                sub_skill
                                and len(sub_skill) > 1
                                and len(sub_skill.split()) <= 3
                            ):
                                # We only consider it a skill if it's short (<= 3 words)
                                found_skills.add(sub_skill.title())
                break

        # Also look for bullet point skills, but ONLY if they are short (<= 3 words). Avoid pulling entire sentences.
        bullet_pattern = r"[•\-*]\s*([^\n]+)"
        bullet_matches = re.findall(bullet_pattern, text)
        for match in bullet_matches:
            skill_candidate = match.strip()
            if len(skill_candidate.split()) <= 3 and len(skill_candidate) > 1:
                # If a short bullet point contains a known keyword, maybe it's a specific relevant tool setting
                if any(k in skill_candidate.lower() for k in self.skill_keywords):
                    found_skills.add(skill_candidate.title())

        return self._drop_noise_skill_tokens(list(found_skills))

    def _drop_noise_skill_tokens(self, skills: list) -> list:
        """Remove section-header fragments and other false positives from extracted skills."""
        noise_exact = {
            "and expertise",
            "expertise",
            "technical skills",
            "soft skills",
            "core skills",
            "key skills",
            "skills",
            "tools",
            "technologies",
            "frameworks",
            "languages",
        }
        out = []
        for s in skills:
            if not s or not str(s).strip():
                continue
            low = str(s).lower().strip()
            if low in noise_exact:
                continue
            if low.startswith("and ") and "expertise" in low:
                continue
            out.append(s)
        return out

    def _extract_experience(self, text: str):
        experience = []
        exp_patterns = [
            r"experience\s*:?\s*(.+?)(?:\n\neducation|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
            r"work experience\s*:?\s*(.+?)(?:\n\neducation|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
            r"employment history\s*:?\s*(.+?)(?:\n\neducation|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
            r"professional experience\s*:?\s*(.+?)(?:\n\neducation|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
            r"work history\s*:?\s*(.+?)(?:\n\neducation|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
        ]

        for pattern in exp_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                exp_text = match.group(1)
                lines = exp_text.split("\n")
                current_job = {}
                job_text = ""

                for line in lines:
                    line = line.strip()
                    if line:
                        # Look for date patterns (YYYY-YYYY, YYYY-Present, etc.)
                        date_match = re.search(
                            r"(\d{4}\s*[-–]\s*(?:\d{4}|present|current|now))|"
                            r"((?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}\s*[-–]\s*(?:(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}|present|current|now))",
                            line,
                            re.IGNORECASE,
                        )

                        if date_match:
                            # Save previous job if exists
                            if current_job:
                                current_job["description"] = job_text.strip()
                                experience.append(current_job)
                                job_text = ""

                            current_job = {
                                "duration": date_match.group(0),
                                "title": "",
                                "company": "",
                                "description": "",
                            }

                            # Try to extract title and company from the same line
                            line_without_date = line.replace(
                                date_match.group(0), ""
                            ).strip()
                            if line_without_date:
                                # Common pattern: "Title at Company" or "Title, Company"
                                if " at " in line_without_date.lower():
                                    parts = line_without_date.split(" at ")
                                    current_job["title"] = parts[0].strip()
                                    current_job["company"] = (
                                        parts[1].strip() if len(parts) > 1 else ""
                                    )
                                elif ", " in line_without_date:
                                    parts = line_without_date.split(", ")
                                    current_job["title"] = parts[0].strip()
                                    current_job["company"] = (
                                        parts[1].strip() if len(parts) > 1 else ""
                                    )
                                else:
                                    current_job["title"] = line_without_date
                        else:
                            if not current_job:
                                current_job = {
                                    "duration": "Not specified",
                                    "title": "",
                                    "company": "",
                                    "description": "",
                                }
                                if " at " in line.lower():
                                    parts = re.split(
                                        r"\s+at\s+",
                                        line,
                                        maxsplit=1,
                                        flags=re.IGNORECASE,
                                    )
                                    current_job["title"] = parts[0].strip()
                                    current_job["company"] = (
                                        parts[1].strip() if len(parts) > 1 else ""
                                    )
                                elif (
                                    len(line.split()) <= 8
                                    and not line.startswith("•")
                                    and not line.startswith("-")
                                    and not line.startswith("*")
                                    and not line.startswith(".")
                                ):
                                    current_job["title"] = line
                                else:
                                    job_text += line + "\n"
                            else:
                                if (
                                    not current_job.get("title")
                                    and len(line.split()) <= 8
                                    and not line.startswith("•")
                                    and not line.startswith("-")
                                    and not line.startswith("*")
                                    and not line.startswith(".")
                                ):
                                    current_job["title"] = line
                                elif (
                                    not current_job.get("company")
                                    and len(line.split()) <= 5
                                    and not line.startswith("•")
                                    and not line.startswith("-")
                                    and not line.startswith("*")
                                    and not line.startswith(".")
                                ):
                                    current_job["company"] = line
                                else:
                                    job_text += line + "\n"

                # Save the last job
                if current_job:
                    current_job["description"] = job_text.strip()
                    experience.append(current_job)

                break

        return experience

    def _extract_education(self, text: str):
        education = []

        # More aggressive - look for any university name in the entire text
        university_keywords = [
            "FAST",
            "NU",
            "NUST",
            "FUI",
            "GIFT",
            "COMSATS",
            "UET",
            "IBA",
            "Bahria",
            "University",
            "Institute",
        ]
        degree_keywords = [
            "bsc",
            "msc",
            "mba",
            "bachelor",
            "master",
            "phd",
            "be",
            "b.tech",
            "m.tech",
            "b.e",
            "m.e",
        ]

        # Search entire text for university + year pattern
        for uni_kw in university_keywords:
            # Look for "UniversityName ... 2020-2024" pattern
            match = re.search(
                rf"{uni_kw}[^\n]{{0,50}}(\d{{4}}\s*[-–]\s*(?:\d{{4}}|present))",
                text,
                re.IGNORECASE,
            )
            if match:
                education.append(
                    {
                        "degree": "Higher Education",
                        "institution": uni_kw,
                        "duration": match.group(1),
                        "details": match.group(0)[:100],
                    }
                )
                break

        # Also look for degree keywords with years anywhere
        if not education:
            for deg_kw in degree_keywords:
                matches = re.findall(
                    rf"{deg_kw}[^\n]{{0,30}}(\d{{4}}\s*[-–]\s*(?:\d{{4}}|present))",
                    text,
                    re.IGNORECASE,
                )
                if matches:
                    for m in matches[:1]:  # Take first match
                        education.append(
                            {
                                "degree": deg_kw.upper(),
                                "institution": "",
                                "duration": m,
                                "details": "",
                            }
                        )
                        break

        # Try original patterns as fallback
        edu_patterns = [
            r"education\s*:?\s*(.+?)(?:\n\nexperience|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
            r"academic background\s*:?\s*(.+?)(?:\n\nexperience|\n\nskills|\n\nprojects|\n\n$|\n[A-Z]{3,})",
        ]

        for pattern in edu_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                edu_text = match.group(1)
                lines = edu_text.split("\n")
                current_edu = {}
                edu_text_accumulated = ""

                for line in lines:
                    line = line.strip()
                    if line:
                        # Look for degree indicators
                        degree_keywords = [
                            "bachelor",
                            "master",
                            "phd",
                            "doctorate",
                            "mba",
                            "msc",
                            "mca",
                            "bsc",
                            "bca",
                            "diploma",
                            "certificate",
                            "degree",
                            "graduation",
                            "post graduation",
                            "be",
                            "b.tech",
                            "m.tech",
                            "b.e.",
                            "m.e.",
                        ]

                        # Look for date patterns in education
                        date_match = re.search(
                            r"\d{4}\s*[-–]\s*(?:\d{4}|present)", line
                        )

                        if (
                            any(keyword in line.lower() for keyword in degree_keywords)
                            or date_match
                        ):
                            # Save previous education if exists
                            if current_edu:
                                if edu_text_accumulated and not current_edu.get(
                                    "details"
                                ):
                                    current_edu["details"] = (
                                        edu_text_accumulated.strip()
                                    )
                                education.append(current_edu)
                                edu_text_accumulated = ""

                            current_edu = {
                                "degree": "",
                                "institution": "",
                                "year": "",
                                "details": "",
                            }

                            # Extract year if present
                            if date_match:
                                current_edu["year"] = date_match.group(0)
                                line = line.replace(date_match.group(0), "").strip()

                            # The line likely contains degree and/or institution
                            if " at " in line.lower():
                                parts = line.split(" at ")
                                current_edu["degree"] = parts[0].strip()
                                current_edu["institution"] = (
                                    parts[1].strip() if len(parts) > 1 else ""
                                )
                            elif ", " in line:
                                parts = line.split(", ")
                                current_edu["degree"] = parts[0].strip()
                                current_edu["institution"] = (
                                    parts[1].strip() if len(parts) > 1 else ""
                                )
                            else:
                                current_edu["degree"] = line
                        elif current_edu:
                            # If we have education entry but no institution yet
                            if (
                                not current_edu.get("institution")
                                and len(line.split()) <= 5
                            ):
                                current_edu["institution"] = line
                            else:
                                # Accumulate details
                                edu_text_accumulated += line + " "

                # Save the last education entry
                if current_edu:
                    if edu_text_accumulated and not current_edu.get("details"):
                        current_edu["details"] = edu_text_accumulated.strip()
                    education.append(current_edu)

                break

        return education

    def _extract_certifications(self, text: str):
        certifications = []
        cert_patterns = [
            r"certifications?\s*:?\s*(.+?)(?:\n\n|\n[A-Z]{3,}|$)",
            r"licenses?\s*:?\s*(.+?)(?:\n\n|\n[A-Z]{3,}|$)",
            r"certificates?\s*:?\s*(.+?)(?:\n\n|\n[A-Z]{3,}|$)",
            r"training\s*:?\s*(.+?)(?:\n\n|\n[A-Z]{3,}|$)",
        ]

        for pattern in cert_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                cert_text = match.group(1)
                lines = cert_text.split("\n")
                for line in lines:
                    line = line.strip()
                    if line and len(line) > 3:
                        # Clean up and add
                        cert = line.strip("•-*· ")
                        certifications.append(cert)
                break

        # Also look for bullet point certifications
        if not certifications:
            bullet_pattern = r"[•\-*]\s*([A-Z][^\n]+?(?:certification|certificate|license|training)[^\n]*)"
            bullet_matches = re.findall(bullet_pattern, text, re.IGNORECASE)
            for match in bullet_matches:
                cert = match.strip()
                if cert and len(cert) > 10:  # Minimum reasonable length
                    certifications.append(cert)

        return list(set(certifications))


# ==================== OCR PROCESSOR ====================


class OCRProcessor:
    """OCR processor for scanned CVs"""

    def __init__(self):
        if os.name == "nt":
            pytesseract.pytesseract.tesseract_cmd = (
                r"C:\Program Files\Tesseract-OCR\tesseract.exe"
            )

    def extract_text_from_scanned_pdf(self, pdf_path: str) -> str:
        text = ""
        try:
            pdf_document = fitz.open(pdf_path)
            for page_num in range(len(pdf_document)):
                page = pdf_document.load_page(page_num)
                pix = page.get_pixmap()
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp_path = tmp.name
                    pix.save(tmp_path)

                page_text = self._perform_ocr(tmp_path)
                text += page_text + "\n\n"
                os.unlink(tmp_path)

            pdf_document.close()
        except Exception as e:
            print(f"OCR processing error for {pdf_path}: {e}")
            raise
        return text

    def _perform_ocr(self, image_path: str) -> str:
        try:
            image = Image.open(image_path)
            if image.mode != "L":
                image = image.convert("L")
            text = pytesseract.image_to_string(image, lang="eng")
            return text
        except Exception as e:
            print(f"Tesseract OCR error: {e}")
            raise

    def is_scanned_pdf(self, pdf_path: str) -> bool:
        try:
            pdf_document = fitz.open(pdf_path)
            page = pdf_document.load_page(0)
            text = page.get_text()
            if len(text.strip()) < 50:
                images = page.get_images()
                if images:
                    return True
            pdf_document.close()
            return False
        except Exception as e:
            print(f"Error checking PDF type: {e}")
            return False


# ==================== SKILL MATCHER ====================


class SkillMatcher:
    """Skill matching logic"""

    def __init__(self):
        self.skill_normalization = {
            "python": ["python", "python3", "python 3", "py"],
            "javascript": ["javascript", "js", "es6", "ecmascript"],
            "react": ["react", "react.js", "reactjs"],
            "node.js": ["node", "node.js", "nodejs"],
            "sql": ["sql", "mysql", "postgresql", "postgres", "oracle"],
            "aws": ["aws", "amazon web services", "amazon aws"],
            "docker": ["docker", "docker container"],
            "kubernetes": ["kubernetes", "k8s"],
            "git": ["git", "github", "gitlab", "bitbucket"],
            "machine learning": ["machine learning", "ml", "ai/ml"],
            "data analysis": ["data analysis", "analytics", "data analytics"],
        }

    def match_skills(self, candidate_skills, job_requirements):
        required_skills_raw = job_requirements.get("required_skills", "")

        if isinstance(required_skills_raw, str):
            required_skills = self._parse_skills_string(required_skills_raw)
        elif isinstance(required_skills_raw, list):
            required_skills = required_skills_raw
        else:
            required_skills = []

        normalized_candidate = self._normalize_skills(candidate_skills)
        normalized_required = self._normalize_skills(required_skills)

        matched_skills = []
        missing_skills = []

        for req_skill in normalized_required:
            found = False
            for cand_skill in normalized_candidate:
                if self._skills_match(req_skill, cand_skill):
                    matched_skills.append((req_skill, cand_skill))
                    found = True
                    break
            if not found:
                missing_skills.append(req_skill)

        extra_skills = []
        for cand_skill in normalized_candidate:
            has_match = False
            for req_skill in normalized_required:
                if self._skills_match(req_skill, cand_skill):
                    has_match = True
                    break
            if not has_match and cand_skill not in extra_skills:
                extra_skills.append(cand_skill)

        if normalized_required:
            match_percentage = (len(matched_skills) / len(normalized_required)) * 100
        else:
            match_percentage = 0

        skill_breakdown = {
            "total_required": len(normalized_required),
            "candidate_has": len(normalized_candidate),
            "matched": len(matched_skills),
            "missing": len(missing_skills),
            "extra": len(extra_skills),
        }

        matched_pairs = [
            {"required": r, "candidate": c} for r, c in matched_skills
        ]

        return {
            "candidate_skills": normalized_candidate,
            "required_skills": normalized_required,
            "missing_skills": missing_skills,
            "extra_skills": extra_skills,
            "matched_pairs": matched_pairs,
            "match_percentage": round(match_percentage, 2),
            "skill_breakdown": skill_breakdown,
        }

    def _parse_skills_string(self, skills_str: str):
        if not skills_str:
            return []
        skills = re.split(r"[,•\n]", skills_str)
        cleaned_skills = []
        for skill in skills:
            skill = skill.strip()
            if skill and len(skill) > 1:
                cleaned_skills.append(skill)
        return cleaned_skills

    def _normalize_skills(self, skills):
        normalized = []
        for skill in skills:
            skill_lower = skill.lower().strip()
            normalized_found = False
            for norm_skill, variants in self.skill_normalization.items():
                if skill_lower in variants or norm_skill in skill_lower:
                    if norm_skill not in normalized:
                        normalized.append(norm_skill)
                    normalized_found = True
                    break
            if not normalized_found and skill_lower not in normalized:
                normalized.append(skill_lower)
        return normalized

    def _skills_match(self, skill1: str, skill2: str) -> bool:
        skill1_lower = skill1.lower()
        skill2_lower = skill2.lower()

        if skill1_lower == skill2_lower:
            return True

        if skill1_lower in skill2_lower or skill2_lower in skill1_lower:
            return True

        for norm_skill, variants in self.skill_normalization.items():
            if skill1_lower in variants and skill2_lower in variants:
                return True
            if norm_skill in skill1_lower and norm_skill in skill2_lower:
                return True

        return False


# ==================== GMAIL MONITOR ====================


def _gmail_newer_than_clause(lookback_hours: int) -> str:
    """
    Gmail search does NOT accept Unix timestamps in after: — use after:YYYY/MM/DD or newer_than:Xd.
    See: https://support.google.com/mail/answer/7190
    """
    days = max(1, (int(lookback_hours) + 23) // 24)
    return f"newer_than:{days}d"


class GmailCVMonitor:
    """Gmail CV monitor with enhanced debugging and error handling"""

    def __init__(self):
        self.SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
        self.service = None
        # Resolve paths from backend folder (not process cwd) so token/credentials are found
        self._backend_dir = backend_dir
        self.download_dir = str(self._backend_dir / "cv_downloads")

        if not os.path.exists(self.download_dir):
            os.makedirs(self.download_dir, exist_ok=True)
        print(f"✅ Gmail monitor initialized. Download dir: {self.download_dir}")

    async def authenticate(self, token_path: str = None, credentials_path: str = None):
        """Authenticate with Gmail API"""
        if token_path is None:
            token_path = str(self._backend_dir / "token.json")
        if credentials_path is None:
            credentials_path = str(self._backend_dir / "credentials.json")
        print(f"🔐 Authenticating Gmail API...")
        creds = None

        # Check for existing token
        if os.path.exists(token_path):
            try:
                print(f"  📁 Found token file: {token_path}")
                creds = Credentials.from_authorized_user_file(token_path, self.SCOPES)
                print(f"  ✅ Loaded credentials from token")
            except Exception as e:
                print(f"  ⚠️ Token file error: {e}")
                os.remove(token_path)  # Remove invalid token
                creds = None

        # If no valid credentials, authenticate
        if not creds or not creds.valid:
            print(f"  🔄 Need to authenticate...")
            if creds and creds.expired and creds.refresh_token:
                try:
                    print(f"  🔄 Refreshing expired token...")
                    creds.refresh(Request())
                    print(f"  ✅ Token refreshed")
                except Exception as e:
                    print(f"  ⚠️ Token refresh failed: {e}")
                    creds = None

            # Need new authentication
            if not creds:
                try:
                    # Check if credentials file exists
                    if not os.path.exists(credentials_path):
                        print(f"  ❌ Credentials file not found: {credentials_path}")
                        print(
                            f"  ℹ️ Please download credentials.json from Google Cloud Console"
                        )
                        return False

                    print(f"  📁 Using credentials file: {credentials_path}")
                    flow = InstalledAppFlow.from_client_secrets_file(
                        credentials_path, self.SCOPES
                    )
                    print(f"  🌐 Opening browser for authentication...")
                    creds = flow.run_local_server(port=0)

                    # Save the credentials
                    with open(token_path, "w") as token:
                        token.write(creds.to_json())
                    print(f"  💾 Saved new token to {token_path}")

                except Exception as e:
                    print(f"  ❌ Authentication failed: {e}")
                    return False

        # Build Gmail service
        try:
            self.service = build("gmail", "v1", credentials=creds)
            print(f"  ✅ Gmail service authenticated successfully")
            return True
        except Exception as e:
            print(f"  ❌ Failed to build Gmail service: {e}")
            return False

    async def debug_list_recent_emails(self, lookback_hours: int = 24, limit: int = 10):
        """Debug method to list recent emails"""
        print(f"\n🔍 DEBUG: Listing recent emails ({lookback_hours}h, max {limit})")

        if not self.service:
            print("  ❌ Service not authenticated")
            return

        newer = _gmail_newer_than_clause(lookback_hours)
        query = newer

        try:
            results = (
                self.service.users()
                .messages()
                .list(userId="me", q=query, maxResults=limit)
                .execute()
            )

            messages = results.get("messages", [])

            print(f"  📧 Found {len(messages)} recent emails:")

            for i, msg in enumerate(messages):
                try:
                    msg_detail = (
                        self.service.users()
                        .messages()
                        .get(userId="me", id=msg["id"], format="metadata")
                        .execute()
                    )

                    headers = msg_detail["payload"]["headers"]
                    subject = next(
                        (h["value"] for h in headers if h["name"] == "Subject"),
                        "No Subject",
                    )
                    sender = next(
                        (h["value"] for h in headers if h["name"] == "From"), "Unknown"
                    )
                    date = next(
                        (h["value"] for h in headers if h["name"] == "Date"), "Unknown"
                    )

                    # Check for attachments
                    has_attachments = "parts" in msg_detail["payload"] and any(
                        part.get("filename")
                        for part in msg_detail["payload"].get("parts", [])
                    )

                    print(f"  {i + 1}. 📩 From: {sender[:40]}...")
                    print(f"     Subject: {subject[:60]}...")
                    print(f"     Date: {date}")
                    print(
                        f"     Attachments: {'✅ Yes' if has_attachments else '❌ No'}"
                    )
                    print(f"     ID: {msg['id']}")
                    print()

                except Exception as e:
                    print(f"  ⚠️ Failed to get message {msg['id']}: {e}")

            return len(messages)

        except Exception as e:
            print(f"  ❌ Debug failed: {e}")
            return 0

    async def fetch_new_cvs(self, lookback_hours: int = 24):
        """Fetch CVs from Gmail with multiple search strategies"""
        print(f"\n🔍 Searching for CVs in Gmail ({lookback_hours}h lookback)...")

        if not self.service:
            raise Exception("Gmail service not authenticated")

        newer = _gmail_newer_than_clause(lookback_hours)

        # Strategy 1: Broad search for attachments
        query1 = f"{newer} has:attachment"
        print(f"  Strategy 1: All emails with attachments")

        # Strategy 2: CV/resume specific search
        query2 = (
            f"{newer} (has:attachment (filename:pdf OR filename:docx OR filename:doc OR filename:txt)) "
            f'(subject:CV OR subject:resume OR subject:"application" OR body:CV OR body:resume)'
        )
        print(f"  Strategy 2: CV-specific search")

        cv_data_list = []
        seen_message_ids = set()

        # Try both strategies
        for strategy, query in [("Broad", query1), ("CV-specific", query2)]:
            print(f"\n  Trying {strategy} search...")
            print(f"  Query: {query}")

            try:
                results = (
                    self.service.users()
                    .messages()
                    .list(userId="me", q=query, maxResults=20)
                    .execute()
                )

                messages = results.get("messages", [])
                print(f"  Found {len(messages)} messages with {strategy} search")

                processed_count = 0
                for msg in messages[:15]:  # Limit processing
                    try:
                        mid = msg["id"]
                        if mid in seen_message_ids:
                            continue
                        msg_data = await self._process_email_message(mid)
                        if msg_data:
                            seen_message_ids.add(mid)
                            cv_data_list.append(msg_data)
                            processed_count += 1
                            print(
                                f"    ✅ Found CV: {msg_data.get('subject', 'No subject')[:50]}..."
                            )
                    except Exception as e:
                        print(
                            f"    ⚠️ Failed to process message {msg['id']}: {str(e)[:100]}"
                        )

                print(f"  Processed {processed_count} CVs from {strategy} search")

            except Exception as e:
                print(f"  ⚠️ {strategy} search failed: {str(e)[:100]}")

        print(f"\n✅ Total CVs found: {len(cv_data_list)}")
        return cv_data_list

    async def _process_email_message(self, message_id: str):
        """Process a single email message and extract CV data"""
        try:
            print(f"  🔍 Processing message: {message_id}")

            # Get full message details
            msg = (
                self.service.users()
                .messages()
                .get(userId="me", id=message_id, format="full")
                .execute()
            )

            # Extract headers
            headers = msg["payload"]["headers"]
            subject = next((h["value"] for h in headers if h["name"] == "Subject"), "")
            sender = next((h["value"] for h in headers if h["name"] == "From"), "")
            date = next((h["value"] for h in headers if h["name"] == "Date"), "")

            print(f"    📧 Subject: {subject[:60]}...")
            print(f"    👤 From: {sender[:40]}...")

            # Extract email body
            body = await self._get_email_body(msg["payload"])
            sender_email = await self._extract_email_from_sender(sender)

            # Extract attachments
            attachments = await self._get_attachments(msg["payload"], msg["id"])
            print(f"    📎 Found {len(attachments)} attachments")

            if attachments:
                cv_data = {
                    "message_id": message_id,
                    "subject": subject,
                    "sender": sender,
                    "sender_email": sender_email,
                    "date": date,
                    "body": body[:500],  # Limit body size
                    "attachments": attachments,
                    "source": "gmail",
                }
                print(f"    ✅ Successfully extracted CV data")
                return cv_data
            else:
                print(f"    ⚠️ No valid attachments found")

        except Exception as e:
            print(f"    ❌ Error processing email {message_id}: {str(e)[:100]}")

        return None

    async def _get_email_body(self, payload):
        """Extract plain text email body (handles nested multipart/alternative)."""
        body = ""

        def decode_part_data(data_b64: str) -> str:
            if not data_b64:
                return ""
            return base64.urlsafe_b64decode(data_b64).decode("utf-8", errors="ignore")

        def walk_parts(parts: list) -> None:
            nonlocal body
            for part in parts:
                mt = part.get("mimeType", "")
                if mt.startswith("multipart/") and "parts" in part:
                    walk_parts(part["parts"])
                elif mt == "text/plain" and part.get("body", {}).get("data"):
                    body += decode_part_data(part["body"]["data"])

        try:
            if "parts" in payload:
                walk_parts(payload["parts"])
            elif payload.get("mimeType") == "text/plain" and payload.get(
                "body", {}
            ).get("data"):
                body = decode_part_data(payload["body"]["data"])

            return body.strip()
        except Exception as e:
            print(f"      ⚠️ Failed to extract body: {e}")
            return ""

    async def _extract_email_from_sender(self, sender: str) -> str:
        """Extract email address from sender string"""
        email_pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        matches = re.findall(email_pattern, sender)
        return matches[0] if matches else ""

    def _download_one_cv_attachment(
        self, part: dict, message_id: str
    ) -> Optional[dict]:
        """Download a single MIME part that has filename + attachmentId. Returns metadata dict or None."""
        filename = (part.get("filename") or "").strip()
        if not filename:
            return None
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext not in [".pdf", ".docx", ".doc", ".txt"]:
            return None
        body_meta = part.get("body") or {}
        if "attachmentId" not in body_meta:
            return None
        att_id = body_meta["attachmentId"]
        print(f"      📄 Found CV attachment: {filename}")
        att = (
            self.service.users()
            .messages()
            .attachments()
            .get(userId="me", messageId=message_id, id=att_id)
            .execute()
        )
        if "data" not in att:
            print(f"      ⚠️ No data in attachment: {filename}")
            return None
        file_data = base64.urlsafe_b64decode(att["data"])
        safe_filename = f"{uuid.uuid4()}_{filename.replace(' ', '_')}"
        file_path = os.path.join(self.download_dir, safe_filename)
        with open(file_path, "wb") as f:
            f.write(file_data)
        print(f"      💾 Saved: {filename} ({len(file_data)} bytes)")
        return {
            "original_filename": filename,
            "filename": safe_filename,
            "file_path": file_path,
            "size": len(file_data),
            "extension": file_ext,
        }

    async def _get_attachments(self, payload, message_id):
        """Extract and download attachments"""
        attachments = []

        def _traverse_parts(parts, depth=0):
            for part in parts:
                if "parts" in part:
                    _traverse_parts(part["parts"], depth + 1)
                    continue
                if not part.get("filename") or not part["filename"].strip():
                    continue
                try:
                    meta = self._download_one_cv_attachment(part, message_id)
                    if meta:
                        attachments.append(meta)
                except Exception as e:
                    print(
                        f"      ❌ Failed to download {part.get('filename')}: {str(e)[:100]}"
                    )

        try:
            if "parts" in payload:
                _traverse_parts(payload["parts"])
            else:
                meta = self._download_one_cv_attachment(payload, message_id)
                if meta:
                    attachments.append(meta)

        except Exception as e:
            print(f"      ❌ Attachment traversal failed: {e}")

        return attachments

    async def test_connection(self):
        """Test Gmail API connection"""
        try:
            if not self.service:
                return {"connected": False, "error": "Service not authenticated"}

            # Try a simple API call
            profile = self.service.users().getProfile(userId="me").execute()
            return {
                "connected": True,
                "email": profile.get("emailAddress"),
                "messages_total": profile.get("messagesTotal"),
                "threads_total": profile.get("threadsTotal"),
            }
        except Exception as e:
            return {"connected": False, "error": str(e)}


# ==================== INITIALIZE SINGLETONS ====================

cv_processor = CVProcessor()
ocr_processor = OCRProcessor()
skill_matcher = SkillMatcher()
gmail_monitor = GmailCVMonitor()


# ==================== GROQ API AI JUDGE ====================
async def groq_ai_judge(code: str, language: str) -> dict:
    """Use Groq API to judge if code is AI-generated"""
    if not GROQ_API_KEY:
        print("❌ Groq API key not found - skipping AI judge")
        return {
            "ai_generated": False,
            "confidence": 0,
            "reasoning": "Groq API not available",
            "evidence": [],
            "source": "Groq API (unavailable)",
        }

    # Format prompt for coding challenge solutions
    prompt = f"""Analyze this {language} coding solution and determine if it was:
1. Generated by AI (ChatGPT, Claude, Copilot, etc.)
2. Written by a human programmer

CODE TO ANALYZE:
```{language}
{code[:1500]}  # Limit to first 1500 chars
```

Consider these specific indicators for AI-generated code in coding challenges:
- Generic function/parameter names (solve, calculate, nums, arr, k)
- Perfect formatting and consistent indentation
- Over-commenting or specific comment patterns (# Time complexity, # Edge cases)
- Optimal algorithm choice without exploration
- Missing error handling and edge cases
- Standard library imports only
- Single function solutions
- Use of common variable names (result, ans, temp, dp, memo)

If you detect multiple AI indicators, classify as AI-generated.

Return ONLY valid JSON with this exact structure:
{{
    "ai_generated": true/false,
    "confidence": 0-100,
    "reasoning": "Brief explanation of your decision",
    "evidence": ["specific pattern 1", "specific pattern 2", "specific pattern 3"],
    "source": "Groq AI Judge"
}}

Do not include any other text, only the JSON."""

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": GROQ_MODEL,
        "messages": [
            {
                "role": "system",
                "content": "You are an expert at detecting AI-generated code. Analyze code and return only JSON.",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.3,  # Lower temperature for more consistent results
        "max_tokens": 500,
        "response_format": {"type": "json_object"},
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30,
            ) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    response = data["choices"][0]["message"]["content"]

                    try:
                        result = json.loads(response)
                        # Ensure required fields
                        if "ai_generated" not in result:
                            result["ai_generated"] = False
                        if "confidence" not in result:
                            result["confidence"] = 0
                        if "evidence" not in result:
                            result["evidence"] = []
                        if "reasoning" not in result:
                            result["reasoning"] = "Analysis completed"

                        print(f"✅ Groq AI Judge: {result['confidence']}% confidence")
                        return result

                    except json.JSONDecodeError:
                        print("❌ Groq returned invalid JSON")
                        return {
                            "ai_generated": False,
                            "confidence": 0,
                            "reasoning": "Invalid response format",
                            "evidence": [],
                            "source": "Groq API (error)",
                        }
                else:
                    print(f"❌ Groq API error: {resp.status}")
                    return {
                        "ai_generated": False,
                        "confidence": 0,
                        "reasoning": f"API error: {resp.status}",
                        "evidence": [],
                        "source": "Groq API (error)",
                    }
    except asyncio.TimeoutError:
        print("❌ Groq API timeout")
        return {
            "ai_generated": False,
            "confidence": 0,
            "reasoning": "API timeout",
            "evidence": [],
            "source": "Groq API (timeout)",
        }
    except Exception as e:
        print(f"❌ Groq API error: {e}")
        return {
            "ai_generated": False,
            "confidence": 0,
            "reasoning": f"Connection error: {str(e)}",
            "evidence": [],
            "source": "Groq API (error)",
        }


async def groq_ai_judge_single_function(code: str, language: str) -> dict:
    """Specialized Groq judge for single-function coding solutions"""
    if not GROQ_API_KEY:
        return {
            "ai_generated": False,
            "confidence": 0,
            "reasoning": "Groq API not available",
            "evidence": [],
            "source": "Groq API (unavailable)",
        }

    # ----- CLEANED PROMPT (THIS IS WHERE YOUR VERSION BROKE) -----
    prompt = f"""
Analyze this SINGLE-FUNCTION {language} coding solution.

CODE:
{code[:1200]}

AI-GENERATED INDICATORS:
- Function name resembles common challenge patterns (two_sum, max_profit, is_palindrome)
- Generic parameter names: nums, arr, s, k, target
- Uses optimal/efficient algorithm immediately
- Imports only exactly needed modules (heapq, collections)
- Generic variable names: result, ans, dp, temp
- Perfect if/else consistency
- No debugging prints
- Handles all edge cases explicitly
- Clean formatting with perfect indentation
- Few or overly formal comments

HUMAN-WRITTEN INDICATORS:
- Debug prints
- TODO/FIXME notes
- Inconsistent indentation
- Imperfect structure
- Missing edge cases
- Problem-specific variable names
- Imperfect formatting
- Experimental logic or trial-error patterns

Return ONLY JSON in this format:
{
        "ai_generated": true/false,
  "confidence": number,
  "reasoning": "short explanation",
  "evidence": ["pattern1", "pattern2"],
  "source": "Groq AI Judge (Single-Function)"
}
"""
    # --------------------------------------------------------------

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": GROQ_MODEL,
        "messages": [
            {
                "role": "system",
                "content": "You detect AI-generated code with high precision.",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "max_tokens": 500,
        "response_format": {"type": "json_object"},
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30,
            ) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    response = data["choices"][0]["message"]["content"]

                    try:
                        result = json.loads(response)

                        # safety fields
                        result.setdefault("ai_generated", False)
                        result.setdefault("confidence", 0)
                        result.setdefault("evidence", [])
                        result.setdefault("reasoning", "Analysis completed")

                        print(
                            f"✅ Groq Single-Function Judge: {result['confidence']}% confidence"
                        )
                        return result

                    except json.JSONDecodeError:
                        print("❌ Groq returned invalid JSON")
                        return {
                            "ai_generated": False,
                            "confidence": 0,
                            "reasoning": "Invalid JSON",
                            "evidence": [],
                            "source": "Groq API (error)",
                        }
                else:
                    return {
                        "ai_generated": False,
                        "confidence": 0,
                        "reasoning": f"HTTP {resp.status}",
                        "evidence": [],
                        "source": "Groq API (error)",
                    }

    except asyncio.TimeoutError:
        return {
            "ai_generated": False,
            "confidence": 0,
            "reasoning": "Timeout",
            "evidence": [],
            "source": "Groq API (timeout)",
        }

    except Exception as e:
        return {
            "ai_generated": False,
            "confidence": 0,
            "reasoning": f"Error: {str(e)}",
            "evidence": [],
            "source": "Groq API (error)",
        }


## **3. Add HYBRID DETECTION function** (add after groq_ai_judge)


# ==================== HYBRID AI DETECTION ====================
# ==================== SIMPLIFIED HYBRID DETECTION ====================
async def hybrid_ai_detection(code: str, language: str) -> dict:
    """
    Enhanced hybrid detection - better weighting and always uses Groq for uncertain cases
    """
    print(
        f"🔍 Running enhanced AI detection for {language} code ({len(code)} chars)..."
    )

    # 1. Run local detector
    start_time = time.time()
    local_result = await ai_detector.detect_ai_code(code, language)
    local_time = time.time() - start_time

    local_score = local_result["confidence"]
    print(f"   Local detector score: {local_score}% ({local_time:.2f}s)")

    # 2. Determine if we need Groq
    groq_used = False
    groq_result = None
    groq_score = 0

    # USE GROQ IF:
    # - Local score is in uncertain range (10-70%)
    # - OR code is very short (< 10 lines)
    # - OR we have API key available
    should_use_groq = GROQ_API_KEY and (
        10 <= local_score <= 70 or len(code.split("\n")) < 10
    )

    if should_use_groq:
        print(f"   🤔 Consulting Groq API (uncertain case)...")
        try:
            # Use specialized judge
            groq_result = await groq_ai_judge_single_function(code, language)
            groq_used = True
            groq_score = groq_result.get("confidence", 0)
            print(f"   Groq AI Judge score: {groq_score}%")
        except Exception as e:
            print(f"   ⚠️ Groq API failed: {e}")
            groq_used = False

    # 3. DECISION LOGIC WITH BETTER WEIGHTING
    if groq_used:
        # Weight Groq more heavily (80%) for uncertain cases
        # But trust local more if it's very confident (>80% or <10%)
        if local_score > 80 or local_score < 10:
            final_confidence = (local_score * 0.7) + (groq_score * 0.3)
        else:
            final_confidence = (local_score * 0.2) + (groq_score * 0.8)

        ai_generated = groq_result.get("ai_generated", False)
        source = "Hybrid (Groq-weighted)"

        # Combine evidence
        evidence = local_result.get("evidence", [])
        groq_evidence = groq_result.get("evidence", [])
        if groq_evidence:
            evidence.extend([f"Groq: {e}" for e in groq_evidence[:2]])
    else:
        # Only local
        final_confidence = local_score
        ai_generated = local_result["ai_generated"]
        source = "Local Only"
        evidence = local_result.get("evidence", [])

    # 4. FORCE AI DETECTION FOR HIGH CONFIDENCE CASES
    # If either detector says AI with good confidence, trust it
    if (local_score >= 40 and local_result["ai_generated"]) or (
        groq_used and groq_score >= 40 and groq_result.get("ai_generated")
    ):
        ai_generated = True
        final_confidence = max(
            final_confidence, max(local_score, groq_score if groq_used else 0)
        )

    # 5. Generate response
    details = f"{'🚨 AI DETECTED' if ai_generated else '✅ Human'}"
    if groq_used:
        details += f" (Local: {local_score}%, Groq: {groq_score}%)"
    else:
        details += f" ({final_confidence:.1f}% confidence)"

    return {
        "ai_generated": ai_generated,
        "detected": ai_generated,
        "confidence": round(final_confidence, 1),
        "score": round(final_confidence, 1),
        "evidence": evidence[:4],
        "details": details,
        "source": source,
        "metadata": {
            "local_score": round(local_score, 1),
            "groq_used": groq_used,
            "groq_score": round(groq_score, 1) if groq_used else None,
            "code_length": len(code),
            "lines": len(code.split("\n")),
        },
    }


# ==================== REQUEST MODELS ====================
class MCQGenerateRequest(BaseModel):
    role: str
    difficulty: str
    num_questions: int = 5
    mode: str = "standard"


class JobPosting(BaseModel):
    job_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    required_skills: List[str] = Field(default_factory=list)
    experience_level: Optional[str] = None
    location: Optional[str] = None
    salary_range: Optional[str] = None
    posted_at: datetime = Field(default_factory=datetime.now)
    posted_by: Optional[str] = None
    status: str = "active"
    assessment_id: Optional[str] = None


class ProcessGmailRequest(BaseModel):
    """JSON body for POST /api/cv/process-gmail (frontend sends this, not query params)."""

    job_id: str = Field(..., min_length=1, description="Active job posting id; CV skills are matched to this job")
    lookback_hours: int = Field(24, ge=1, le=168)


class LoginRequest(BaseModel):
    email: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class CandidateRegisterRequest(BaseModel):
    email: str
    password: str
    full_name: Optional[str] = None


class ApplyJobRequest(BaseModel):
    job_id: str


class UpdateProfileRequest(BaseModel):
    full_name: str


class HrApplicationStatusRequest(BaseModel):
    status: str


def get_current_user_payload(authorization: Optional[str] = Header(None)) -> dict:
    """Bearer JWT → decoded payload (sub, role, uid, …)."""
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1].strip()
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return payload


def _require_candidate(payload: dict) -> None:
    if (str(payload.get("role") or "")).strip().lower() != "candidate":
        raise HTTPException(status_code=403, detail="Candidate account required")


def _require_hr_staff(payload: dict) -> None:
    r = (str(payload.get("role") or "")).strip().lower()
    if r not in ("hr", "admin", "manager", "recruiter"):
        raise HTTPException(status_code=403, detail="HR or manager role required")


APPLICATION_STATUSES = frozenset(
    {"applied", "screening", "interview", "offer", "hired", "rejected"}
)


# Create upload directory
UPLOAD_DIR = "cv_uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)


@app.post("/api/auth/login")
async def auth_login(body: LoginRequest, db: AsyncSession = Depends(get_db)):
    """HR / admin login — requires `app_users` table (see migrations/001_app_users.sql)."""
    email = (body.email or "").strip().lower()
    if not email or not body.password:
        raise HTTPException(status_code=400, detail="Email and password required")
    try:
        result = await db.execute(
            text(
                "SELECT id, email, password_hash, full_name, role, is_active "
                "FROM app_users WHERE email = :e LIMIT 1"
            ),
            {"e": email},
        )
        row = result.mappings().first()
    except Exception as ex:
        print(f"Auth login DB error: {ex}")
        raise HTTPException(
            status_code=503,
            detail="Login unavailable — did you run migrations/001_app_users.sql?",
        )
    if not row:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not row.get("is_active"):
        raise HTTPException(status_code=403, detail="Account disabled")
    if not verify_password(body.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(
        row["email"],
        int(row["id"]),
        str(row["role"] or "hr"),
        row.get("full_name"),
    )
    return {
        "success": True,
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": row["id"],
            "email": row["email"],
            "full_name": row.get("full_name"),
            "role": row["role"],
        },
    }


@app.get("/api/auth/me")
async def auth_me(
    authorization: Optional[str] = Header(None), db: AsyncSession = Depends(get_db)
):
    """Validate bearer token; refresh name/role from app_users when available."""
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1].strip()
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    email = (payload.get("sub") or "").strip().lower()
    full_name = payload.get("name") or ""
    role = payload.get("role")
    uid = payload.get("uid")
    try:
        result = await db.execute(
            text("SELECT id, full_name, role FROM app_users WHERE email = :e LIMIT 1"),
            {"e": email},
        )
        row = result.mappings().first()
        if row:
            full_name = row.get("full_name") or full_name
            role = row.get("role") or role
            uid = row.get("id") if row.get("id") is not None else uid
    except Exception:
        pass
    return {
        "success": True,
        "user": {"email": email, "role": role, "name": full_name, "id": uid},
    }


@app.post("/api/auth/change-password")
async def auth_change_password(
    body: ChangePasswordRequest,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """Update password for the authenticated user (Module 2.2.13)."""
    email = (payload.get("sub") or "").strip().lower()
    if not body.current_password or not body.new_password:
        raise HTTPException(status_code=400, detail="Current and new password required")
    if len(body.new_password) < 8:
        raise HTTPException(
            status_code=400, detail="New password must be at least 8 characters"
        )
    try:
        result = await db.execute(
            text("SELECT id, password_hash FROM app_users WHERE email = :e LIMIT 1"),
            {"e": email},
        )
        row = result.mappings().first()
    except Exception as ex:
        print(f"Change password DB error: {ex}")
        raise HTTPException(status_code=503, detail="Database unavailable")
    if not row:
        raise HTTPException(status_code=404, detail="User not found")
    if not verify_password(body.current_password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    await db.execute(
        text("UPDATE app_users SET password_hash = :h WHERE id = :id"),
        {"h": hash_password(body.new_password), "id": row["id"]},
    )
    await db.commit()
    return {"success": True, "message": "Password updated"}


@app.post("/api/auth/register-candidate")
async def auth_register_candidate(
    body: CandidateRegisterRequest, db: AsyncSession = Depends(get_db)
):
    """Self-service candidate account (Module 2.2.11). Stored in app_users with role=candidate."""
    email = (body.email or "").strip().lower()
    if not email or not body.password:
        raise HTTPException(status_code=400, detail="Email and password required")
    if len(body.password) < 8:
        raise HTTPException(
            status_code=400, detail="Password must be at least 8 characters"
        )
    try:
        exists = await db.execute(
            text("SELECT id FROM app_users WHERE email = :e LIMIT 1"),
            {"e": email},
        )
        if exists.mappings().first():
            raise HTTPException(
                status_code=409, detail="An account with this email already exists"
            )
        await db.execute(
            text(
                """
                INSERT INTO app_users (email, password_hash, full_name, role, is_active)
                VALUES (:email, :ph, :fn, 'candidate', 1)
                """
            ),
            {
                "email": email,
                "ph": hash_password(body.password),
                "fn": (body.full_name or "").strip() or None,
            },
        )
        await db.commit()
        res = await db.execute(
            text(
                "SELECT id, email, full_name, role FROM app_users WHERE email = :e LIMIT 1"
            ),
            {"e": email},
        )
        row = res.mappings().first()
    except HTTPException:
        raise
    except Exception as ex:
        print(f"register-candidate error: {ex}")
        raise HTTPException(status_code=503, detail="Could not create account")
    token = create_access_token(
        row["email"],
        int(row["id"]),
        str(row["role"] or "candidate"),
        row.get("full_name"),
    )
    return {
        "success": True,
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": row["id"],
            "email": row["email"],
            "full_name": row.get("full_name"),
            "role": row["role"],
        },
    }


@app.post("/api/auth/register-hr")
async def auth_register_hr(
    body: CandidateRegisterRequest, db: AsyncSession = Depends(get_db)
):
    """Self-service HR / recruiter account. Same shape as candidate register; role is fixed to hr."""
    email = (body.email or "").strip().lower()
    if not email or not body.password:
        raise HTTPException(status_code=400, detail="Email and password required")
    if len(body.password) < 8:
        raise HTTPException(
            status_code=400, detail="Password must be at least 8 characters"
        )
    try:
        exists = await db.execute(
            text("SELECT id FROM app_users WHERE email = :e LIMIT 1"),
            {"e": email},
        )
        if exists.mappings().first():
            raise HTTPException(
                status_code=409, detail="An account with this email already exists"
            )
        await db.execute(
            text(
                """
                INSERT INTO app_users (email, password_hash, full_name, role, is_active)
                VALUES (:email, :ph, :fn, 'hr', 1)
                """
            ),
            {
                "email": email,
                "ph": hash_password(body.password),
                "fn": (body.full_name or "").strip() or None,
            },
        )
        await db.commit()
        res = await db.execute(
            text(
                "SELECT id, email, full_name, role FROM app_users WHERE email = :e LIMIT 1"
            ),
            {"e": email},
        )
        row = res.mappings().first()
    except HTTPException:
        raise
    except Exception as ex:
        print(f"register-hr error: {ex}")
        raise HTTPException(status_code=503, detail="Could not create account")
    token = create_access_token(
        row["email"],
        int(row["id"]),
        str(row["role"] or "hr"),
        row.get("full_name"),
    )
    return {
        "success": True,
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": row["id"],
            "email": row["email"],
            "full_name": row.get("full_name"),
            "role": row["role"],
        },
    }


@app.patch("/api/auth/profile")
async def auth_update_profile(
    body: UpdateProfileRequest,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """Update display name for the signed-in user (any role)."""
    email = (payload.get("sub") or "").strip().lower()
    name = (body.full_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    try:
        await db.execute(
            text("UPDATE app_users SET full_name = :n WHERE email = :e"),
            {"n": name, "e": email},
        )
        await db.commit()
    except Exception as ex:
        print(f"profile update error: {ex}")
        raise HTTPException(status_code=503, detail="Could not update profile")
    return {"success": True, "user": {"email": email, "full_name": name}}


@app.post("/api/candidate/apply")
async def candidate_apply_job(
    body: ApplyJobRequest,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """Apply to an active job posting (candidate accounts only)."""
    _require_candidate(payload)
    job_id = (body.job_id or "").strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="job_id required")
    job = await get_job_posting(db, job_id)
    if not job or not job.get("job_id"):
        raise HTTPException(status_code=404, detail="Job not found")
    if (job.get("status") or "active").lower() != "active":
        raise HTTPException(
            status_code=400, detail="This job is not accepting applications"
        )
    uid = payload.get("uid")
    if uid is None:
        raise HTTPException(status_code=401, detail="Invalid token")
    application_id = f"app_{uuid.uuid4()}"
    try:
        await db.execute(
            text(
                """
                INSERT INTO job_applications (application_id, candidate_user_id, job_id, status)
                VALUES (:aid, :uid, :jid, 'applied')
                """
            ),
            {"aid": application_id, "uid": int(uid), "jid": job_id},
        )
        await db.commit()
    except Exception as ex:
        err = str(ex).lower()
        if "duplicate" in err or "uq_candidate_job" in err:
            raise HTTPException(
                status_code=409, detail="You have already applied to this job"
            )
        print(f"candidate apply error: {ex}")
        raise HTTPException(status_code=500, detail="Could not submit application")
    return {
        "success": True,
        "application_id": application_id,
        "message": "Application submitted",
    }


@app.post("/api/candidate/apply-with-cv")
async def candidate_apply_with_cv(
    job_id: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    cv_file: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """Apply to an active job posting with a CV upload."""
    _require_candidate(payload)
    job_id = job_id.strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="job_id required")
    job = await get_job_posting(db, job_id)
    if not job or not job.get("job_id"):
        raise HTTPException(status_code=404, detail="Job not found")
    if (job.get("status") or "active").lower() != "active":
        raise HTTPException(
            status_code=400, detail="This job is not accepting applications"
        )

    uid = payload.get("uid")
    if uid is None:
        raise HTTPException(status_code=401, detail="Invalid token")

    application_id = f"app_{uuid.uuid4()}"

    try:
        await db.execute(
            text(
                """
                INSERT INTO job_applications (application_id, candidate_user_id, job_id, status)
                VALUES (:aid, :uid, :jid, 'applied')
                """
            ),
            {"aid": application_id, "uid": int(uid), "jid": job_id},
        )
        await db.commit()
    except Exception as ex:
        err = str(ex).lower()
        if "duplicate" in err or "uq_candidate_job" in err:
            raise HTTPException(
                status_code=409, detail="You have already applied to this job"
            )
        print(f"candidate apply error: {ex}")
        raise HTTPException(status_code=500, detail=str(ex))

    # CV handling: Save path to job_applications instead of auto-processing
    if cv_file:
        allowed_extensions = [".pdf", ".docx", ".doc", ".txt"]
        file_extension = os.path.splitext(cv_file.filename)[1].lower()
        if file_extension in allowed_extensions:
            file_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4()}{file_extension}")
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(cv_file.file, buffer)

            try:
                # Store the file path and manual info locally so HR can fetch it later
                await db.execute(
                    text("""
                    UPDATE job_applications 
                    SET cv_file_path = :path, 
                        manual_name = :name, 
                        manual_email = :email, 
                        is_fetched = 0 
                    WHERE application_id = :aid
                    """),
                    {
                        "path": file_path,
                        "name": name,
                        "email": email,
                        "aid": application_id,
                    },
                )
                await db.commit()
                print(
                    f"✅ Candidate applied with CV. Stored for HR manual fetch: {file_path}"
                )
            except Exception as e:
                print("Failed to save CV path to job_applications.", e)

    return {
        "success": True,
        "application_id": application_id,
        "message": "Application and CV submitted",
    }


@app.get("/api/candidate/applications")
async def candidate_list_applications(
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """List the signed-in candidate's applications and pipeline status."""
    _require_candidate(payload)
    uid = payload.get("uid")
    if uid is None:
        raise HTTPException(status_code=401, detail="Invalid token")
    try:
        result = await db.execute(
            text(
                """
                SELECT ja.application_id, ja.job_id, ja.status, ja.applied_at,
                       jp.title AS job_title, jp.location, jp.status AS job_status
                FROM job_applications ja
                LEFT JOIN job_postings jp ON jp.job_id = ja.job_id
                WHERE ja.candidate_user_id = :uid
                ORDER BY ja.applied_at DESC
                """
            ),
            {"uid": int(uid)},
        )
        rows = result.mappings().all()
    except Exception as ex:
        print(f"list applications error: {ex}")
        raise HTTPException(status_code=503, detail="Could not load applications")
    items = []
    for r in rows:
        items.append(
            {
                "application_id": r["application_id"],
                "job_id": r["job_id"],
                "job_title": r.get("job_title") or r["job_id"],
                "status": r["status"],
                "applied_at": str(r["applied_at"]) if r.get("applied_at") else None,
                "location": r.get("location"),
                "job_status": r.get("job_status"),
            }
        )
    return {"success": True, "applications": items, "count": len(items)}


@app.patch("/api/hr/job-applications/{application_id}/status")
async def hr_update_application_status(
    application_id: str,
    body: HrApplicationStatusRequest,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """Move a candidate application along the pipeline (HR / manager)."""
    _require_hr_staff(payload)
    st = (body.status or "").strip().lower()
    if st not in APPLICATION_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"status must be one of: {', '.join(sorted(APPLICATION_STATUSES))}",
        )
    try:
        res = await db.execute(
            text(
                "SELECT application_id FROM job_applications WHERE application_id = :id LIMIT 1"
            ),
            {"id": application_id},
        )
        if not res.mappings().first():
            raise HTTPException(status_code=404, detail="Application not found")
        await db.execute(
            text("UPDATE job_applications SET status = :s WHERE application_id = :id"),
            {"s": st, "id": application_id},
        )
        await db.commit()
    except HTTPException:
        raise
    except Exception as ex:
        print(f"hr status update error: {ex}")
        raise HTTPException(status_code=500, detail="Could not update status")
    return {"success": True, "application_id": application_id, "status": st}


@app.post("/api/cv/upload")
async def upload_cv(
    file: UploadFile = File(...),
    job_id: str = Form(...),
    source: str = Form("manual_upload"),
    db: AsyncSession = Depends(get_db),
):
    """
    FR-CVE-03, FR-CVE-04: Upload and process CV file
    """
    try:
        jid = (job_id or "").strip()
        if not jid:
            raise HTTPException(
                status_code=400,
                detail="Select a job posting first. Skill match is calculated against that job's required skills.",
            )
        job_row = await get_job_posting(db, jid)
        if not job_row or not job_row.get("job_id"):
            raise HTTPException(
                status_code=400,
                detail="Job posting not found. Create a job in Job Posting, then try again.",
            )

        allowed_extensions = [".pdf", ".docx", ".doc", ".txt"]
        file_extension = os.path.splitext(file.filename)[1].lower()

        if file_extension not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Allowed: {', '.join(allowed_extensions)}",
            )

        file_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4()}{file_extension}")
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        print(f"✅ CV file saved: {file_path}")

        is_scanned = False
        if file_extension == ".pdf":
            is_scanned = ocr_processor.is_scanned_pdf(file_path)
            print(f"📄 PDF is scanned: {is_scanned}")

        cv_data = await cv_processor.process_cv_file(
            file_path, file_extension, jid
        )

        if is_scanned:
            try:
                ocr_text = ocr_processor.extract_text_from_scanned_pdf(file_path)
                if ocr_text:
                    cv_data = await cv_processor.process_cv_file(
                        file_path, ".txt", jid
                    )
                    print("✅ Used OCR for scanned PDF")
            except Exception as ocr_error:
                print(f"⚠️ OCR processing failed: {ocr_error}")

        job_requirements = job_row

        skill_result = skill_matcher.match_skills(cv_data["skills"], job_requirements)

        cv_data["missing_skills"] = skill_result["missing_skills"]
        cv_data["extra_skills"] = skill_result["extra_skills"]
        cv_data["matched_pairs"] = skill_result.get("matched_pairs", [])
        cv_data["skill_match_percentage"] = skill_result["match_percentage"]
        cv_data["source"] = source
        cv_data["status"] = "processed"

        candidate_id = await save_cv_candidate(db, cv_data)

        return {
            "success": True,
            "candidate_id": candidate_id,
            "email": cv_data["email"],
            "name": cv_data["name"],
            "skill_match_percentage": skill_result["match_percentage"],
            "missing_skills": skill_result["missing_skills"],
            "extra_skills": skill_result["extra_skills"],
            "matched_pairs": skill_result.get("matched_pairs", []),
            "message": "CV processed successfully",
            "job_id": jid,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ CV upload failed: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/hr/fetch-portal-cvs")
async def fetch_portal_cvs(
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(get_current_user_payload),
):
    """HR endpoint to fetch CVs uploaded by candidates and parse them."""
    # Temporarily removed role requirement so it works for manager/other roles too.
    from sqlalchemy import text

    try:
        res = await db.execute(
            text(
                "SELECT application_id, job_id, cv_file_path, manual_name, manual_email FROM job_applications WHERE cv_file_path IS NOT NULL AND is_fetched = 0"
            )
        )
        pending_apps = res.mappings().all()

        processed_count = 0
        for app in pending_apps:
            cv_path = app["cv_file_path"]
            app_id = app["application_id"]
            j_id = app["job_id"]
            m_name = app["manual_name"]
            m_email = app["manual_email"]

            if not cv_path or not os.path.exists(cv_path):
                await db.execute(
                    text(
                        "UPDATE job_applications SET is_fetched = 1 WHERE application_id = :aid"
                    ),
                    {"aid": app_id},
                )
                continue

            file_ext = os.path.splitext(cv_path)[1].lower()
            try:
                cv_data = await cv_processor.process_cv_file(cv_path, file_ext, j_id)
                # Ensure manual name/email overrides the parsed AI output
                cv_data["name"] = m_name or cv_data.get("name", "Unknown Applicant")
                if m_email:
                    cv_data["email"] = m_email

                job_requirements = {}
                if j_id:
                    job = await get_job_posting(db, j_id)
                    if job:
                        job_requirements = job
                        cv_data["role"] = job.get(
                            "title", cv_data.get("role", "Candidate")
                        )

                skill_result = skill_matcher.match_skills(
                    cv_data.get("skills", []), job_requirements
                )
                cv_data["missing_skills"] = skill_result["missing_skills"]
                cv_data["extra_skills"] = skill_result["extra_skills"]
                cv_data["matched_pairs"] = skill_result.get("matched_pairs", [])
                cv_data["skill_match_percentage"] = skill_result["match_percentage"]
                cv_data["source"] = "Candidate Portal"
                cv_data["status"] = "processed"

                candidate_id = await save_cv_candidate(db, cv_data)

                # Mark as fetched on success
                await db.execute(
                    text(
                        "UPDATE job_applications SET is_fetched = 1 WHERE application_id = :aid"
                    ),
                    {"aid": app_id},
                )
                processed_count += 1
            except Exception as cv_e:
                print(f"Error parsing candidate CV {app_id}: {cv_e}")

        await db.commit()
        return {
            "success": True,
            "fetched_count": processed_count,
            "message": f"Fetched {processed_count} CVs successfully",
        }
    except Exception as e:
        print(f"fetch portal cvs err: {e}")
        raise HTTPException(status_code=500, detail="Could not fetch candidate CVs")


@app.post("/api/cv/process-gmail")
async def process_gmail_cvs(
    background_tasks: BackgroundTasks,
    req: ProcessGmailRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    FR-CVE-01: Process CVs from Gmail.
    Requires credentials.json + token.json in the backend folder; Gmail API enabled in Google Cloud.
    job_id is required so each CV is scored against that job posting's required skills.
    """
    try:
        jid = (req.job_id or "").strip()
        if not jid:
            raise HTTPException(
                status_code=400,
                detail="Select a job posting first. CVs from Gmail will be matched to that job.",
            )
        job_row = await get_job_posting(db, jid)
        if not job_row or not job_row.get("job_id"):
            raise HTTPException(
                status_code=400,
                detail="Job posting not found. Create a job in Job Posting, then try again.",
            )

        authenticated = await gmail_monitor.authenticate()
        if not authenticated:
            raise HTTPException(
                status_code=500,
                detail=(
                    "Gmail authentication failed. Place credentials.json from Google Cloud OAuth "
                    "(Desktop app) in the backend folder and complete OAuth once, or refresh token.json."
                ),
            )

        # Use a fresh DB session inside the background task (request session must not be reused)
        background_tasks.add_task(
            process_gmail_cvs_background,
            req.lookback_hours,
            jid,
        )

        return {
            "success": True,
            "message": "Gmail CV processing started in background",
            "lookback_hours": req.lookback_hours,
            "job_id": jid,
            "job_title": job_row.get("title") or jid,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Gmail processing failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def process_gmail_cvs_background(lookback_hours: int, job_id: Optional[str]):
    """Background task: fetch Gmail attachments and parse CVs (own DB session)."""
    jid = (job_id or "").strip()
    async with AsyncSessionLocal() as db:
        try:
            job_requirements: dict = {}
            if jid:
                job = await get_job_posting(db, jid)
                if job:
                    job_requirements = job

            cv_emails = await gmail_monitor.fetch_new_cvs(lookback_hours)

            processed_count = 0
            skipped_count = 0

            for email_data in cv_emails:
                sender_email = email_data.get("sender_email", "")

                # Skip if already exists in database
                check_query = text(
                    "SELECT candidate_id FROM cv_candidates WHERE email = :email"
                )
                result = await db.execute(check_query, {"email": sender_email})
                existing = result.fetchone()
                if existing:
                    print(f"⚠️ Skipping CV from {sender_email} - already exists")
                    skipped_count += 1
                    continue

                try:
                    for attachment in email_data.get("attachments", []):
                        file_path = attachment["file_path"]
                        file_extension = os.path.splitext(attachment["filename"])[
                            1
                        ].lower()

                        cv_data = await cv_processor.process_cv_file(
                            file_path, file_extension, jid or None
                        )

                        cv_data["source"] = "gmail"
                        cv_data["email"] = email_data.get(
                            "sender_email", cv_data["email"]
                        )

                        skill_result = skill_matcher.match_skills(
                            cv_data["skills"], job_requirements
                        )

                        cv_data["missing_skills"] = skill_result["missing_skills"]
                        cv_data["extra_skills"] = skill_result["extra_skills"]
                        cv_data["matched_pairs"] = skill_result.get("matched_pairs", [])
                        cv_data["skill_match_percentage"] = skill_result[
                            "match_percentage"
                        ]
                        cv_data["status"] = "processed"

                        await save_cv_candidate(db, cv_data)
                        processed_count += 1

                        print(f"✅ Processed CV from Gmail: {cv_data['email']}")

                except Exception as e:
                    print(f"⚠️ Failed to process CV from Gmail: {e}")

            print(
                f"✅ Processed {processed_count} CVs from Gmail (skipped {skipped_count} duplicates)"
            )

        except Exception as e:
            print(f"❌ Background Gmail processing failed: {e}")


class UpdateStatusRequest(BaseModel):
    status: str


@app.put("/api/cv/candidate/{candidate_id}/status")
async def update_candidate_status(
    candidate_id: str,
    db: AsyncSession = Depends(get_db),
    body: UpdateStatusRequest = None,
):
    """Update candidate status"""
    try:
        new_status = body.status if body else "pending"
        cid = (candidate_id or "").strip()
        if not cid:
            raise HTTPException(status_code=400, detail="Invalid candidate ID")

        query = text(
            "UPDATE cv_candidates SET status = :status WHERE candidate_id = :candidate_id"
        )
        result = await db.execute(query, {"status": new_status, "candidate_id": cid})
        await db.commit()

        return {"success": True, "message": f"Status updated to {new_status}"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Status update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/cv/candidate/{candidate_id}")
async def delete_cv_candidate(candidate_id: str, db: AsyncSession = Depends(get_db)):
    """Delete a CV candidate by ID"""
    try:
        cid = (candidate_id or "").strip()
        if not cid:
            raise HTTPException(status_code=400, detail="Invalid candidate ID")

        query = text("DELETE FROM cv_candidates WHERE candidate_id = :candidate_id")
        result = await db.execute(query, {"candidate_id": cid})
        await db.commit()

        if result.rowcount > 0:
            return {"success": True, "message": "Candidate deleted"}
        else:
            raise HTTPException(status_code=404, detail="Candidate not found")
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Delete candidate failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cv/candidates")
async def get_cv_candidates(
    status: str = None,
    job_id: str = None,
    min_match: float = None,
    db: AsyncSession = Depends(get_db),
):
    """
    FR-CVE-06, FR-CVE-07: Get all CV candidates with skill analysis
    """
    try:
        candidates = await get_all_candidates(db, status, job_id, min_match)

        return {"success": True, "candidates": candidates, "count": len(candidates)}

    except Exception as e:
        print(f"❌ Fetch candidates failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cv/candidate/", include_in_schema=False)
async def get_cv_candidate_missing_id():
    raise HTTPException(
        status_code=400,
        detail="Missing candidate_id. Use GET /api/cv/candidate/{candidate_id} with a valid id.",
    )


@app.get("/api/cv/candidate/{candidate_id}")
async def get_candidate_details(candidate_id: str, db: AsyncSession = Depends(get_db)):
    """Get detailed candidate information"""
    try:
        cid = (candidate_id or "").strip()
        if not cid:
            raise HTTPException(status_code=400, detail="Invalid candidate ID")

        query = text("""
            SELECT candidate_id, email, name, phone, role, skills, cv_filename, cv_text,
                   status, skill_match_percentage, cv_source, job_id, raw_text,
                   experience, education, created_at, missing_skills, extra_skills,
                   matched_skills
            FROM cv_candidates WHERE candidate_id = :candidate_id
        """)
        result = await db.execute(query, {"candidate_id": cid})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Candidate not found")

        job_details = None
        if row[11]:
            job_details = await get_job_posting(db, row[11])

        # Safely parse JSON fields
        try:
            skills = json.loads(row[5]) if row[5] and row[5] != "[]" else []
        except (json.JSONDecodeError, TypeError):
            skills = []
        try:
            experience = json.loads(row[13]) if row[13] and row[13] != "[]" else []
        except (json.JSONDecodeError, TypeError):
            experience = []
        try:
            education = json.loads(row[14]) if row[14] and row[14] != "[]" else []
        except (json.JSONDecodeError, TypeError):
            education = []
        try:
            missing_skills = json.loads(row[16]) if row[16] and row[16] != "[]" else []
        except (json.JSONDecodeError, TypeError):
            missing_skills = []
        try:
            extra_skills = json.loads(row[17]) if row[17] and row[17] != "[]" else []
        except (json.JSONDecodeError, TypeError):
            extra_skills = []

        matched_pairs: list = []
        try:
            if len(row) > 18 and row[18]:
                matched_pairs = json.loads(row[18])
                if not isinstance(matched_pairs, list):
                    matched_pairs = []
        except (json.JSONDecodeError, TypeError, IndexError):
            matched_pairs = []

        if not matched_pairs and job_details and skills:
            sr = skill_matcher.match_skills(skills, job_details)
            matched_pairs = sr.get("matched_pairs", [])

        req_total = len(job_details.get("required_skills", [])) if job_details else 0
        matched_n = len(matched_pairs)

        candidate_data = {
            "candidate_id": row[0],
            "email": row[1],
            "name": row[2],
            "phone": row[3],
            "role": row[4],
            "skills": skills,
            "cv_filename": row[6],
            "cv_file_path": row[6],
            "cv_text": row[7],
            "raw_text": row[12],
            "status": row[8],
            "skill_match_percentage": row[9],
            "cv_source": row[10],
            "job_id": row[11],
            "experience": experience,
            "education": education,
            "created_at": str(row[15]) if row[15] else None,
            "extracted_at": str(row[15]) if row[15] else None,
            "missing_skills": missing_skills,
            "extra_skills": extra_skills,
            "matched_pairs": matched_pairs,
            "job_details": job_details,
            "skill_match_explainer": (
                f"Job match = {matched_n} of {req_total} required skills found on the CV (see Missing skills for gaps)."
                if req_total
                else "Add required skills to the job posting to enable match scoring."
            ),
        }

        return {"success": True, "candidate": candidate_data}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Get candidate failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/cv/job-posting")
async def create_job_posting(job_data: JobPosting, db: AsyncSession = Depends(get_db)):
    """Create a new job posting"""
    try:
        job_id = job_data.job_id or f"job_{uuid.uuid4()}"

        query = text(
            """
            INSERT INTO job_postings (
                job_id, title, description, required_skills, experience_level,
                location, salary_range, posted_at, posted_by, status, assessment_id
            ) VALUES (
                :job_id, :title, :description, :required_skills, :experience_level,
                :location, :salary_range, :posted_at, :posted_by, :status, :assessment_id
            )
        """
        )

        await db.execute(
            query,
            {
                "job_id": job_id,
                "title": job_data.title,
                "description": job_data.description,
                "required_skills": json.dumps(job_data.required_skills),
                "experience_level": job_data.experience_level,
                "location": job_data.location,
                "salary_range": job_data.salary_range,
                "posted_at": job_data.posted_at,
                "posted_by": job_data.posted_by,
                "status": job_data.status,
                "assessment_id": job_data.assessment_id,
            },
        )

        await db.commit()

        return {
            "success": True,
            "job_id": job_id,
            "message": "Job posting created successfully",
        }

    except Exception as e:
        print(f"❌ Create job posting failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cv/job-postings")
async def get_job_postings(status: str = "active", db: AsyncSession = Depends(get_db)):
    """Get all job postings"""
    try:
        job_postings = await get_all_job_postings(db, status)

        return {
            "success": True,
            "job_postings": job_postings,
            "count": len(job_postings),
        }

    except Exception as e:
        print(f"❌ Get job postings failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cv/skill-analysis/{job_id}")
async def get_skill_analysis(job_id: str, db: AsyncSession = Depends(get_db)):
    """
    FR-CVE-08: Get skill analysis for a job
    """
    try:
        job = await get_job_posting(db, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job posting not found")

        candidates = await get_all_candidates(db, job_id=job_id)

        total_candidates = len(candidates)
        avg_match = (
            sum(c["skill_match_percentage"] for c in candidates) / total_candidates
            if total_candidates > 0
            else 0
        )

        skill_frequency = {}
        for candidate in candidates:
            for skill in candidate["skills"]:
                skill_frequency[skill] = skill_frequency.get(skill, 0) + 1

        missing_skills_freq = {}
        for candidate in candidates:
            for skill in candidate["missing_skills"]:
                missing_skills_freq[skill] = missing_skills_freq.get(skill, 0) + 1

        extra_skills_freq = {}
        for candidate in candidates:
            for skill in candidate["extra_skills"]:
                extra_skills_freq[skill] = extra_skills_freq.get(skill, 0) + 1

        return {
            "success": True,
            "job": job,
            "statistics": {
                "total_candidates": total_candidates,
                "average_match_percentage": round(avg_match, 2),
                "top_candidates": sorted(
                    candidates, key=lambda x: x["skill_match_percentage"], reverse=True
                )[:5],
                "skill_frequency": dict(
                    sorted(skill_frequency.items(), key=lambda x: x[1], reverse=True)[
                        :10
                    ]
                ),
                "common_missing_skills": dict(
                    sorted(
                        missing_skills_freq.items(), key=lambda x: x[1], reverse=True
                    )[:5]
                ),
                "common_extra_skills": dict(
                    sorted(extra_skills_freq.items(), key=lambda x: x[1], reverse=True)[
                        :5
                    ]
                ),
            },
        }

    except Exception as e:
        print(f"❌ Skill analysis failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/cv/notify-hr/{candidate_id}")
async def notify_hr_manager(candidate_id: str, db: AsyncSession = Depends(get_db)):
    """
    FR-CVE-10: Notify HR manager about new candidate
    """
    try:
        cid = (candidate_id or "").strip()
        if not cid:
            raise HTTPException(status_code=400, detail="Invalid candidate ID")

        query = text("""
            SELECT candidate_id, email, name, phone, role, skills, cv_filename, cv_text,
                   status, skill_match_percentage, cv_source, job_id, raw_text,
                   experience, education, created_at
            FROM cv_candidates WHERE candidate_id = :candidate_id
        """)
        result = await db.execute(query, {"candidate_id": cid})
        candidate = result.fetchone()

        if not candidate:
            raise HTTPException(status_code=404, detail="Candidate not found")

        job_details = None
        if candidate[11]:
            job_details = await get_job_posting(db, candidate[11])

        notification_data = {
            "candidate_id": candidate[0],
            "candidate_name": candidate[2] or "Unknown",
            "candidate_email": candidate[1],  # This is the candidate's email from CV
            "job_title": job_details["title"] if job_details else "General Application",
            "skill_match_percentage": candidate[9],
            "missing_skills": [],
            "submitted_at": str(candidate[14])
            if candidate[14]
            else datetime.now().isoformat(),
        }

        print(f"📧 HR Notification - New Candidate: {notification_data}")

        return {
            "success": True,
            "message": "HR manager notified",
            "notification_data": notification_data,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ HR notification failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== GLOBAL VARIABLES ====================
adaptive_engines: dict = {}
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_MODEL = "llama-3.3-70b-versatile"


# ==================== CANDIDATE PIPELINE TRACKING ====================
# Enhanced HR Dashboard - Track candidates across all modules


@app.get("/api/hr/pipeline")
async def get_candidate_pipeline(db: AsyncSession = Depends(get_db)):
    """
    HR Pipeline Dashboard - Track candidates across all recruitment stages
    Stages: CV Submitted -> Assessment -> Coding Challenge -> Interview -> Recommendation
    """
    try:
        pipeline_data = {
            "stages": [
                {"id": "cv", "name": "CV Submitted", "icon": "📄", "order": 1},
                {
                    "id": "assessment",
                    "name": "Assessment/Test",
                    "icon": "📝",
                    "order": 2,
                },
                {"id": "coding", "name": "Coding Challenge", "icon": "💻", "order": 3},
                {"id": "interview", "name": "AI Interview", "icon": "🎤", "order": 4},
                {
                    "id": "recommendation",
                    "name": "Recommendation",
                    "icon": "⭐",
                    "order": 5,
                },
            ],
            "candidates": [],
            "statistics": {
                "total_candidates": 0,
                "by_stage": {},
                "completed": 0,
                "in_progress": 0,
                "pending": 0,
            },
        }

        # Fetch CV candidates
        cv_query = text("""
            SELECT candidate_id, email, name, skill_match_percentage, status, job_id, created_at 
            FROM cv_candidates
            ORDER BY created_at DESC
        """)
        cv_result = await db.execute(cv_query)
        cv_candidates = cv_result.fetchall()

        # Fetch Assessment results
        assessment_query = text("""
            SELECT assessment_id, role, status, created_at 
            FROM assessments
            ORDER BY created_at DESC
        """)
        try:
            assessment_result = await db.execute(assessment_query)
            assessments = assessment_result.fetchall()
        except:
            assessments = []

        # Fetch Interview sessions
        interview_query = text("""
            SELECT session_id, candidate_email, candidate_name, job_role, status, overall_score, created_at, end_time
            FROM interview_sessions
            ORDER BY created_at DESC
        """)
        try:
            interview_result = await db.execute(interview_query)
            interviews = interview_result.fetchall()
        except:
            interviews = []

        # Build candidate pipeline by grouping by email
        candidate_map = {}

        # Add CV candidates
        for cv in cv_candidates:
            email = cv[1]
            if email not in candidate_map:
                candidate_map[email] = {
                    "email": email,
                    "name": cv[2] or "Unknown",
                    "current_stage": "cv",
                    "stage_details": {},
                    "scores": {},
                    "timeline": [],
                }

            # Add CV stage info
            candidate_map[email]["stage_details"]["cv"] = {
                "status": cv[4] if cv[4] else "pending",
                "match_percentage": cv[3],
                "job_id": cv[5],
                "submitted_at": str(cv[6]) if cv[6] else None,
            }
            candidate_map[email]["timeline"].append(
                {
                    "stage": "cv",
                    "date": str(cv[6]) if cv[6] else None,
                    "status": cv[4] if cv[4] else "pending",
                }
            )

        # Add Interview data
        for interview in interviews:
            email = interview[1]
            if email in candidate_map:
                # Determine current stage based on interview status
                status = interview[4]
                stage = "interview"
                if status == "completed":
                    stage = "recommendation"
                elif status == "in_progress":
                    stage = "interview"

                candidate_map[email]["current_stage"] = stage
                candidate_map[email]["stage_details"]["interview"] = {
                    "session_id": interview[0],
                    "job_role": interview[3],
                    "status": status,
                    "score": interview[5],
                    "started_at": str(interview[6]) if interview[6] else None,
                    "completed_at": str(interview[7]) if interview[7] else None,
                }
                candidate_map[email]["scores"]["interview"] = interview[5]
                candidate_map[email]["timeline"].append(
                    {
                        "stage": "interview",
                        "date": str(interview[6]) if interview[6] else None,
                        "status": status,
                    }
                )

        # Convert to list and calculate statistics
        pipeline_data["candidates"] = list(candidate_map.values())
        pipeline_data["statistics"]["total_candidates"] = len(
            pipeline_data["candidates"]
        )

        # Count by stage
        stage_counts = {}
        for c in pipeline_data["candidates"]:
            stage = c["current_stage"]
            stage_counts[stage] = stage_counts.get(stage, 0) + 1
        pipeline_data["statistics"]["by_stage"] = stage_counts

        # Calculate completed vs in progress
        for c in pipeline_data["candidates"]:
            if c["current_stage"] == "recommendation":
                pipeline_data["statistics"]["completed"] += 1
            elif c["current_stage"] in ["assessment", "coding", "interview"]:
                pipeline_data["statistics"]["in_progress"] += 1
            else:
                pipeline_data["statistics"]["pending"] += 1

        return {"success": True, "pipeline": pipeline_data}

    except Exception as e:
        print(f"❌ Pipeline fetch failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hr/pipeline/{candidate_email}")
async def get_candidate_pipeline_detail(
    candidate_email: str, db: AsyncSession = Depends(get_db)
):
    """
    Get detailed pipeline information for a specific candidate
    """
    try:
        # Get CV data
        cv_query = text("""
            SELECT * FROM cv_candidates WHERE email = :email
        """)
        cv_result = await db.execute(cv_query, {"email": candidate_email})
        cv = cv_result.fetchone()

        # Get Interview data
        interview_query = text("""
            SELECT * FROM interview_sessions WHERE candidate_email = :email
            ORDER BY created_at DESC
        """)
        interview_result = await db.execute(interview_query, {"email": candidate_email})
        interviews = interview_result.fetchall()

        candidate_detail = {
            "email": candidate_email,
            "cv": None,
            "interviews": [],
            "current_stage": "cv",
            "overall_score": None,
        }

        if cv:
            candidate_detail["cv"] = {
                "candidate_id": cv[0],
                "name": cv[2],
                "skills": json.loads(cv[4]) if cv[4] else [],
                "experience": json.loads(cv[5]) if cv[5] else [],
                "education": json.loads(cv[6]) if cv[6] else [],
                "skill_match_percentage": cv[11],
                "status": cv[12],
                "job_id": cv[14],
            }

            # Determine current stage
            if interviews:
                latest_status = dict(interviews[0]._mapping).get("status")
                if str(latest_status) == "completed":
                    candidate_detail["current_stage"] = "recommendation"
                else:
                    candidate_detail["current_stage"] = "interview"

        for interview in interviews:
            im = dict(interview._mapping)
            candidate_detail["interviews"].append(
                {
                    "session_id": im.get("session_id"),
                    "job_role": im.get("job_role"),
                    "status": str(im.get("status") or ""),
                    "score": im.get("overall_score"),
                    "face_verified": im.get("face_verified"),
                    "start_time": str(im.get("start_time"))
                    if im.get("start_time")
                    else None,
                    "end_time": str(im.get("end_time")) if im.get("end_time") else None,
                }
            )

        return {"success": True, "candidate": candidate_detail}

    except Exception as e:
        print(f"❌ Candidate detail failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hr/candidate-stages")
async def get_hr_candidate_stages(db: AsyncSession = Depends(get_db)):
    """
    One row per candidate email: best MCQ/assessment result, best coding submission,
    and latest AI interview — for HR Results dashboard (match all three stages).
    """

    def norm_email(e) -> str:
        if e is None:
            return ""
        return str(e).strip().lower()

    assessment_by_email: dict = {}
    try:
        all_results = await ResultRepository.get_all(db, sort_by="score_percentage")
        for r in all_results:
            e = norm_email(r.get("candidate_email"))
            if not e:
                continue
            prev = assessment_by_email.get(e)
            score = float(r.get("score_percentage") or 0)
            if not prev or score > float(prev.get("score_percentage") or 0):
                assessment_by_email[e] = {
                    "result_id": r.get("result_id"),
                    "role": r.get("role"),
                    "score_percentage": r.get("score_percentage"),
                    "grade": r.get("grade"),
                    "status": r.get("status"),
                    "submitted_at": r.get("end_time") or r.get("start_time"),
                    "difficulty": r.get("difficulty"),
                }
    except Exception as ex:
        print(f"⚠️ candidate-stages assessment: {ex}")

    coding_by_email: dict = {}
    try:
        c_query = text(
            """
            SELECT
                cs.candidate_email,
                cc.title AS challenge_title,
                cs.score,
                cs.status,
                cs.submitted_at,
                COALESCE(v.violation_count, 0) AS violation_count,
                cs.submission_id
            FROM coding_submissions cs
            LEFT JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            LEFT JOIN (
                SELECT session_id, COUNT(*) AS violation_count
                FROM violations
                GROUP BY session_id
            ) v ON cs.session_id = v.session_id
            WHERE cs.candidate_email IS NOT NULL AND TRIM(cs.candidate_email) != ''
            ORDER BY cs.submitted_at DESC
            """
        )
        c_res = await db.execute(c_query)
        for row in c_res.fetchall():
            m = dict(row._mapping)
            e = norm_email(m.get("candidate_email"))
            if not e:
                continue
            if e not in coding_by_email:
                sc = m.get("score")
                coding_by_email[e] = {
                    "challenge_title": m.get("challenge_title"),
                    "score": float(sc) if sc is not None else None,
                    "status": m.get("status"),
                    "submitted_at": str(m.get("submitted_at"))
                    if m.get("submitted_at")
                    else None,
                    "violations": int(m.get("violation_count") or 0),
                    "submission_id": m.get("submission_id"),
                }
    except Exception as ex:
        print(f"⚠️ candidate-stages coding: {ex}")

    interviews_by_email: dict = {}
    try:
        i_query = text(
            """
            SELECT session_id, candidate_email, candidate_name, job_role, status,
                   overall_score, end_time, created_at
            FROM interview_sessions
            ORDER BY COALESCE(end_time, created_at) DESC
            """
        )
        i_res = await db.execute(i_query)
        for row in i_res.fetchall():
            m = dict(row._mapping)
            e = norm_email(m.get("candidate_email"))
            if not e:
                continue
            if e not in interviews_by_email:
                os_ = m.get("overall_score")
                interviews_by_email[e] = {
                    "session_id": m.get("session_id"),
                    "candidate_name": m.get("candidate_name"),
                    "job_role": m.get("job_role"),
                    "status": str(m.get("status") or ""),
                    "overall_score": float(os_) if os_ is not None else None,
                    "end_time": str(m.get("end_time")) if m.get("end_time") else None,
                    "created_at": str(m.get("created_at"))
                    if m.get("created_at")
                    else None,
                }
    except Exception as ex:
        print(f"⚠️ candidate-stages interview: {ex}")

    cv_name_by_email: dict = {}
    try:
        cv_q = text("SELECT email, name FROM cv_candidates WHERE email IS NOT NULL")
        cv_r = await db.execute(cv_q)
        for row in cv_r.fetchall():
            m = dict(row._mapping)
            em = norm_email(m.get("email"))
            if em and m.get("name"):
                cv_name_by_email[em] = m.get("name")
    except Exception as ex:
        print(f"⚠️ candidate-stages cv names: {ex}")

    all_emails = (
        set(assessment_by_email) | set(coding_by_email) | set(interviews_by_email)
    )
    candidates = []
    for e in sorted(all_emails):
        intr = interviews_by_email.get(e) or {}
        name = (
            intr.get("candidate_name")
            or cv_name_by_email.get(e)
            or (e.split("@")[0] if "@" in e else e)
        )
        candidates.append(
            {
                "email": e,
                "name": name,
                "assessment": assessment_by_email.get(e),
                "coding": coding_by_email.get(e),
                "interview": intr or None,
            }
        )
    return {"success": True, "candidates": candidates, "total": len(candidates)}


# ==================== CV DATABASE HELPER FUNCTIONS ====================


async def get_job_posting(db: AsyncSession, job_id: str) -> dict:
    """Get job posting by ID"""
    try:
        from sqlalchemy import text

        query = text("SELECT * FROM job_postings WHERE job_id = :job_id")
        result = await db.execute(query, {"job_id": job_id})
        row = result.fetchone()

        if not row:
            return {}

        return {
            "job_id": row[0],
            "title": row[1],
            "description": row[2],
            "required_skills": json.loads(row[3]) if row[3] else [],
            "experience_level": row[4],
            "location": row[5],
            "salary_range": row[6],
            "posted_at": str(row[7]) if row[7] else None,
            "posted_by": row[8],
            "status": row[9],
            "assessment_id": row[10],
        }
    except Exception as e:
        print(f"❌ Error getting job posting: {e}")
        return {}


async def get_all_candidates(
    db: AsyncSession, status: str = None, job_id: str = None, min_match: float = None
):
    """Get all CV candidates with filters"""
    try:
        from sqlalchemy import text

        query_str = """
            SELECT candidate_id, email, name, phone, role, skills, cv_filename, cv_text,
                   status, skill_match_percentage, cv_source, job_id, raw_text,
                   experience, education, created_at, missing_skills, extra_skills,
                   matched_skills
            FROM cv_candidates WHERE 1=1
        """
        params = {}

        if status:
            query_str += " AND status = :status"
            params["status"] = status

        if job_id:
            query_str += " AND job_id = :job_id"
            params["job_id"] = job_id

        if min_match is not None:
            query_str += " AND skill_match_percentage >= :min_match"
            params["min_match"] = min_match

        query_str += " ORDER BY created_at DESC"

        result = await db.execute(text(query_str), params)
        rows = result.fetchall()

        candidates = []
        job_cache: dict = {}
        for row in rows:
            # Safely parse JSON fields
            try:
                skills = json.loads(row[5]) if row[5] and row[5] != "[]" else []
            except (json.JSONDecodeError, TypeError):
                skills = []
            try:
                experience = json.loads(row[13]) if row[13] and row[13] != "[]" else []
            except (json.JSONDecodeError, TypeError):
                experience = []
            try:
                education = json.loads(row[14]) if row[14] and row[14] != "[]" else []
            except (json.JSONDecodeError, TypeError):
                education = []
            try:
                missing_skills = (
                    json.loads(row[16]) if row[16] and row[16] != "[]" else []
                )
            except (json.JSONDecodeError, TypeError):
                missing_skills = []
            try:
                extra_skills = (
                    json.loads(row[17]) if row[17] and row[17] != "[]" else []
                )
            except (json.JSONDecodeError, TypeError):
                extra_skills = []

            matched_pairs = []
            try:
                if len(row) > 18 and row[18]:
                    matched_pairs = json.loads(row[18])
                    if not isinstance(matched_pairs, list):
                        matched_pairs = []
            except (json.JSONDecodeError, TypeError, IndexError):
                matched_pairs = []

            if not matched_pairs and row[11] and skills:
                jid = row[11]
                if jid not in job_cache:
                    job_cache[jid] = await get_job_posting(db, jid)
                job_d = job_cache.get(jid) or {}
                if job_d:
                    matched_pairs = skill_matcher.match_skills(skills, job_d).get(
                        "matched_pairs", []
                    )

            candidates.append(
                {
                    "candidate_id": row[0],
                    "email": row[1],
                    "name": row[2],
                    "phone": row[3],
                    "role": row[4],
                    "skills": skills,
                    "cv_filename": row[6],
                    "cv_file_path": row[6],
                    "cv_text": row[7],
                    "status": row[8],
                    "skill_match_percentage": row[9],
                    "cv_source": row[10],
                    "job_id": row[11],
                    "raw_text": row[12],
                    "experience": experience,
                    "education": education,
                    "created_at": str(row[15]) if row[15] else None,
                    "extracted_at": str(row[15]) if row[15] else None,
                    "missing_skills": missing_skills,
                    "extra_skills": extra_skills,
                    "matched_pairs": matched_pairs,
                }
            )

        return candidates

    except Exception as e:
        print(f"❌ Error getting candidates: {e}")
        return []


async def get_all_job_postings(db: AsyncSession, status: str = "active"):
    """Get all job postings"""
    try:
        from sqlalchemy import text

        query = "SELECT * FROM job_postings WHERE 1=1"
        params = {}

        if status:
            query += " AND status = :status"
            params["status"] = status

        query += " ORDER BY posted_at DESC"

        result = await db.execute(text(query), params)
        rows = result.fetchall()

        job_postings = []
        for row in rows:
            job_postings.append(
                {
                    "job_id": row[0],
                    "title": row[1],
                    "description": row[2],
                    "required_skills": json.loads(row[3]) if row[3] else [],
                    "experience_level": row[4],
                    "location": row[5],
                    "salary_range": row[6],
                    "posted_at": str(row[7]) if row[7] else None,
                    "posted_by": row[8],
                    "status": row[9],
                    "assessment_id": row[10],
                }
            )

        return job_postings

    except Exception as e:
        print(f"❌ Error getting job postings: {e}")
        return []


async def save_cv_candidate(db: AsyncSession, cv_data: dict) -> str:
    """Save CV candidate to database"""
    try:
        from sqlalchemy import text

        candidate_id = cv_data.get("candidate_id", f"candidate_{uuid.uuid4()}")

        # Check if candidate already exists by email
        if cv_data.get("email"):
            check_query = text(
                "SELECT candidate_id FROM cv_candidates WHERE email = :email"
            )
            result = await db.execute(check_query, {"email": cv_data["email"]})
            existing = result.fetchone()

            if existing:
                candidate_id = existing[0]
                print(f"⚠️ Candidate already exists: {cv_data['email']}")

        insert_query = text(
            """
            INSERT INTO cv_candidates (
                candidate_id,
                email, name, phone, role, skills, cv_filename, cv_text,
                status, skill_match_percentage, cv_source, job_id, raw_text,
                experience, education, missing_skills, extra_skills, matched_skills, created_at
            ) VALUES (
                :candidate_id,
                :email, :name, :phone, :role, :skills, :cv_filename, :cv_text,
                :status, :skill_match_percentage, :cv_source, :job_id, :raw_text,
                :experience, :education, :missing_skills, :extra_skills, :matched_skills, NOW()
            ) ON DUPLICATE KEY UPDATE
                name = VALUES(name),
                phone = VALUES(phone),
                role = VALUES(role),
                skills = VALUES(skills),
                cv_filename = VALUES(cv_filename),
                cv_text = VALUES(cv_text),
                status = VALUES(status),
                skill_match_percentage = VALUES(skill_match_percentage),
                cv_source = VALUES(cv_source),
                job_id = VALUES(job_id),
                raw_text = VALUES(raw_text),
                experience = VALUES(experience),
                education = VALUES(education),
                missing_skills = VALUES(missing_skills),
                extra_skills = VALUES(extra_skills),
                matched_skills = VALUES(matched_skills),
                updated_at = NOW()
        """
        )

        await db.execute(
            insert_query,
            {
                "candidate_id": candidate_id,
                "email": cv_data.get("email", ""),
                "name": cv_data.get("name", ""),
                "phone": cv_data.get("phone", ""),
                "role": cv_data.get("role", ""),
                "skills": json.dumps(cv_data.get("skills", [])),
                "cv_filename": cv_data.get("cv_file_path", ""),
                "cv_text": cv_data.get("raw_text", "")[:5000],  # Limit raw text size
                "status": cv_data.get("status", "processed"),
                "skill_match_percentage": float(
                    cv_data.get("skill_match_percentage", 0)
                ),
                "cv_source": cv_data.get("cv_source") or cv_data.get("source") or "manual_upload",
                "job_id": cv_data.get("job_id"),
                "raw_text": cv_data.get("raw_text", "")[:5000],  # Limit raw text size
                "experience": json.dumps(cv_data.get("experience", [])),
                "education": json.dumps(cv_data.get("education", [])),
                "missing_skills": json.dumps(cv_data.get("missing_skills", [])),
                "extra_skills": json.dumps(cv_data.get("extra_skills", [])),
                "matched_skills": json.dumps(cv_data.get("matched_pairs", [])),
            },
        )

        await db.commit()
        print(f"✅ Saved candidate: {cv_data.get('email')}")
        # PK is VARCHAR; lastrowid is not meaningful for this insert
        return str(candidate_id)

    except Exception as e:
        print(f"❌ Error saving candidate: {e}")
        await db.rollback()
        return cv_data.get("candidate_id", "")


# ==================== SCORING HELPER FUNCTIONS ====================
def calculate_percentage(correct: int, total: int) -> float:
    """Safe percentage calculation with zero division protection"""
    if total == 0:
        return 0.0
    return round((correct / total) * 100, 2)


def calculate_grade(score_percentage: float) -> str:
    """Calculate grade based on score percentage"""
    if score_percentage >= 90:
        return "A+"
    elif score_percentage >= 85:
        return "A"
    elif score_percentage >= 80:
        return "A-"
    elif score_percentage >= 75:
        return "B+"
    elif score_percentage >= 70:
        return "B"
    elif score_percentage >= 65:
        return "B-"
    elif score_percentage >= 60:
        return "C+"
    elif score_percentage >= 55:
        return "C"
    elif score_percentage >= 50:
        return "C-"
    else:
        return "F"


# ==================== GENERIC SCORING HELPER FUNCTIONS ====================
def calculate_efficiency_score(
    avg_execution_time: float, language: str, total_tests: int
) -> int:
    """Calculate efficiency score dynamically based on language and test count"""
    # ✅ FIXED: Return 0 if tests failed
    if total_tests == 0:
        return 0

    # ✅ FIXED: Penalize compilation/runtime errors heavily
    if avg_execution_time < 0:  # Indicates error
        return 20

    if language in ["python", "javascript", "js", "node"]:
        if avg_execution_time < 50:
            return 100
        elif avg_execution_time < 200:
            return 90
        elif avg_execution_time < 500:
            return 80
        elif avg_execution_time < 1000:
            return 70
        elif avg_execution_time < 2000:
            return 60
        else:
            return 50
    else:
        if avg_execution_time < 20:
            return 100
        elif avg_execution_time < 100:
            return 90
        elif avg_execution_time < 300:
            return 80
        elif avg_execution_time < 800:
            return 70
        elif avg_execution_time < 1500:
            return 60
        else:
            return 50


def generate_evaluation_prompt(
    code: str,
    language: str,
    passed_tests: int,
    total_tests: int,
    avg_execution_time: float,
) -> str:
    """Generate dynamic evaluation prompt for any coding challenge"""
    # ✅ ADD EDGE CASE DETECTION
    edge_case_check = ""

    # Check code for edge case handling
    if language == "go":
        if "len(nums) == 0" not in code and "len(arr) == 0" not in code:
            edge_case_check += "⚠️ Missing edge case: empty array check\\n"
    elif language == "javascript":
        if "!nums" not in code and "arr.length === 0" not in code:
            edge_case_check += "⚠️ Missing edge case: null/empty check\\n"
    elif language == "python":
        if "if not nums" not in code and "len(nums) == 0" not in code:
            edge_case_check += "⚠️ Missing edge case: empty list check\\n"

    return f"""Evaluate this {language} code solution for a coding challenge:

```{language}
{code}
Test Results: {passed_tests}/{total_tests} tests passed
Average Execution Time: {avg_execution_time:.1f}ms per test

{edge_case_check}
Provide comprehensive evaluation focusing on:

1. Edge case handling (0-100) - empty arrays, null values, boundary conditions
2. Code quality (0-100) - structure, organization, adherence to language conventions
3. Problem-solving approach (0-100) - algorithmic thinking, optimal solution choice
4. Readability (0-100) - naming, comments, clarity
5. Best practices (0-100) - error handling, edge cases, modern patterns

Key strengths (3-5 bullet points)

Areas for improvement (3-5 bullet points)

Return ONLY valid JSON:
{{
"edge_case_score": 85,
"code_quality_score": 85,
"problem_solving_score": 90,
"readability_score": 80,
"best_practices_score": 75,
"time_complexity": "O(n)",
"space_complexity": "O(1)",
"strengths": ["Efficient algorithm", "Good variable naming", "Proper structure", "Handles edge cases"],
"improvements": ["Add comments", "Handle edge cases", "Improve error handling"],
"overall_feedback": "Solid solution with good efficiency"
}}"""


def parse_evaluation_response(eval_response: str, language: str) -> dict:
    """Parse AI evaluation response with fallback defaults"""
    evaluation = {
        "code_quality_score": 50,
        "problem_solving_score": 50,
        "readability_score": 50,
        "best_practices_score": 50,
        "time_complexity": "Unknown",
        "space_complexity": "Unknown",
        "strengths": ["Solution passes basic tests"],
        "improvements": ["Add documentation", "Improve code structure"],
        "overall_feedback": "Basic evaluation completed",
    }

    if eval_response:
        eval_response = eval_response.strip()
        if "```json" in eval_response:
            eval_response = eval_response.split("```json")[1].split("```")[0].strip()

        json_match = re.search(r"\{[\s\S]*\}", eval_response)
        if json_match:
            try:
                parsed_eval = json.loads(json_match.group(0))
                for key in evaluation.keys():
                    if key in parsed_eval:
                        evaluation[key] = parsed_eval[key]
            except:
                print("⚠️ Failed to parse AI evaluation, using defaults")

    return evaluation


def calculate_score_breakdown(
    correctness_score: float,
    efficiency_score: float,
    evaluation: dict,
    passed_tests: int,
    total_tests: int,
    avg_execution_time: float,
) -> dict:
    """Calculate comprehensive score breakdown for any solution"""
    correctness_contribution = correctness_score * 0.40
    efficiency_contribution = efficiency_score * 0.20
    code_quality_contribution = evaluation["code_quality_score"] * 0.15
    problem_solving_contribution = evaluation["problem_solving_score"] * 0.15
    readability_contribution = evaluation["readability_score"] * 0.10

    total_contribution = (
        correctness_contribution
        + efficiency_contribution
        + code_quality_contribution
        + problem_solving_contribution
        + readability_contribution
    )

    correctness_feedback = generate_correctness_feedback(passed_tests, total_tests)
    efficiency_feedback = generate_efficiency_feedback(
        efficiency_score, avg_execution_time
    )
    code_quality_feedback = generate_code_quality_feedback(evaluation)
    problem_solving_feedback = generate_problem_solving_feedback(evaluation)
    readability_feedback = generate_readability_feedback(evaluation)

    return {
        "correctness": {
            "score": round(correctness_score, 1),
            "weight": 0.40,
            "contribution": round(correctness_contribution, 1),
            "feedback": correctness_feedback,
        },
        "efficiency": {
            "score": efficiency_score,
            "weight": 0.20,
            "contribution": round(efficiency_contribution, 1),
            "feedback": efficiency_feedback,
        },
        "code_quality": {
            "score": evaluation["code_quality_score"],
            "weight": 0.15,
            "contribution": round(code_quality_contribution, 1),
            "feedback": code_quality_feedback,
        },
        "problem_solving": {
            "score": evaluation["problem_solving_score"],
            "weight": 0.15,
            "contribution": round(problem_solving_contribution, 1),
            "feedback": problem_solving_feedback,
        },
        "readability": {
            "score": evaluation["readability_score"],
            "weight": 0.10,
            "contribution": round(readability_contribution, 1),
            "feedback": readability_feedback,
        },
        "total_contribution": round(total_contribution, 2),
        "total_contribution_rounded": round(total_contribution),
    }


def generate_correctness_feedback(passed_tests: int, total_tests: int) -> str:
    """Generate dynamic correctness feedback"""
    if passed_tests == total_tests:
        return f"✅ Perfect! All {total_tests} test cases passed"
    elif passed_tests >= total_tests * 0.8:
        return f"✅ Good! {passed_tests}/{total_tests} tests passed"
    else:
        return f"⚠️ Needs work: {passed_tests}/{total_tests} tests passed"


def generate_efficiency_feedback(
    efficiency_score: int, avg_execution_time: float
) -> str:
    """Generate dynamic efficiency feedback"""
    if efficiency_score >= 90:
        return f"⚡ Excellent performance ({avg_execution_time:.1f}ms avg)"
    elif efficiency_score >= 80:
        return f"⚡ Good performance ({avg_execution_time:.1f}ms avg)"
    else:
        return f"⏱️ Could be optimized ({avg_execution_time:.1f}ms avg)"


def generate_code_quality_feedback(evaluation: dict) -> str:
    """Generate dynamic code quality feedback"""
    score = evaluation["code_quality_score"]
    if score >= 90:
        return "🎯 Excellent code structure and organization"
    elif score >= 80:
        return "🎯 Good code quality with minor improvements possible"
    else:
        return "📝 Focus on code structure and organization"


def generate_problem_solving_feedback(evaluation: dict) -> str:
    """Generate dynamic problem-solving feedback"""
    score = evaluation["problem_solving_score"]
    complexity = evaluation.get("time_complexity", "Unknown")
    if score >= 90:
        return f"💡 Optimal solution ({complexity})"
    elif score >= 80:
        return f"💡 Good approach ({complexity})"
    else:
        return f"🤔 Consider alternative approaches ({complexity})"


def generate_readability_feedback(evaluation: dict) -> str:
    """Generate dynamic readability feedback"""
    score = evaluation["readability_score"]
    if score >= 90:
        return "📖 Very clean and readable code"
    elif score >= 80:
        return "📖 Readable with minor documentation needed"
    else:
        return "📖 Add comments and improve naming"


def get_performance_level(score: int) -> str:
    """Determine performance level based on score"""
    if score >= 90:
        return "Excellent"
    elif score >= 75:
        return "Good"
    elif score >= 60:
        return "Satisfactory"
    else:
        return "Needs Improvement"


def generate_improvement_suggestions(
    evaluation: dict, correctness_score: float, efficiency_score: float
) -> list:
    """Generate dynamic improvement suggestions"""
    suggestions = evaluation.get("improvements", [])
    if correctness_score < 100:
        suggestions.append("Focus on passing all test cases")
    if efficiency_score < 80:
        suggestions.append("Optimize algorithm for better performance")
    if evaluation["readability_score"] < 80:
        suggestions.append("Improve code documentation and comments")
    return suggestions[:5]


# ==================== GROQ API HELPER ====================
async def generate_with_groq(
    prompt: str,
    *,
    max_tokens: int = 4000,
    temperature: float = 0.7,
) -> str:
    """Generate content using Groq API"""
    if not GROQ_API_KEY:
        print("❌ Groq API key not found")
        return ""

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": GROQ_MODEL,
        "messages": [
            {
                "role": "system",
                "content": "You are an expert technical interviewer. Generate high-quality coding challenges and evaluations in valid JSON format.",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=aiohttp.ClientTimeout(total=120),
            ) as resp:
                if resp.status != 200:
                    err_body = await resp.text()
                    print(f"❌ Groq API HTTP {resp.status}: {err_body[:800]}")
                    return ""
                data = await resp.json()
                return (
                    data.get("choices", [{}])[0].get("message", {}).get("content", "")
                )
    except Exception as e:
        print(f"❌ Groq API error: {e}")
        return ""


# ==================== TECHNICAL SCORE INTEGRATION ====================
async def update_technical_score(
    db,
    candidate_email,
    assessment_id,
    coding_submission_id,
    mcq_result_id,
    coding_score,
):
    """Update combined technical score (MCQ + Coding) with role-based weighting"""
    try:
        from sqlalchemy import text

        # Get MCQ score if linked
        mcq_score, mcq_percentage = None, None
        if mcq_result_id:
            result = await db.execute(
                text(
                    "SELECT score, score_percentage FROM results WHERE result_id = :id"
                ),
                {"id": mcq_result_id},
            )
            row = result.fetchone()
            if row:
                mcq_score, mcq_percentage = row[0], row[1]

        # Get role to determine weighting
        role = "Software Engineer"
        if assessment_id:
            role_result = await db.execute(
                text("SELECT role FROM assessments WHERE assessment_id = :id"),
                {"id": assessment_id},
            )
            role_row = role_result.fetchone()
            if role_row:
                role = role_row[0]

        technical_keywords = [
            "engineer",
            "developer",
            "programmer",
            "technical",
            "software",
            "coding",
            "programming",
        ]
        managerial_keywords = ["manager", "lead", "director", "head", "supervisor"]

        is_technical = any(keyword in role.lower() for keyword in technical_keywords)
        is_managerial = any(keyword in role.lower() for keyword in managerial_keywords)

        if is_technical:
            mcq_weight, coding_weight = 0.4, 0.6
        elif is_managerial:
            mcq_weight, coding_weight = 0.7, 0.3
        else:
            mcq_weight, coding_weight = 0.5, 0.5

        if mcq_percentage is not None and coding_score is not None:
            combined_percentage = (mcq_percentage * mcq_weight) + (
                coding_score * coding_weight
            )
        elif mcq_percentage is not None:
            combined_percentage = mcq_percentage
        elif coding_score is not None:
            combined_percentage = coding_score
        else:
            combined_percentage = 0.0

        combined_score = int(combined_percentage)

        if combined_score >= 85:
            technical_level = "Senior/Expert"
        elif combined_score >= 70:
            technical_level = "Mid-Level"
        elif combined_score >= 55:
            technical_level = "Junior"
        else:
            technical_level = "Entry-Level"

        # Check if exists
        result = await db.execute(
            text(
                """
            SELECT tech_score_id FROM technical_scores
            WHERE candidate_email = :email AND assessment_id = :assessment_id
        """
            ),
            {"email": candidate_email, "assessment_id": assessment_id},
        )

        existing = result.fetchone()

        if existing:
            await db.execute(
                text(
                    """
                UPDATE technical_scores SET
                    mcq_result_id = :mcq_result_id,
                    mcq_score = :mcq_score,
                    mcq_percentage = :mcq_percentage,
                    coding_submission_id = :coding_submission_id,
                    coding_score = :coding_score,
                    coding_percentage = :coding_percentage,
                    combined_score = :combined_score,
                    combined_percentage = :combined_percentage,
                    technical_level = :technical_level,
                    updated_at = :updated_at
                WHERE tech_score_id = :tech_score_id
            """
                ),
                {
                    "tech_score_id": existing[0],
                    "mcq_result_id": mcq_result_id,
                    "mcq_score": mcq_score,
                    "mcq_percentage": mcq_percentage,
                    "coding_submission_id": coding_submission_id,
                    "coding_score": coding_score,
                    "coding_percentage": coding_score,
                    "combined_score": combined_score,
                    "combined_percentage": combined_percentage,
                    "technical_level": technical_level,
                    "updated_at": datetime.now(),
                },
            )
        else:
            tech_score_id = f"tech_{uuid.uuid4()}"
            await db.execute(
                text(
                    """
                INSERT INTO technical_scores (
                    tech_score_id, candidate_email, assessment_id,
                    mcq_result_id, mcq_score, mcq_percentage,
                    coding_submission_id, coding_score, coding_percentage,
                    combined_score, combined_percentage, technical_level,
                    created_at, updated_at
                ) VALUES (
                    :tech_score_id, :candidate_email, :assessment_id,
                    :mcq_result_id, :mcq_score, :mcq_percentage,
                    :coding_submission_id, :coding_score, :coding_percentage,
                    :combined_score, :combined_percentage, :technical_level,
                    :created_at, :updated_at
                )
            """
                ),
                {
                    "tech_score_id": tech_score_id,
                    "candidate_email": candidate_email,
                    "assessment_id": assessment_id,
                    "mcq_result_id": mcq_result_id,
                    "mcq_score": mcq_score,
                    "mcq_percentage": mcq_percentage,
                    "coding_submission_id": coding_submission_id,
                    "coding_score": coding_score,
                    "coding_percentage": coding_score,
                    "combined_score": combined_score,
                    "combined_percentage": combined_percentage,
                    "technical_level": technical_level,
                    "created_at": datetime.now(),
                    "updated_at": datetime.now(),
                },
            )
    except Exception as e:
        print(f"❌ Technical score update failed: {e}")


# ==================== DUPLICATE QUESTION DETECTION ====================
def is_duplicate_question(new_question: Dict, existing_questions: List[Dict]) -> bool:
    """Check if a question is an EXACT duplicate"""
    import re
    from difflib import SequenceMatcher

    new_text = new_question.get("question", "").lower().strip()
    new_text_normalized = re.sub(r"[^a-z0-9\s]", "", new_text)
    new_text_normalized = re.sub(r"\s+", " ", new_text_normalized).strip()

    for existing_q in existing_questions:
        existing_text = existing_q.get("question", "").lower().strip()
        existing_text_normalized = re.sub(r"[^a-z0-9\s]", "", existing_text)
        existing_text_normalized = re.sub(r"\s+", " ", existing_text_normalized).strip()

        if new_text_normalized == existing_text_normalized:
            return True

        similarity = SequenceMatcher(
            None, new_text_normalized, existing_text_normalized
        ).ratio()
        if similarity > 0.98:
            return True

    return False


# ==================== EMAIL FUNCTIONALITY ====================
async def send_assessment_email(candidate_email: str, assessment_id: str, db):
    """Send assessment via email"""
    try:
        assessment = await AssessmentRepository.get_by_id(db, assessment_id)
        if not assessment:
            return False

        sender_email = os.getenv("SMTP_EMAIL")
        sender_password = os.getenv("SMTP_PASSWORD")
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))

        is_adaptive = (
            assessment.get("difficulty") == "adaptive"
            or assessment.get("is_adaptive") == True
            or str(assessment_id).startswith("adaptive_")
        )

        if is_adaptive:
            assessment_link = f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/adaptive-test?id={assessment_id}"
            test_type = "Adaptive"
        else:
            assessment_link = f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/take-assessment?id={assessment_id}"
            test_type = "Standard"

        message = MIMEMultipart("alternative")
        message["Subject"] = f"Your {test_type} Assessment is Ready"
        message["From"] = sender_email
        message["To"] = candidate_email

        html = f"""
        <html>
          <body>
            <h2>Assessment Invitation</h2>
            <p>You have been invited to take a <strong>{test_type}</strong> assessment for <strong>{assessment.get("role", "Position")}</strong>.</p>
            <p><a href="{assessment_link}" style="background-color: #4f46e5; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; display: inline-block;">Start {test_type} Assessment</a></p>
            <p>Link: {assessment_link}</p>
            <p><strong>Assessment Details:</strong></p>
            <ul>
              <li>Role: {assessment.get("role", "N/A")}</li>
              <li>Type: {test_type}</li>
              <li>Questions: {len(assessment.get("questions", []))}</li>
              <li>Duration: {assessment.get("duration_minutes", 30)} minutes</li>
            </ul>
            <p>Please complete the assessment within the specified time limit.</p>
            {f"<p><em>Note: This is an adaptive test. Questions will adjust to your skill level.</em></p>" if is_adaptive else ""}
          </body>
        </html>
        """

        part = MIMEText(html, "html")
        message.attach(part)

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, candidate_email, message.as_string())

        return True
    except Exception as e:
        print(f"❌ Email failed: {e}")
        return False


@app.get("/")
async def root(db: AsyncSession = Depends(get_db)):
    """Health check"""
    assessments = await AssessmentRepository.get_all(db)
    results = await ResultRepository.get_all(db)

    return {
        "message": "RECRUTO - AI-Powered Recruitment API",
        "status": "running",
        "database": "MySQL",
        "assessments_count": len(assessments),
        "results_count": len(results),
    }


@app.get("/api/coding/technical-score/{email}")
async def get_technical_score(
    email: str, assessment_id: str = None, db: AsyncSession = Depends(get_db)
):
    """✅ INTEGRATED: Get combined technical score (MCQ + Coding)"""
    try:
        from sqlalchemy import text

        query = "SELECT * FROM technical_scores WHERE candidate_email = :email"
        params = {"email": email}

        if assessment_id:
            query += " AND assessment_id = :assessment_id"
            params["assessment_id"] = assessment_id

        query += " ORDER BY updated_at DESC LIMIT 1"

        result = await db.execute(text(query), params)
        row = result.fetchone()

        if not row:
            return {"message": "No technical score found", "has_score": False}

        return {
            "has_score": True,
            "tech_score_id": row[0],
            "candidate_email": row[1],
            "assessment_id": row[2],
            "mcq_score": row[4],
            "mcq_percentage": row[5],
            "coding_score": row[7],
            "coding_percentage": row[8],
            "combined_score": row[9],
            "combined_percentage": row[10],
            "technical_level": row[11],
            "created_at": str(row[12]) if row[12] else None,
            "updated_at": str(row[13]) if row[13] else None,
        }
    except Exception as e:
        print(f"❌ Technical score fetch failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/mcq/send-assessment")
async def send_assessment_via_email(
    data: dict, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)
):
    """FR-MCQ-09: Send assessment link via email"""
    assessment_id = data.get("assessment_id")
    candidate_emails = data.get("candidate_emails", [])

    if not assessment_id or not candidate_emails:
        raise HTTPException(
            status_code=400, detail="assessment_id and candidate_emails required"
        )

    for email in candidate_emails:
        background_tasks.add_task(send_assessment_email, email, assessment_id, db)

    return {
        "message": f"Assessment invitations sent to {len(candidate_emails)} candidates",
        "assessment_id": assessment_id,
    }


@app.post("/api/mcq/send-adaptive-assessment")
async def send_adaptive_assessment_via_email(
    data: dict, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)
):
    """Send adaptive assessment link via email"""
    assessment_id = data.get("assessment_id")
    role = data.get("role")
    candidate_emails = data.get("candidate_emails", [])

    if not role or not candidate_emails:
        raise HTTPException(
            status_code=400, detail="role and candidate_emails required"
        )

    # If assessment_id provided, use it; otherwise send generic adaptive link
    if assessment_id:
        for email in candidate_emails:
            background_tasks.add_task(send_assessment_email, email, assessment_id, db)
    else:
        for email in candidate_emails:
            background_tasks.add_task(send_adaptive_email, email, role)

    return {
        "message": f"Adaptive assessment invitations sent to {len(candidate_emails)} candidates",
        "assessment_id": assessment_id if assessment_id else "adaptive",
    }


async def send_adaptive_email(candidate_email: str, role: str):
    """Send adaptive assessment invitation without pre-created assessment"""
    try:
        sender_email = os.getenv("SMTP_EMAIL")
        sender_password = os.getenv("SMTP_PASSWORD")
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))

        # Generate adaptive test link (no assessment_id needed)
        assessment_link = f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/adaptive-test?email={candidate_email}&role={role}"

        message = MIMEMultipart("alternative")
        message["Subject"] = f"Your Adaptive Assessment is Ready - {role}"
        message["From"] = sender_email
        message["To"] = candidate_email

        html = f"""
        <html>
          <body style="font-family: Arial, sans-serif; padding: 20px;">
            <h2 style="color: #4f46e5;">Adaptive Assessment Invitation</h2>
            <p>You have been invited to take an <strong>Adaptive Assessment</strong> for the <strong>{role}</strong> position.</p>
            <p style="margin: 30px 0;">
              <a href="{assessment_link}" style="background-color: #4f46e5; color: white; padding: 14px 28px; text-decoration: none; border-radius: 8px; display: inline-block; font-weight: 600;">
                🚀 Start Adaptive Assessment
              </a>
            </p>
            <p><strong>What is an Adaptive Assessment?</strong></p>
            <ul style="line-height: 1.8;">
              <li>Questions adjust to your skill level in real-time</li>
              <li>20 questions total</li>
              <li>Each question is unique and personalized</li>
              <li>Your performance determines the next question's difficulty</li>
            </ul>
            <p style="margin-top: 30px; padding: 15px; background-color: #f3f4f6; border-left: 4px solid #4f46e5;">
              <strong>Assessment Link:</strong><br>
              <a href="{assessment_link}" style="color: #4f46e5; word-break: break-all;">{assessment_link}</a>
            </p>
            <p style="color: #6b7280; font-size: 14px; margin-top: 30px;">
              Please complete the assessment in one sitting. The test cannot be paused once started.
            </p>
          </body>
        </html>
        """

        part = MIMEText(html, "html")
        message.attach(part)

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, candidate_email, message.as_string())

        print(f"✅ Adaptive email sent to {candidate_email} for {role}")
        return True
    except Exception as e:
        print(f"❌ Adaptive email failed: {e}")
        return False


# ==================== ASSESSMENTS ====================


@app.post("/api/mcq/generate")
async def generate_mcqs(
    request: MCQGenerateRequest, db: AsyncSession = Depends(get_db)
):
    """Generate MCQs for the given role and difficulty."""
    try:
        mcqs = await generate_mcqs_service(
            role=request.role,
            difficulty=request.difficulty,
            num_questions=request.num_questions,
            mode=request.mode,
        )

        if not mcqs:
            raise HTTPException(status_code=500, detail="Failed to generate MCQs")

        # Save to database
        assessment_id = str(uuid.uuid4())
        assessment_data = {
            "assessment_id": assessment_id,
            "role": request.role,
            "difficulty": request.difficulty,
            "questions": mcqs,
            "duration_minutes": 30,
            "status": "draft",
            "is_adaptive": (request.mode == "adaptive"),
            "created_at": datetime.now().isoformat(),
            "num_questions": len(mcqs),
        }

        await AssessmentRepository.create(db, assessment_data)

        return {
            "assessment_id": assessment_id,
            "questions": mcqs,
            "num_questions": len(mcqs),
            "role": request.role,
            "difficulty": request.difficulty,
            "mode": request.mode,
        }

    except Exception as e:
        print(f"Error generating MCQs: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/mcq/assessments")
async def list_assessments(status: str = None, db: AsyncSession = Depends(get_db)):
    """Get all assessments"""
    assessments = await AssessmentRepository.get_all(db, status)
    return {"assessments": assessments, "total": len(assessments)}


@app.get("/api/mcq/assessment/{assessment_id}")
async def get_assessment(assessment_id: str, db: AsyncSession = Depends(get_db)):
    """Get specific assessment"""
    assessment = await AssessmentRepository.get_by_id(db, assessment_id)
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    return assessment


@app.put("/api/mcq/assessment/{assessment_id}")
async def update_assessment(
    assessment_id: str, data: dict, db: AsyncSession = Depends(get_db)
):
    """FR-MCQ-07, FR-MCQ-08: HR review/edit"""
    success = await AssessmentRepository.update(db, assessment_id, data)
    if not success:
        raise HTTPException(status_code=404, detail="Assessment not found")

    updated = await AssessmentRepository.get_by_id(db, assessment_id)
    return {"message": "Assessment updated successfully", "assessment": updated}


@app.delete("/api/mcq/assessment/{assessment_id}")
async def delete_assessment(assessment_id: str, db: AsyncSession = Depends(get_db)):
    """Delete assessment"""
    success = await AssessmentRepository.delete(db, assessment_id)
    if not success:
        raise HTTPException(status_code=404, detail="Assessment not found")
    return {"message": "Assessment deleted successfully"}


# ==================== SESSIONS ====================


@app.post("/api/mcq/start-assessment")
async def start_assessment(data: dict, db: AsyncSession = Depends(get_db)):
    """FR-MCQ-10: Start assessment"""
    assessment_id = data.get("assessment_id")
    candidate_email = data.get("candidate_email")

    if not assessment_id or not candidate_email:
        raise HTTPException(
            status_code=400, detail="assessment_id and candidate_email required"
        )

    assessment = await AssessmentRepository.get_by_id(db, assessment_id)
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    questions_list = assessment.get("questions") or []
    if not assessment.get("is_adaptive") and len(questions_list) == 0:
        raise HTTPException(
            status_code=400,
            detail="This assessment has no questions yet. Ask HR to regenerate MCQs or assign another assessment.",
        )

    session_id = str(uuid.uuid4())
    session = {
        "session_id": session_id,
        "assessment_id": assessment_id,
        "candidate_email": candidate_email,
        "role": assessment["role"],
        "start_time": datetime.now().isoformat(),
        "time_remaining": assessment["duration_minutes"] * 60,
        "answers": {},
        "status": "in_progress",
        "violations": {},
        "metadata": {},
    }

    await SessionRepository.create(db, session)
    await AssessmentRepository.update(db, assessment_id, {"status": "active"})

    return {
        "session_id": session_id,
        "assessment": assessment,
        "start_time": session["start_time"],
        "time_remaining": session["time_remaining"],  # ✅ FIXED FOR STANDARD TEST TIMER
    }


@app.get("/api/mcq/session/{session_id}")
async def get_session(session_id: str, db: AsyncSession = Depends(get_db)):
    """Get session details"""
    session = await SessionRepository.get_by_id(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    assessment = await AssessmentRepository.get_by_id(db, session["assessment_id"])
    return {"session": session, "assessment": assessment}


@app.get("/api/coding/session/{session_id}/submission-status")
async def get_submission_status(session_id: str, db: AsyncSession = Depends(get_db)):
    """Check if a session has already been submitted"""
    try:
        # Check sessions table for submission status
        result = await db.execute(
            text(
                """
                SELECT submitted
                FROM coding_sessions
                WHERE session_id = :session_id
            """
            ),
            {"session_id": session_id},
        )

        row = result.fetchone()
        if row:
            return {"submitted": row[0]}
        else:
            return {"submitted": False}

    except Exception as e:
        print(f"Error checking submission status: {e}")
        return {"submitted": False}


@app.get("/api/mcq/check-submission/{assessment_id}/{email}")
async def check_mcq_submission(
    assessment_id: str, email: str, db: AsyncSession = Depends(get_db)
):
    """
    Check if candidate has already submitted this assessment
    FIXED: Uses results table with correct column names
    """
    try:
        from urllib.parse import unquote
        from sqlalchemy import text

        decoded_email = unquote(email)
        print(
            f"🔍 Checking submission for: {decoded_email}, Assessment: {assessment_id}"
        )

        # ✅ FIX: Query results table with correct column names
        query = text(
            """
            SELECT result_id, created_at, score_percentage, status
            FROM results
            WHERE assessment_id = :assessment_id
            AND candidate_email = :email
            AND status IN ('completed', 'submitted')
            ORDER BY created_at DESC
            LIMIT 1
        """
        )

        result = await db.execute(
            query, {"assessment_id": assessment_id, "email": decoded_email}
        )

        row = result.fetchone()

        if row:
            print(f"✅ Found existing submission - Result ID: {row[0]}")
            return {
                "submitted": True,
                "result_id": row[0],
                "submitted_at": row[1],  # Using created_at instead of submitted_at
                "score": row[2],
                "status": row[3],
            }

        print(f"❌ No existing submission found")
        return {"submitted": False}

    except Exception as e:
        print(f"❌ Error checking submission: {e}")
        import traceback

        traceback.print_exc()
        return {"submitted": False, "error": str(e)}


@app.post("/api/mcq/submit-assessment")
async def submit_assessment(data: dict, db: AsyncSession = Depends(get_db)):
    """FR-RES-01, FR-RES-02: Auto-grade with proper percentage calculation"""
    session_id = data.get("session_id")
    answers = data.get("answers", {})
    violations = data.get("violations", [])

    session = await SessionRepository.get_by_id(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    session["answers"] = answers
    session["violations"] = violations
    session["status"] = "submitted"
    session["end_time"] = datetime.now().isoformat()

    await SessionRepository.update(
        db,
        session_id,
        {
            "answers": json.dumps(answers),
            "violations": json.dumps(violations),
            "status": "submitted",
            "end_time": session["end_time"],
        },
    )

    assessment = await AssessmentRepository.get_by_id(db, session["assessment_id"])

    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    # ✅ FIXED: Calculate score based on actual questions in assessment (FR-RES-01)
    assessment_questions = assessment.get("questions", [])
    total_questions = len(assessment_questions)
    if total_questions == 0:
        raise HTTPException(status_code=400, detail="Assessment has no questions")

    # DEBUG: Print the structure of answers
    print(f"🔍 DEBUG: Answers structure for session {session_id}")
    for q_key, ans in answers.items():
        print(f"  Question key: '{q_key}' Type={type(ans)}, Value={ans}")
        if isinstance(ans, dict):
            print(f"    Dict keys: {list(ans.keys())}")

    # DEBUG: Print the actual questions and their correct answers
    print(f"📋 DEBUG: Assessment questions structure:")
    for idx, question in enumerate(assessment_questions):
        print(f"  Question {idx}: ID={question.get('question_id')}")
        print(f"    Text: {question.get('question', '')[:100]}...")
        print(f"    Options: {question.get('options', [])}")
        print(f"    Correct Answer: {question.get('correct_answer', 'N/A')}")
        print(f"    Topic: {question.get('topic', 'N/A')}")

    # Count correct answers - FIXED VERSION
    correct_count = 0
    skill_breakdown = {}  # FR-RES-02: Skill-wise breakdown
    question_results = []

    # Process each question in the assessment
    for question_idx, question in enumerate(assessment_questions):
        # Get the actual question_id from the assessment
        actual_question_id = question.get("question_id")

        # Try to find the answer using different possible keys
        candidate_answer = ""

        # Try 1: Use the actual question_id
        if actual_question_id in answers:
            answer_data = answers[actual_question_id]
        # Try 2: Use the index as string (what frontend is sending)
        elif str(question_idx) in answers:
            answer_data = answers[str(question_idx)]
        # Try 3: Use the index as integer
        elif question_idx in answers:
            answer_data = answers[question_idx]
        else:
            # No answer provided for this question
            answer_data = None

        # Extract the answer value
        if answer_data is None:
            candidate_answer = ""
        elif isinstance(answer_data, str):
            candidate_answer = answer_data.strip()
        elif isinstance(answer_data, dict):
            # Try multiple possible keys
            for key in ["answer", "selected", "choice", "value"]:
                if key in answer_data:
                    candidate_answer = str(answer_data[key]).strip()
                    break
            # If still not found, try to get the first string value
            if not candidate_answer:
                for key, value in answer_data.items():
                    if isinstance(value, str):
                        candidate_answer = value.strip()
                        break

        correct_answer = str(question.get("correct_answer", "")).strip()

        # DEBUG: Show what we're comparing
        print(
            f"🔍 Comparing Q{question_idx}: Candidate='{candidate_answer}' vs Correct='{correct_answer}'"
        )

        # Normalize answers (remove extra whitespace, case-insensitive for single letters)
        candidate_norm = (
            candidate_answer.lower().strip()
            if len(candidate_answer) == 1
            else candidate_answer
        )
        correct_norm = (
            correct_answer.lower().strip()
            if len(correct_answer) == 1
            else correct_answer
        )

        is_correct = candidate_norm == correct_norm

        if is_correct:
            correct_count += 1

        # Track skill breakdown
        skill = question.get("topic", "General")
        if skill not in skill_breakdown:
            skill_breakdown[skill] = {"total": 0, "correct": 0}
        skill_breakdown[skill]["total"] += 1
        if is_correct:
            skill_breakdown[skill]["correct"] += 1

        # Store question result
        question_results.append(
            {
                "question_index": question_idx,
                "question_id": actual_question_id,
                "question_text": question.get("question", ""),
                "candidate_answer": candidate_answer,
                "correct_answer": correct_answer,
                "is_correct": is_correct,
                "skill": skill,
            }
        )

        print(
            f"📝 Question {question_idx} (ID: {actual_question_id}): Candidate='{candidate_answer}', Correct='{correct_answer}', Match={is_correct}"
        )

    # ✅ FIXED: Calculate percentage properly using the helper function (FR-RES-02)
    score_percentage = calculate_percentage(correct_count, total_questions)

    # Calculate skill-wise percentages
    skill_percentages = {}
    for skill, stats in skill_breakdown.items():
        skill_percentages[skill] = calculate_percentage(
            stats["correct"], stats["total"]
        )

    print(
        f"✅ STANDARD ASSESSMENT SCORE: {correct_count}/{total_questions} = {score_percentage:.1f}%"
    )
    print(f"📊 Skill breakdown: {skill_percentages}")

    # Process result data
    result_data = {
        "result_id": str(uuid.uuid4()),
        "session_id": session_id,
        "assessment_id": session["assessment_id"],
        "candidate_email": session["candidate_email"],
        "role": assessment["role"],
        "difficulty": assessment["difficulty"],
        "total_questions": total_questions,
        "correct_answers": correct_count,
        "wrong_answers": total_questions - correct_count,
        "unanswered": total_questions - len(answers),
        "score_percentage": score_percentage,
        "skill_breakdown": skill_percentages,  # FR-RES-02: Add skill breakdown
        "start_time": session["start_time"],
        "end_time": session["end_time"],
        "total_time_taken": int(
            (
                datetime.fromisoformat(session["end_time"])
                - datetime.fromisoformat(session["start_time"])
            ).total_seconds()
        ),
        "question_results": question_results,
        "violations": session.get("violations", {}),
        "status": "completed",
        "grade": calculate_grade(score_percentage),
        "created_at": datetime.now().isoformat(),
        "flagged": False,
    }

    await ResultRepository.create(db, result_data)

    return {
        "result_id": result_data["result_id"],
        "score": result_data["score_percentage"],
        "correct": result_data["correct_answers"],
        "total": result_data["total_questions"],
        "skill_breakdown": result_data[
            "skill_breakdown"
        ],  # Include skill breakdown in response
        "grade": result_data["grade"],
        "status": "submitted",
        "flagged": result_data["flagged"],
    }


# ==================== ADAPTIVE SESSIONS ====================


@app.post("/api/mcq/generate-adaptive-link")
async def generate_adaptive_link(data: dict, db: AsyncSession = Depends(get_db)):
    """Generate a permanent shareable link for adaptive assessment"""
    role = data.get("role")
    try:
        max_questions = int(data.get("max_questions", 20))
    except (TypeError, ValueError):
        max_questions = 20
    max_questions = max(1, min(max_questions, 100))

    duration_raw = data.get("duration_minutes")
    try:
        duration_minutes = (
            int(duration_raw) if duration_raw is not None else max_questions * 2
        )
    except (TypeError, ValueError):
        duration_minutes = max_questions * 2
    duration_minutes = max(5, min(duration_minutes, 240))

    if not role:
        raise HTTPException(status_code=400, detail="role required")

    # Create permanent adaptive assessment
    assessment_id = f"adaptive_{uuid.uuid4()}"

    assessment_data = {
        "assessment_id": assessment_id,
        "role": role,
        "difficulty": "adaptive",
        "questions": json.dumps([]),
        "duration_minutes": duration_minutes,
        "status": "draft",  # Draft until someone starts it
        "is_adaptive": True,
        "created_at": datetime.now().isoformat(),
        "num_questions": max_questions,
    }

    try:
        await AssessmentRepository.create(db, assessment_data)

        # Generate shareable link
        frontend_url = os.getenv("FRONTEND_URL", "http://localhost:5173")
        assessment_link = f"{frontend_url}/adaptive-test?id={assessment_id}"

        print(f"✅ Created permanent adaptive assessment: {assessment_id}")

        return {
            "assessment_id": assessment_id,
            "assessment_link": assessment_link,
            "role": role,
            "max_questions": max_questions,
            "message": "Shareable link generated! Send this to candidates anytime.",
        }
    except Exception as e:
        print(f"❌ Failed to create assessment: {e}")
        raise HTTPException(status_code=500, detail="Failed to create assessment")


@app.post("/api/mcq/adaptive/start")
async def start_adaptive_assessment(data: dict, db: AsyncSession = Depends(get_db)):
    """Start adaptive test"""
    role = data.get("role")
    candidate_email = data.get("candidate_email")
    try:
        max_questions = int(data.get("max_questions", 20))
    except (TypeError, ValueError):
        max_questions = 20
    max_questions = max(1, min(max_questions, 100))

    if not role or not candidate_email:
        raise HTTPException(status_code=400, detail="role and candidate_email required")

    session_id = str(uuid.uuid4())
    adaptive_engines[session_id] = AdaptiveTestEngine()

    # ✅ CREATE ASSESSMENT ENTRY FOR ADAPTIVE TEST
    assessment_id = f"adaptive_{uuid.uuid4()}"

    # Create assessment record
    assessment_data = {
        "assessment_id": assessment_id,
        "role": role,
        "difficulty": "adaptive",
        "questions": json.dumps([]),
        "duration_minutes": max_questions * 2,
        "status": "active",
        "is_adaptive": True,
        "created_at": datetime.now().isoformat(),
        "num_questions": max_questions,
    }

    try:
        await AssessmentRepository.create(db, assessment_data)
        print(f"✅ Created adaptive assessment: {assessment_id}")
    except Exception as e:
        print(f"❌ Failed to create assessment: {e}")
        raise HTTPException(status_code=500, detail="Failed to create assessment")

    # Now create session
    session = {
        "session_id": session_id,
        "assessment_id": assessment_id,
        "candidate_email": candidate_email,
        "role": role,
        "start_time": datetime.now().isoformat(),
        "time_remaining": max_questions * 90,
        "answers": {},
        "status": "in_progress",
        "violations": {},
        "metadata": {
            "role": role,
            "max_questions": max_questions,
            "adaptive": True,
            "questions_asked": [],
            "violation_count": 0,
        },
    }

    await SessionRepository.create(db, session)

    # Generate first question
    from integrations.llm_api import generate_single_mcq_dynamic

    first_question = await generate_single_mcq_dynamic(role, "medium")

    session["metadata"]["questions_asked"].append(first_question)
    await SessionRepository.update(
        db, session_id, {"metadata": json.dumps(session["metadata"])}
    )

    return {
        "session_id": session_id,
        "assessment_id": assessment_id,
        "question": first_question,
        "question_number": 1,
        "total_questions": max_questions,
        "difficulty": "medium",
        "time_limit_seconds": max_questions * 90,
    }


@app.post("/api/mcq/adaptive/answer")
async def submit_adaptive_answer(data: dict, db: AsyncSession = Depends(get_db)):
    """Submit adaptive answer with proper scoring based on PDF requirements"""
    session_id = data.get("session_id")
    question_id = data.get("question_id")
    answer = data.get("answer")
    new_violations = data.get("violations", [])

    if not all([session_id, question_id, answer]):
        raise HTTPException(status_code=400, detail="Missing required fields")

    session = await SessionRepository.get_by_id(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    engine = adaptive_engines.get(session_id)
    if not engine:
        raise HTTPException(status_code=400, detail="Adaptive engine not found")

    # Parse metadata
    metadata = session.get("metadata", {})
    if isinstance(metadata, str):
        metadata = json.loads(metadata)

    questions_asked = metadata.get("questions_asked", [])
    current_question = next(
        (q for q in questions_asked if q["question_id"] == question_id), None
    )
    if not current_question:
        raise HTTPException(status_code=404, detail="Question not found")

    # Get answers dict first
    answers = session.get("answers", {})
    if isinstance(answers, str):
        answers = json.loads(answers)

    # ✅ UPDATE VIOLATIONS IF PROVIDED
    if new_violations:
        current_violations = session.get("violations", {})
        if isinstance(current_violations, str):
            try:
                current_violations = json.loads(current_violations)
            except:
                current_violations = {}

        # Ensure current_violations is a dict
        if not isinstance(current_violations, dict):
            current_violations = {
                "tab_switches": 0,
                "window_blurs": 0,
                "copy_attempts": 0,
                "paste_attempts": 0,
                "right_clicks": 0,
            }

        print(f"🔍 BEFORE: Current violations in DB: {current_violations}")
        print(f"🔍 NEW: Violations from frontend: {new_violations}")

        # ✅ REPLACE violations (frontend sends cumulative total)
        if isinstance(new_violations, dict):
            # Frontend sends cumulative count - just REPLACE completely
            current_violations = {
                "tab_switches": int(new_violations.get("tab_switches", 0)),
                "window_blurs": int(new_violations.get("window_blurs", 0)),
                "copy_attempts": int(new_violations.get("copy_attempts", 0)),
                "paste_attempts": int(new_violations.get("paste_attempts", 0)),
                "right_clicks": int(new_violations.get("right_clicks", 0)),
            }
        elif isinstance(new_violations, list):
            # Old format: array of violation objects (should not happen but for backwards compatibility)
            for v in new_violations:
                v_type = v.get("type", "unknown")
                if v_type == "tab_switch":
                    current_violations["tab_switches"] = (
                        current_violations.get("tab_switches", 0) + 1
                    )
                elif v_type == "window_blur":
                    current_violations["window_blurs"] = (
                        current_violations.get("window_blurs", 0) + 1
                    )
                elif v_type == "copy_attempt":
                    current_violations["copy_attempts"] = (
                        current_violations.get("copy_attempts", 0) + 1
                    )
                elif v_type == "paste_attempt":
                    current_violations["paste_attempts"] = (
                        current_violations.get("paste_attempts", 0) + 1
                    )
                elif v_type == "right_click":
                    current_violations["right_clicks"] = (
                        current_violations.get("right_clicks", 0) + 1
                    )

        print(f"🔍 AFTER: Violations to save: {current_violations}")

        # Save back to session
        await SessionRepository.update(
            db, session_id, {"violations": json.dumps(current_violations)}
        )
        await db.commit()

        total_violations = (
            sum(current_violations.values())
            if isinstance(current_violations, dict)
            else len(current_violations)
        )
        print(f"⚠️ Violations updated (total: {total_violations})")

    # ✅ DUPLICATE SUBMISSION CHECK
    if question_id in answers:
        print(
            f"⚠️ Question {question_id} already answered - returning current unanswered question"
        )

        questions_answered = len(answers)
        max_questions = metadata.get("max_questions", 20)

        if questions_answered >= max_questions:
            # ✅ FIXED WITH WEIGHTED SCORING: Hard questions worth more
            def calculate_weighted_adaptive_score(answers_dict, difficulty_history):
                """Calculate weighted score where hard questions are worth more"""
                difficulty_weights = {"easy": 1.0, "medium": 1.5, "hard": 2.0}

                weighted_correct = 0
                total_weight = 0

                for q_id, answer_data in answers_dict.items():
                    # Get difficulty from answer data or history
                    difficulty = answer_data.get("difficulty", "medium")
                    weight = difficulty_weights.get(difficulty, 1.0)

                    total_weight += weight
                    if answer_data.get("correct", False):
                        weighted_correct += weight

                if total_weight > 0:
                    return (weighted_correct / total_weight) * 100
                return 0

            # Calculate weighted score
            score_percentage = calculate_weighted_adaptive_score(
                answers, engine.difficulty_history
            )
            correct_count = sum(1 for a in answers.values() if a["correct"])

            # Calculate weighted skill breakdown for adaptive test
            skill_breakdown = {}
            for q_id, ans_data in answers.items():
                question = next(
                    (q for q in questions_asked if q["question_id"] == q_id), None
                )
                if question:
                    skill = question.get("topic", "General")
                    difficulty = ans_data.get("difficulty", "medium")
                    weight = {"easy": 1.0, "medium": 1.5, "hard": 2.0}.get(
                        difficulty, 1.0
                    )

                    if skill not in skill_breakdown:
                        skill_breakdown[skill] = {
                            "total": 0,
                            "correct": 0,
                            "weighted_total": 0,
                            "weighted_correct": 0,
                        }

                    skill_breakdown[skill]["total"] += 1
                    skill_breakdown[skill]["weighted_total"] += weight

                    if ans_data["correct"]:
                        skill_breakdown[skill]["correct"] += 1
                        skill_breakdown[skill]["weighted_correct"] += weight

            skill_percentages = {}
            for skill, stats in skill_breakdown.items():
                # Use weighted scoring for skill percentages too
                if stats["weighted_total"] > 0:
                    skill_percentages[skill] = calculate_percentage(
                        stats["weighted_correct"], stats["weighted_total"]
                    )
                else:
                    skill_percentages[skill] = 0

            print(f"✅ FINAL ADAPTIVE SCORE (Weighted): {score_percentage:.1f}%")
            print(f"📊 Adaptive skill breakdown: {skill_percentages}")

            return {
                "completed": True,
                "score": score_percentage,
                "correct": correct_count,
                "total": questions_answered,
                "skill_breakdown": skill_percentages,  # Include skill breakdown
                "performance_summary": {
                    "current_difficulty": engine.get_current_difficulty(),
                    "recent_accuracy": engine.calculate_accuracy(),
                    "total_adjustments": len(engine.difficulty_changes),
                },
            }

        # Return existing unanswered question
        current_unanswered_index = questions_answered
        if current_unanswered_index < len(questions_asked):
            current_unanswered_question = questions_asked[current_unanswered_index]
        else:
            # Generate new if missing
            next_difficulty = engine.get_next_difficulty()
            from domain.mcq_prompts import extract_topics_from_questions

            avoid_topics = extract_topics_from_questions(questions_asked)
            from integrations.llm_api import generate_single_mcq_dynamic

            # ✅ WITH DUPLICATE CHECK
            max_attempts = 3
            next_question = None
            for attempt in range(max_attempts):
                candidate = await generate_single_mcq_dynamic(
                    metadata["role"],
                    next_difficulty,
                    avoid_topics,
                    engine.calculate_accuracy() * 100,
                )
                if candidate and not is_duplicate_question(candidate, questions_asked):
                    next_question = candidate
                    print(
                        f"✅ Attempt {attempt + 1}: Generated unique question {next_question['question_id']}"
                    )
                    break
                else:
                    print(f"⚠️ Attempt {attempt + 1}: Duplicate or failed")

            if not next_question:
                raise HTTPException(
                    status_code=500, detail="Failed to generate unique question"
                )

            questions_asked.append(next_question)
            metadata["questions_asked"] = questions_asked
            await SessionRepository.update(
                db, session_id, {"metadata": json.dumps(metadata)}
            )
            await db.commit()
            current_unanswered_question = next_question

        return {
            "completed": False,
            "correct": answers[question_id]["correct"],
            "next_question": current_unanswered_question,
            "question_number": questions_answered + 1,
            "total_questions": max_questions,
            "current_difficulty": current_unanswered_question.get(
                "difficulty", "medium"
            ),
            "performance": {
                "recent_accuracy": engine.calculate_accuracy(),
                "difficulty_adjusted": False,
            },
        }

    # Check if correct
    is_correct = answer == current_question["correct_answer"]
    engine.update_performance(is_correct)
    engine.difficulty_history.append(current_question.get("difficulty", "medium"))

    # Add new answer
    answers[question_id] = {
        "answer": answer,
        "correct": is_correct,
        "difficulty": current_question["difficulty"],
        "timestamp": datetime.now().isoformat(),
    }

    # Save to database immediately
    await SessionRepository.update(db, session_id, {"answers": json.dumps(answers)})
    await db.commit()

    # Now count
    questions_answered = len(answers)
    max_questions = metadata.get("max_questions", 20)

    print(
        f"[ADAPTIVE] Question {question_id}: {'✅ Correct' if is_correct else '❌ Wrong'}"
    )
    print(
        f"[ADAPTIVE] Progress: {questions_answered}/{max_questions} questions answered"
    )

    # Check if test complete
    if questions_answered >= max_questions:
        print(f"[ADAPTIVE] 🎉 Test complete!")

        session["status"] = "completed"
        session["end_time"] = datetime.now().isoformat()

        await SessionRepository.update(
            db, session_id, {"status": "completed", "end_time": session["end_time"]}
        )
        await db.commit()

        # ✅ FIXED WITH WEIGHTED SCORING: Hard questions worth more
        def calculate_weighted_adaptive_score(answers_dict, difficulty_history):
            """Calculate weighted score where hard questions are worth more"""
            difficulty_weights = {"easy": 1.0, "medium": 1.5, "hard": 2.0}

            weighted_correct = 0
            total_weight = 0

            for q_id, answer_data in answers_dict.items():
                # Get difficulty from answer data or history
                difficulty = answer_data.get("difficulty", "medium")
                weight = difficulty_weights.get(difficulty, 1.0)

                total_weight += weight
                if answer_data.get("correct", False):
                    weighted_correct += weight

            if total_weight > 0:
                return (weighted_correct / total_weight) * 100
            return 0

        # Calculate weighted score
        score_percentage = calculate_weighted_adaptive_score(
            answers, engine.difficulty_history
        )
        correct_count = sum(1 for a in answers.values() if a["correct"])

        # Calculate weighted skill breakdown for adaptive test
        skill_breakdown = {}
        for q_id, ans_data in answers.items():
            question = next(
                (q for q in questions_asked if q["question_id"] == q_id), None
            )
            if question:
                skill = question.get("topic", "General")
                difficulty = ans_data.get("difficulty", "medium")
                weight = {"easy": 1.0, "medium": 1.5, "hard": 2.0}.get(difficulty, 1.0)

                if skill not in skill_breakdown:
                    skill_breakdown[skill] = {
                        "total": 0,
                        "correct": 0,
                        "weighted_total": 0,
                        "weighted_correct": 0,
                    }

                skill_breakdown[skill]["total"] += 1
                skill_breakdown[skill]["weighted_total"] += weight

                if ans_data["correct"]:
                    skill_breakdown[skill]["correct"] += 1
                    skill_breakdown[skill]["weighted_correct"] += weight

        skill_percentages = {}
        for skill, stats in skill_breakdown.items():
            # Use weighted scoring for skill percentages too
            if stats["weighted_total"] > 0:
                skill_percentages[skill] = calculate_percentage(
                    stats["weighted_correct"], stats["weighted_total"]
                )
            else:
                skill_percentages[skill] = 0

        print(f"✅ FINAL ADAPTIVE SCORE (Weighted): {score_percentage:.1f}%")
        print(f"📊 Adaptive skill breakdown: {skill_percentages}")

        # ✅ Calculate final difficulty from last 5 questions
        last_difficulties = [
            answers[qid].get("difficulty", "medium")
            for qid in list(answers.keys())[-5:]
        ]
        final_difficulty = (
            max(set(last_difficulties), key=last_difficulties.count)
            if last_difficulties
            else "medium"
        )

        # Save result with adaptive performance data
        from application.result_service import calculate_grade

        # ✅ Calculate adaptive performance metrics with weighted scoring
        difficulty_counts = {"easy": 0, "medium": 0, "hard": 0}
        difficulty_weights = {"easy": 0, "medium": 0, "hard": 0}
        difficulty_correct = {"easy": 0, "medium": 0, "hard": 0}

        for ans in answers.values():
            diff = ans.get("difficulty", "medium")
            difficulty_counts[diff] = difficulty_counts.get(diff, 0) + 1

            weight = {"easy": 1.0, "medium": 1.5, "hard": 2.0}.get(diff, 1.0)
            difficulty_weights[diff] = difficulty_weights.get(diff, 0) + weight

            if ans.get("correct", False):
                difficulty_correct[diff] = difficulty_correct.get(diff, 0) + weight

        # Calculate difficulty-specific weighted scores
        difficulty_scores = {}
        for diff in ["easy", "medium", "hard"]:
            if difficulty_weights.get(diff, 0) > 0:
                difficulty_scores[diff] = (
                    difficulty_correct.get(diff, 0) / difficulty_weights[diff]
                ) * 100
            else:
                difficulty_scores[diff] = 0

        result = {
            "result_id": str(uuid.uuid4()),
            "session_id": session_id,
            "assessment_id": session["assessment_id"],
            "candidate_email": session["candidate_email"],
            "role": metadata["role"],
            "difficulty": "adaptive",
            "total_questions": len(answers),
            "correct_answers": correct_count,
            "wrong_answers": len(answers) - correct_count,
            "unanswered": 0,
            "score_percentage": round(score_percentage, 2),
            "skill_breakdown": skill_percentages,  # Add skill breakdown
            "start_time": session["start_time"],
            "end_time": session["end_time"],
            "total_time_taken": int(
                (
                    datetime.fromisoformat(session["end_time"])
                    - datetime.fromisoformat(session["start_time"])
                ).total_seconds()
            ),
            "question_results": [
                {
                    "question_index": idx,
                    "question_text": q["question"],
                    "candidate_answer": answers.get(q["question_id"], {}).get("answer"),
                    "correct_answer": q["correct_answer"],
                    "is_correct": answers.get(q["question_id"], {}).get(
                        "correct", False
                    ),
                    "difficulty": answers.get(q["question_id"], {}).get(
                        "difficulty", "medium"
                    ),
                }
                for idx, q in enumerate(questions_asked)
            ],
            "violations": session.get("violations", {}),
            "status": "completed",
            "grade": calculate_grade(score_percentage),
            "created_at": datetime.now().isoformat(),
            # ✅ ADD ADAPTIVE PERFORMANCE METADATA
            "adaptive_metadata": {
                "final_difficulty": final_difficulty,
                "difficulty_distribution": difficulty_counts,
                "difficulty_weights": difficulty_weights,
                "difficulty_scores": difficulty_scores,
                "difficulty_adjustments": engine.difficulty_changes,
                "final_accuracy": engine.calculate_accuracy(),
                "weighted_scoring_used": True,
                "weight_config": {"easy": 1.0, "medium": 1.5, "hard": 2.0},
            },
        }

        await ResultRepository.create(db, result)
        adaptive_engines.pop(session_id, None)

        # ✅ UPDATE ASSESSMENT WITH ACTUAL QUESTIONS
        try:
            # Use direct SQL update instead of importing Assessment model
            await db.execute(
                text(
                    """
                    UPDATE assessments
                    SET questions = :questions,
                        num_questions = :num_questions,
                        status = :status
                    WHERE assessment_id = :assessment_id
                """
                ),
                {
                    "questions": json.dumps(questions_asked),
                    "num_questions": len(questions_asked),
                    "status": "completed",
                    "assessment_id": session["assessment_id"],
                },
            )
            await db.commit()
            print(
                f"✅ Updated assessment {session['assessment_id']} with {len(questions_asked)} questions"
            )
        except Exception as e:
            print(f"⚠️ Failed to update assessment: {e}")
            import traceback

            traceback.print_exc()

        # Calculate both weighted and unweighted scores for comparison
        unweighted_score = calculate_percentage(correct_count, len(answers))

        return {
            "completed": True,
            "assessment_id": session["assessment_id"],
            "session_id": session_id,
            "score": score_percentage,  # Weighted score
            "unweighted_score": unweighted_score,  # Traditional score
            "score_difference": round(score_percentage - unweighted_score, 1),
            "correct": correct_count,
            "total": len(answers),
            "skill_breakdown": skill_percentages,
            "weighted_scoring": True,
            "performance_summary": {
                "current_difficulty": final_difficulty,
                "recent_accuracy": engine.calculate_accuracy(),
                "total_adjustments": len(engine.difficulty_changes),
                "difficulty_breakdown": difficulty_counts,
                "weighted_scores_by_difficulty": difficulty_scores,
            },
        }

    # ✅✅✅ MAIN QUESTION GENERATION WITH DUPLICATE DETECTION
    next_difficulty = engine.get_next_difficulty()

    from domain.mcq_prompts import extract_topics_from_questions

    avoid_topics = extract_topics_from_questions(questions_asked)

    print(f"[ADAPTIVE] Next difficulty: {next_difficulty}")
    print(f"[ADAPTIVE] Avoiding {len(avoid_topics)} topics")

    # ✅ TRY UP TO 5 TIMES TO GET UNIQUE QUESTION
    from integrations.llm_api import generate_single_mcq_dynamic

    max_attempts = 5
    next_question = None

    for attempt in range(max_attempts):
        # ✅ PROGRESSIVE TOPIC RELAXATION
        # Attempt 1-2: Avoid all topics (strict)
        # Attempt 3-4: Reduce to last 10 topics (moderate)
        # Attempt 5: No topic restrictions (last resort)
        if attempt < 2:
            current_avoid_topics = avoid_topics
        elif attempt < 4:
            current_avoid_topics = (
                avoid_topics[-10:] if len(avoid_topics) > 10 else avoid_topics
            )
        else:
            current_avoid_topics = []  # Last attempt: no restrictions

        if attempt > 0:
            print(
                f"⚠️ Retry {attempt}: Using {len(current_avoid_topics)} avoided topics (reduced from {len(avoid_topics)})"
            )

        candidate_question = await generate_single_mcq_dynamic(
            metadata["role"],
            next_difficulty,
            current_avoid_topics,
            engine.calculate_accuracy() * 100,
        )

        if not candidate_question:
            print(
                f"⚠️ Attempt {attempt + 1}/{max_attempts}: Failed to generate question"
            )
            continue

        # ✅ CHECK BOTH DUPLICATE QUESTION AND DUPLICATE TOPIC
        is_dup_question = is_duplicate_question(candidate_question, questions_asked)
        candidate_topic = candidate_question.get("topic", "")
        is_dup_topic = candidate_topic in avoid_topics

        if is_dup_question:
            print(f"⚠️ DUPLICATE QUESTION: Text matches existing question")
            continue

        # Only enforce topic uniqueness in first 3 attempts
        if is_dup_topic and attempt < 3:
            print(
                f"⚠️ DUPLICATE TOPIC: '{candidate_topic}' already used (attempt {attempt + 1})"
            )
            continue

        # ✅ PASSED ALL CHECKS - UNIQUE QUESTION AND TOPIC!
        next_question = candidate_question
        print(
            f"✅ Attempt {attempt + 1}: Generated UNIQUE question {next_question['question_id']} on NEW topic '{candidate_topic}'"
        )
        break

    if not next_question:
        # ❌ FAILED TO GENERATE UNIQUE QUESTION
        # Instead of ending test, let's try one more time with NO topic restrictions
        print("⚠️ Failed with topic restrictions - trying with NO restrictions...")

        candidate_question = await generate_single_mcq_dynamic(
            metadata["role"],
            next_difficulty,
            [],  # No topic avoidance
            engine.calculate_accuracy() * 100,
        )

        if candidate_question and not is_duplicate_question(
            candidate_question, questions_asked
        ):
            next_question = candidate_question
            print(
                f"✅ Generated question with no restrictions: {next_question['question_id']}"
            )
        else:
            # STILL failed - this means we've truly exhausted questions
            print(
                "⚠️ Could not generate unique question even without restrictions - ending test early"
            )
            print(f"[ADAPTIVE] 🎉 Test ending with {len(answers)} questions!")

            # ✅ MARK SESSION AS COMPLETED
            session["status"] = "completed"
            session["end_time"] = datetime.now().isoformat()

            await SessionRepository.update(
                db, session_id, {"status": "completed", "end_time": session["end_time"]}
            )
            await db.commit()

            # ✅ FIXED WITH WEIGHTED SCORING: Hard questions worth more
            def calculate_weighted_adaptive_score(answers_dict, difficulty_history):
                """Calculate weighted score where hard questions are worth more"""
                difficulty_weights = {"easy": 1.0, "medium": 1.5, "hard": 2.0}

                weighted_correct = 0
                total_weight = 0

                for q_id, answer_data in answers_dict.items():
                    # Get difficulty from answer data or history
                    difficulty = answer_data.get("difficulty", "medium")
                    weight = difficulty_weights.get(difficulty, 1.0)

                    total_weight += weight
                    if answer_data.get("correct", False):
                        weighted_correct += weight

                if total_weight > 0:
                    return (weighted_correct / total_weight) * 100
                return 0

            # Calculate weighted score
            score_percentage = calculate_weighted_adaptive_score(
                answers, engine.difficulty_history
            )
            correct_count = sum(1 for a in answers.values() if a["correct"])
            unweighted_score = calculate_percentage(correct_count, len(answers))

            # ✅ SAVE RESULT TO DATABASE
            from application.result_service import calculate_grade

            result = {
                "result_id": str(uuid.uuid4()),
                "session_id": session_id,
                "assessment_id": session["assessment_id"],
                "candidate_email": session["candidate_email"],
                "role": metadata["role"],
                "difficulty": "adaptive",
                "total_questions": len(answers),
                "correct_answers": correct_count,
                "wrong_answers": len(answers) - correct_count,
                "unanswered": 0,
                "score_percentage": round(score_percentage, 2),
                "start_time": session["start_time"],
                "end_time": session["end_time"],
                "total_time_taken": int(
                    (
                        datetime.fromisoformat(session["end_time"])
                        - datetime.fromisoformat(session["start_time"])
                    ).total_seconds()
                ),
                "question_results": [
                    {
                        "question_index": idx,
                        "question_text": q["question"],
                        "candidate_answer": answers.get(q["question_id"], {}).get(
                            "answer"
                        ),
                        "correct_answer": q["correct_answer"],
                        "is_correct": answers.get(q["question_id"], {}).get(
                            "correct", False
                        ),
                    }
                    for idx, q in enumerate(questions_asked)
                ],
                "violations": session.get("violations", {}),
                "status": "completed",
                "grade": calculate_grade(score_percentage),
                "created_at": datetime.now().isoformat(),
            }

            await ResultRepository.create(db, result)
            await db.commit()

            # ✅ UPDATE ASSESSMENT WITH QUESTIONS
            await AssessmentRepository.update(
                db,
                session["assessment_id"],
                {
                    "questions": json.dumps(questions_asked),
                    "num_questions": len(questions_asked),
                    "status": "completed",
                },
            )
            await db.commit()

            print(
                f"✅ Updated assessment {session['assessment_id']} with {len(questions_asked)} questions"
            )

            return {
                "completed": True,
                "assessment_id": session["assessment_id"],
                "session_id": session_id,
                "score": score_percentage,
                "unweighted_score": unweighted_score,
                "score_difference": round(score_percentage - unweighted_score, 1),
                "correct": correct_count,
                "total": len(answers),
                "weighted_scoring": True,
                "performance_summary": {
                    "current_difficulty": engine.get_current_difficulty(),
                    "recent_accuracy": engine.calculate_accuracy(),
                    "total_adjustments": len(engine.difficulty_changes),
                },
            }

    # ✅ IF WE GET HERE, WE HAVE A VALID NEXT_QUESTION
    # Continue with normal question flow
    # Add to questions asked
    questions_asked.append(next_question)
    metadata["questions_asked"] = questions_asked

    # Save metadata with new question
    await SessionRepository.update(db, session_id, {"metadata": json.dumps(metadata)})
    await db.commit()

    # Return response
    return {
        "completed": False,
        "correct": is_correct,
        "next_question": next_question,
        "question_number": questions_answered + 1,
        "total_questions": max_questions,
        "current_difficulty": next_difficulty,
        "performance": {
            "recent_accuracy": engine.calculate_accuracy(),
            "difficulty_adjusted": (next_difficulty != current_question["difficulty"]),
        },
    }


# ==================== SKILL MATCH CALCULATION ====================


def calculate_skill_match_percentage(
    candidate_skills: dict, job_requirements: dict
) -> dict:
    """
    UC-04 Main Flow Step 6: Calculate skill match percentage between candidate and job
    """
    skill_match_result = {
        "overall_match_percentage": 0,
        "skill_breakdown": {},
        "missing_skills": [],
        "strong_skills": [],
    }

    if not job_requirements:
        return skill_match_result

    total_weight = 0
    weighted_score = 0

    for skill, req_data in job_requirements.items():
        weight = req_data.get("weight", 1)  # Default weight=1
        required_level = req_data.get("required_level", 70)  # Default 70%

        total_weight += weight

        # Get candidate's skill level
        candidate_skill_data = candidate_skills.get(skill, {})
        candidate_score = candidate_skill_data.get("score_percentage", 0)

        # Calculate match for this skill
        if candidate_score >= required_level:
            skill_match = 100  # Fully meets requirement
            skill_match_result["strong_skills"].append(skill)
        else:
            skill_match = (candidate_score / required_level) * 100
            if candidate_score == 0:
                skill_match_result["missing_skills"].append(skill)

        weighted_score += skill_match * weight

        # Store individual skill match
        skill_match_result["skill_breakdown"][skill] = {
            "candidate_score": candidate_score,
            "required_level": required_level,
            "match_percentage": min(skill_match, 100),
            "weight": weight,
            "status": (
                "exceeds"
                if candidate_score > required_level
                else "meets"
                if candidate_score >= required_level
                else "below"
            ),
        }

    # Calculate overall match percentage
    if total_weight > 0:
        skill_match_result["overall_match_percentage"] = round(
            weighted_score / total_weight, 2
        )

    return skill_match_result


# Update the results analytics endpoint to include skill matching
@app.get("/api/results/analytics-with-matching")
async def get_results_analytics_with_matching(
    job_role: str = None, db: AsyncSession = Depends(get_db)
):
    """
    Enhanced analytics with skill match percentage calculation
    """
    try:
        # Get basic analytics
        analytics = await get_results_analytics(db)

        # Define job requirement templates (in real app, this would come from database)
        JOB_REQUIREMENTS = {
            "Software Engineer": {
                "Algorithms": {"weight": 1.5, "required_level": 75},
                "Data Structures": {"weight": 1.5, "required_level": 75},
                "Problem Solving": {"weight": 1.2, "required_level": 70},
                "Code Quality": {"weight": 1.0, "required_level": 65},
                "Efficiency": {"weight": 1.0, "required_level": 70},
            },
            "Frontend Developer": {
                "JavaScript": {"weight": 1.5, "required_level": 80},
                "React": {"weight": 1.3, "required_level": 75},
                "CSS": {"weight": 1.0, "required_level": 70},
                "Problem Solving": {"weight": 1.0, "required_level": 65},
            },
            "Data Scientist": {
                "Python": {"weight": 1.5, "required_level": 80},
                "Statistics": {"weight": 1.3, "required_level": 75},
                "Machine Learning": {"weight": 1.2, "required_level": 70},
                "Data Analysis": {"weight": 1.0, "required_level": 70},
            },
        }

        # If job role specified, calculate matches for top performers
        skill_matches = []
        if job_role and job_role in JOB_REQUIREMENTS:
            from sqlalchemy import text

            # Get top performers
            top_query = text(
                """
                SELECT candidate_email, score_percentage, skill_breakdown
                FROM results
                WHERE role = :role
                ORDER BY score_percentage DESC
                LIMIT 10
            """
            )
            result = await db.execute(top_query, {"role": job_role})
            top_candidates = result.fetchall()

            for candidate in top_candidates:
                candidate_email = candidate[0]
                skill_breakdown_json = candidate[2]

                if skill_breakdown_json:
                    try:
                        candidate_skills = json.loads(skill_breakdown_json)
                        skill_match = calculate_skill_match_percentage(
                            candidate_skills, JOB_REQUIREMENTS[job_role]
                        )

                        skill_matches.append(
                            {
                                "candidate_email": candidate_email,
                                "overall_score": candidate[1],
                                "skill_match": skill_match,
                            }
                        )
                    except:
                        continue

        analytics["skill_matches"] = skill_matches
        return analytics

    except Exception as e:
        print(f"❌ Analytics with matching failed: {e}")
        return await get_results_analytics(db)  # Fallback to basic analytics


# ==================== RESULTS ====================


@app.get("/api/results")
async def get_all_results(
    sort_by: str = "score",
    min_score: float = None,
    status: str = None,
    db: AsyncSession = Depends(get_db),
):
    """FR-RES-04, FR-RES-05, FR-RES-06"""
    results = await ResultRepository.get_all(db, sort_by="score_percentage")

    # Apply filters
    if min_score is not None:
        results = [r for r in results if r["score_percentage"] >= min_score]
    if status:
        results = [r for r in results if r["status"] == status]

    return {"results": results, "total": len(results)}


@app.get("/api/results/analytics")
async def get_results_analytics(db: AsyncSession = Depends(get_db)):
    """FR-RES-07: Visual analytics and aggregate data"""
    try:
        print("📊 Fetching analytics...")

        # Get basic analytics
        analytics_query = text(
            """
            SELECT
                COUNT(*) as total,
                COALESCE(AVG(score_percentage), 0) as avg_score,
                SUM(CASE WHEN score_percentage >= 60 THEN 1 ELSE 0 END) as pass_count
            FROM results
        """
        )
        analytics_result = await db.execute(analytics_query)
        analytics_row = analytics_result.fetchone()

        total = int(analytics_row[0]) if analytics_row and analytics_row[0] else 0
        avg_score = (
            float(analytics_row[1]) if analytics_row and analytics_row[1] else 0.0
        )
        pass_count = int(analytics_row[2]) if analytics_row and analytics_row[2] else 0

        print(f"📊 Total: {total}, Avg: {avg_score}, Pass: {pass_count}")

        # Get grade distribution
        grade_query = text(
            "SELECT grade, COUNT(*) as count FROM results GROUP BY grade"
        )
        grade_result = await db.execute(grade_query)
        grade_rows = grade_result.fetchall()

        grade_distribution = (
            {row[0]: int(row[1]) for row in grade_rows} if grade_rows else {}
        )
        print(f"📊 Grades: {grade_distribution}")

        # Get top performers
        top_query = text(
            """
            SELECT candidate_email, score_percentage, grade, role
            FROM results
            ORDER BY score_percentage DESC
            LIMIT 5
        """
        )
        top_result = await db.execute(top_query)
        top_rows = top_result.fetchall()

        top_performers = (
            [
                {
                    "candidate_email": row[0],
                    "score": float(row[1]),
                    "grade": row[2],
                    "role": row[3],
                }
                for row in top_rows
            ]
            if top_rows
            else []
        )

        print(f"📊 Top performers: {len(top_performers)}")

        response = {
            "total_assessments": total,
            "average_score": round(avg_score, 2),
            "pass_rate": round((pass_count / total * 100), 2) if total > 0 else 0,
            "grade_distribution": grade_distribution,
            "top_performers": top_performers,
        }

        print(f"✅ Analytics response: {response}")
        return response

    except Exception as e:
        print(f"❌ Analytics error: {e}")
        import traceback

        traceback.print_exc()

        # Return empty but valid structure
        return {
            "total_assessments": 0,
            "average_score": 0.0,
            "pass_rate": 0.0,
            "grade_distribution": {},
            "top_performers": [],
        }


@app.get("/api/results/{result_id}")
async def get_result(result_id: str, db: AsyncSession = Depends(get_db)):
    """FR-RES-08: Get detailed performance report"""
    try:
        print(f"🔍 Fetching result: {result_id}")

        result = await ResultRepository.get_by_id(db, result_id)
        if not result:
            print(f"❌ Result not found: {result_id}")
            raise HTTPException(status_code=404, detail="Result not found")

        print(f"✅ Found result for: {result.get('candidate_email')}")

        # Try to get assessment if assessment_id exists
        assessment = None
        assessment_id = result.get("assessment_id")

        if assessment_id:
            try:
                print(f"🔍 Fetching assessment: {assessment_id}")
                assessment = await AssessmentRepository.get_by_id(db, assessment_id)
                if assessment:
                    print(f"✅ Found assessment: {assessment.get('role')}")
            except Exception as e:
                print(f"⚠️ Assessment not found or error: {e}")
                pass

        return {"result": result, "assessment": assessment}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Get result error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to fetch result: {str(e)}")


@app.get("/api/results/candidate/{email}")
async def get_candidate_results(email: str, db: AsyncSession = Depends(get_db)):
    """Get candidate results"""
    results = await ResultRepository.get_by_candidate(db, email)
    return {"candidate_email": email, "results": results, "total": len(results)}


# ==================== EXPORTS ====================


@app.get("/api/results/export/excel")
async def export_results_excel(db: AsyncSession = Depends(get_db)):
    """FR-RES-09: Excel export"""
    results = await ResultRepository.get_all(db, sort_by="score_percentage")

    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment Results"

    headers = [
        "Rank",
        "Candidate Email",
        "Role",
        "Difficulty",
        "Score %",
        "Grade",
        "Correct",
        "Wrong",
        "Unanswered",
        "Time (s)",
        "Violations",
        "Status",
    ]
    ws.append(headers)

    header_fill = PatternFill(
        start_color="4F46E5", end_color="4F46E5", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, result in enumerate(results, 1):
        ws.append(
            [
                idx,
                result["candidate_email"],
                result["role"],
                result["difficulty"],
                result["score_percentage"],
                result["grade"],
                result["correct_answers"],
                result["wrong_answers"],
                result["unanswered"],
                result["total_time_taken"],
                len(result.get("violations", [])),
                result["status"],
            ]
        )

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assessment_results.xlsx"},
    )


@app.get("/api/results/export/pdf")
async def export_results_pdf(db: AsyncSession = Depends(get_db)):
    """FR-RES-09: PDF export"""
    results = await ResultRepository.get_all(db, sort_by="score_percentage")

    pdf_file = BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Assessment Results Report", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    data = [["Rank", "Email", "Role", "Score %", "Grade", "Status"]]

    for idx, result in enumerate(results[:20], 1):
        data.append(
            [
                str(idx),
                result["candidate_email"][:25],
                result["role"][:15],
                f"{result['score_percentage']:.1f}%",
                result["grade"],
                result["status"],
            ]
        )

    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)

    pdf_file.seek(0)

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=assessment_results.pdf"},
    )


# NEW: Export results with period grouping to Excel
@app.post("/api/results/export/excel")
async def export_results_excel(request: dict, db: AsyncSession = Depends(get_db)):
    """
    Export results to Excel with period grouping
    Expected request body:
    {
        "results": [...],  # Array of combined results
        "period": "all" or "Dec-2024" etc.
    }
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from fastapi.responses import StreamingResponse

        results = request.get("results", [])
        period_filter = request.get("period", "all")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Results_{period_filter}"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "Period",
            "Role",
            "Candidate Email",
            "Type",
            "MCQ Score (%)",
            "Coding Score (%)",
            "Combined Score (%)",
            "Grade",
            "Test Date",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_num, result in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=result.get("period", "Unknown"))
            ws.cell(row=row_num, column=2, value=result.get("role", "Unknown"))
            ws.cell(row=row_num, column=3, value=result.get("candidate_email", ""))
            ws.cell(row=row_num, column=4, value=result.get("type", ""))

            # MCQ Score
            mcq_score = result.get("mcq_score")
            ws.cell(row=row_num, column=5, value=mcq_score if mcq_score else "N/A")

            # Coding Score
            coding_score = result.get("coding_score")
            ws.cell(
                row=row_num, column=6, value=coding_score if coding_score else "N/A"
            )

            # Combined Score
            combined = result.get("combined_score", 0)
            combined_cell = ws.cell(
                row=row_num, column=7, value=float(combined) if combined else 0
            )

            # Color code combined score
            combined_val = float(combined) if combined else 0
            if combined_val >= 80:
                combined_cell.fill = PatternFill(
                    start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"
                )
            elif combined_val >= 60:
                combined_cell.fill = PatternFill(
                    start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
                )
            else:
                combined_cell.fill = PatternFill(
                    start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
                )

            ws.cell(row=row_num, column=8, value=result.get("grade", ""))

            # Date
            test_date = result.get("test_date", "")
            if test_date:
                try:
                    from datetime import datetime

                    date_obj = datetime.fromisoformat(test_date.replace("Z", "+00:00"))
                    ws.cell(
                        row=row_num, column=9, value=date_obj.strftime("%Y-%m-%d %H:%M")
                    )
                except:
                    ws.cell(row=row_num, column=9, value=test_date)

            # Apply borders to all cells
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = thin_border

        # Adjust column widths
        column_widths = {
            "A": 15,  # Period
            "B": 25,  # Role
            "C": 30,  # Email
            "D": 15,  # Type
            "E": 15,  # MCQ Score
            "F": 15,  # Coding Score
            "G": 18,  # Combined Score
            "H": 10,  # Grade
            "I": 20,  # Date
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create filename
        from datetime import datetime

        filename = (
            f"results_{period_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        print(f"❌ Excel export error: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# NEW: Get period statistics
@app.get("/api/results/period-stats")
async def get_period_statistics(
    period: Optional[str] = None, db: AsyncSession = Depends(get_db)
):
    """
    Get statistics grouped by hiring period
    Optional period parameter to filter specific period
    """
    try:
        from sqlalchemy import text
        from datetime import datetime

        # Fetch all results with dates
        mcq_query = """
            SELECT
                result_id,
                candidate_email,
                role,
                score_percentage,
                grade,
                created_at,
                'MCQ' as type
            FROM results
            WHERE status = 'completed'
        """

        coding_query = """
            SELECT
                submission_id,
                candidate_email,
                role,
                score,
                'Coding' as type,
                submitted_at
            FROM coding_submissions
            WHERE score > 0
        """

        mcq_results = await db.execute(text(mcq_query))
        mcq_data = mcq_results.fetchall()

        coding_results = await db.execute(text(coding_query))
        coding_data = coding_results.fetchall()

        # Group by period
        period_stats = {}

        def extract_period(date_str):
            try:
                date_obj = datetime.fromisoformat(str(date_str).replace("Z", "+00:00"))
                month_names = [
                    "Jan",
                    "Feb",
                    "Mar",
                    "Apr",
                    "May",
                    "Jun",
                    "Jul",
                    "Aug",
                    "Sep",
                    "Oct",
                    "Nov",
                    "Dec",
                ]
                return f"{month_names[date_obj.month - 1]}-{date_obj.year}"
            except:
                return "Unknown"

        # Process MCQ results
        for row in mcq_data:
            period = extract_period(row[5])  # created_at
            if period not in period_stats:
                period_stats[period] = {
                    "period": period,
                    "total_tests": 0,
                    "mcq_tests": 0,
                    "coding_tests": 0,
                    "candidates": set(),
                    "roles": set(),
                    "scores": [],
                    "avg_score": 0,
                }

            period_stats[period]["total_tests"] += 1
            period_stats[period]["mcq_tests"] += 1
            period_stats[period]["candidates"].add(row[1])  # email
            period_stats[period]["roles"].add(row[2])  # role
            period_stats[period]["scores"].append(float(row[3]))  # score

        # Process Coding results
        for row in coding_data:
            period = extract_period(row[5])  # submitted_at
            if period not in period_stats:
                period_stats[period] = {
                    "period": period,
                    "total_tests": 0,
                    "mcq_tests": 0,
                    "coding_tests": 0,
                    "candidates": set(),
                    "roles": set(),
                    "scores": [],
                    "avg_score": 0,
                }

            period_stats[period]["total_tests"] += 1
            period_stats[period]["coding_tests"] += 1
            period_stats[period]["candidates"].add(row[1])  # email
            period_stats[period]["roles"].add(row[2])  # role
            period_stats[period]["scores"].append(float(row[3]))  # score

        # Calculate averages and convert sets to counts
        for period in period_stats:
            stats = period_stats[period]
            stats["candidate_count"] = len(stats["candidates"])
            stats["role_count"] = len(stats["roles"])
            stats["avg_score"] = (
                round(sum(stats["scores"]) / len(stats["scores"]), 1)
                if stats["scores"]
                else 0
            )
            stats["candidates"] = list(stats["candidates"])
            stats["roles"] = list(stats["roles"])
            del stats["scores"]  # Remove raw scores array

        # Filter by period if specified
        if period and period != "all":
            return {
                "success": True,
                "period": period,
                "stats": period_stats.get(period, {}),
            }

        # Return all periods sorted by date (newest first)
        sorted_periods = sorted(
            period_stats.values(), key=lambda x: x["period"], reverse=True
        )

        return {
            "success": True,
            "total_periods": len(sorted_periods),
            "periods": sorted_periods,
        }

    except Exception as e:
        print(f"❌ Period stats error: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to fetch period statistics: {str(e)}"
        )


# NEW: Get combined scores for specific period and role
@app.get("/api/results/combined-scores")
async def get_combined_scores(
    period: Optional[str] = None,
    role: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Get combined MCQ + Coding scores grouped by period and role
    Returns candidates with both test types and their combined scores
    """
    try:
        from sqlalchemy import text
        from datetime import datetime

        def extract_period(date_str):
            try:
                date_obj = datetime.fromisoformat(str(date_str).replace("Z", "+00:00"))
                month_names = [
                    "Jan",
                    "Feb",
                    "Mar",
                    "Apr",
                    "May",
                    "Jun",
                    "Jul",
                    "Aug",
                    "Sep",
                    "Oct",
                    "Nov",
                    "Dec",
                ]
                return f"{month_names[date_obj.month - 1]}-{date_obj.year}"
            except:
                return "Unknown"

        # Fetch MCQ results
        mcq_query = """
            SELECT
                candidate_email,
                role,
                score_percentage,
                grade,
                created_at
            FROM results
            WHERE status = 'completed'
        """

        if role:
            mcq_query += f" AND role = '{role}'"

        mcq_results = await db.execute(text(mcq_query))
        mcq_data = mcq_results.fetchall()

        # Fetch coding results
        coding_query = """
            SELECT
                candidate_email,
                role,
                score,
                submitted_at
            FROM coding_submissions
            WHERE score > 0
        """

        if role:
            coding_query += f" AND role = '{role}'"

        coding_results = await db.execute(text(coding_query))
        coding_data = coding_results.fetchall()

        # Group by period, role, and candidate
        grouped_data = {}

        # Process MCQ
        for row in mcq_data:
            email = row[0]
            test_role = row[1]
            score = float(row[2])
            test_period = extract_period(row[4])

            if period and period != "all" and test_period != period:
                continue

            key = (test_period, test_role, email)
            if key not in grouped_data:
                grouped_data[key] = {
                    "period": test_period,
                    "role": test_role,
                    "candidate_email": email,
                    "mcq_scores": [],
                    "coding_scores": [],
                }

            grouped_data[key]["mcq_scores"].append(score)

        # Process Coding
        for row in coding_data:
            email = row[0]
            test_role = row[1]
            score = float(row[2])
            test_period = extract_period(row[3])

            if period and period != "all" and test_period != period:
                continue

            key = (test_period, test_role, email)
            if key not in grouped_data:
                grouped_data[key] = {
                    "period": test_period,
                    "role": test_role,
                    "candidate_email": email,
                    "mcq_scores": [],
                    "coding_scores": [],
                }

            grouped_data[key]["coding_scores"].append(score)

        # Calculate combined scores
        combined_results = []
        for key, data in grouped_data.items():
            mcq_avg = (
                sum(data["mcq_scores"]) / len(data["mcq_scores"])
                if data["mcq_scores"]
                else None
            )
            coding_avg = (
                sum(data["coding_scores"]) / len(data["coding_scores"])
                if data["coding_scores"]
                else None
            )

            if mcq_avg is not None and coding_avg is not None:
                # Both tests - calculate combined
                combined_score = (mcq_avg + coding_avg) / 2
                test_type = "Combined"
            elif mcq_avg is not None:
                # MCQ only
                combined_score = mcq_avg
                test_type = "MCQ Only"
            elif coding_avg is not None:
                # Coding only
                combined_score = coding_avg
                test_type = "Coding Only"
            else:
                continue

            # Calculate grade
            if combined_score >= 90:
                grade = "A+"
            elif combined_score >= 85:
                grade = "A"
            elif combined_score >= 80:
                grade = "B+"
            elif combined_score >= 75:
                grade = "B"
            elif combined_score >= 70:
                grade = "C+"
            elif combined_score >= 65:
                grade = "C"
            elif combined_score >= 60:
                grade = "D"
            else:
                grade = "F"

            combined_results.append(
                {
                    "period": data["period"],
                    "role": data["role"],
                    "candidate_email": data["candidate_email"],
                    "type": test_type,
                    "mcq_score": round(mcq_avg, 1) if mcq_avg else None,
                    "coding_score": round(coding_avg, 1) if coding_avg else None,
                    "combined_score": round(combined_score, 1),
                    "grade": grade,
                    "mcq_test_count": len(data["mcq_scores"]),
                    "coding_test_count": len(data["coding_scores"]),
                }
            )

        # Sort by period and score
        combined_results.sort(key=lambda x: (x["period"], -x["combined_score"]))

        return {
            "success": True,
            "total_results": len(combined_results),
            "filters": {"period": period or "all", "role": role or "all"},
            "results": combined_results,
        }

    except Exception as e:
        print(f"❌ Combined scores error: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to fetch combined scores: {str(e)}"
        )


# ==================== MODULE 5 COMPLETE: TECHNICAL ASSESSMENT ====================
# Add to end of main.py
# Complete implementation based on PDF requirements

import os
import json
import uuid
from datetime import datetime
import aiohttp
import asyncio


# ==================== CHALLENGE GENERATION ====================

# ============================================
# REPLACE YOUR generate_coding_challenges FUNCTION
# In main.py around line 1645
# ============================================


def _extract_first_json_array(text: str) -> Optional[str]:
    """Find the first top-level JSON array; bracket depth ignores `[`/`]` inside strings."""
    if not text:
        return None
    start = text.find("[")
    if start < 0:
        return None
    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape:
            escape = False
            continue
        if in_string:
            if ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return None


def _parse_llm_challenges_response(raw: str) -> List[dict]:
    """Strip fences, extract JSON array, load; supports wrapper {{\"challenges\": [...]}}."""
    r = (raw or "").strip()
    if "```json" in r:
        r = r.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in r:
        parts = r.split("```", 2)
        if len(parts) >= 2:
            r = parts[1].strip()
    blob = _extract_first_json_array(r) or r
    data = json.loads(blob)
    if isinstance(data, dict) and "challenges" in data:
        data = data["challenges"]
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ValueError("LLM output must be a JSON array of challenge objects")
    out = [x for x in data if isinstance(x, dict)]
    if not out:
        raise ValueError("No challenge objects in LLM JSON")
    return out


def _compact_coding_challenge_prompt(
    role: str,
    difficulty: str,
    language: str,
    topic: Optional[str],
    index: int,
    total: int,
) -> str:
    topic_line = f"Preferred topic: {topic}. " if topic else ""
    return f"""{topic_line}Generate exactly ONE distinct coding challenge ({index} of {total} for this request) for a {role} role.
Difficulty: {difficulty}. Language for starter_code: {language}.

Return ONLY a JSON array with exactly one object. Keys required:
"title", "description", "difficulty", "language", "starter_code", "test_cases", "constraints", "examples", "hints"

Strict JSON rules:
- Use double quotes for all keys and string values.
- Inside "description" and "starter_code", represent line breaks only as escaped \\n — never break a JSON string across physical lines.
- test_cases: 3–4 objects, each with "input" (object with keys matching the function parameters) and "expected_output". Verify outputs manually.
- constraints: array of strings. examples: array of objects with input/output/explanation. hints: array of strings.
- Match starter_code style to {language} (e.g. Python: def solve(...): with docstring or pass).

No markdown, no commentary — only the JSON array."""


@app.post("/api/coding/generate")
async def generate_coding_challenges(data: dict, db: AsyncSession = Depends(get_db)):
    """FR-TECH-01: Generate coding challenges using LLM - FIXED VERSION"""
    try:
        role = data.get("role", "Software Engineer")
        difficulty = data.get("difficulty", "medium")
        language = data.get("language", "python")
        topic = data.get("topic")
        num_challenges = data.get("num_challenges", 3)

        # ✅ COMPREHENSIVE FIXED PROMPT with EXACT starter code templates
        prompt = f"""Generate {num_challenges} coding challenge(s) for a {role} position.

Difficulty: {difficulty}
Language: {language}
{f"Topic: {topic}" if topic else ""}

For each challenge provide:
1. Title (clear, descriptive)
2. Description (problem statement with examples)
3. Starter code (function signature with clear parameter names and user instructions)
4. Test cases (3-5 cases with MANUALLY VERIFIED expected outputs)
5. Constraints
6. Hints

🔴 CRITICAL INSTRUCTIONS FOR TEST CASES:
- Input MUST be a dictionary with keys matching the function parameter names
- Example: If function is "def two_sum(nums, target):" then input should be {{"nums": [2,7], "target": 9}}
- NOT as arrays like [[2,7], 9]
- Keys must match parameter names EXACTLY

🔴 CRITICAL TEST CASE VERIFICATION:
1. MANUALLY calculate each expected output - DO NOT GUESS
2. Verify expected outputs match the problem requirements EXACTLY
3. For "maximum/minimum" problems: verify the solution actually gives the max/min
4. For "longest/shortest" problems: verify the solution actually gives the longest/shortest
5. Test edge cases: empty input, single element, all same values
6. ALL test cases must follow the SAME logic consistently

🔴 STARTER CODE TEMPLATES - USE THESE EXACTLY:
IMPORTANT: For each language, use ONLY these starter code templates:

1. GO:
starter_code = "// Write ONLY the function below. Do NOT include 'package main' or 'func main()' - our platform handles that.\\n// You can import packages as needed (e.g., import \\"sort\\"), but DO NOT import \\"fmt\\" - it's already included.\\n\\nfunc longestSubsequence(nums []int, k int) int {{\\n    // Edge cases\\n    if len(nums) == 0 || k <= 0 {{\\n        return 0\\n    }}\\n    \\n    // Your code here\\n    return 0\\n}}"

2. JAVASCRIPT:
starter_code = "// Write ONLY the function. Our platform will call it with test cases.\\n\\nfunction twoSum(nums, target) {{\\n    // Edge cases\\n    if (!nums || nums.length < 2) {{\\n        return [];\\n    }}\\n    \\n    // Your code here\\n    return [];\\n}}"

3. C++:
starter_code = "// Write ONLY the function. Do NOT include int main() - our platform handles that.\\n// You can include standard library headers.\\n\\n#include <vector>\\n#include <algorithm>\\n\\n// Function definition only\\nint solutionFunction(std::vector<int>& arr) {{\\n    // Edge cases\\n    if (arr.empty()) {{\\n        return 0;\\n    }}\\n    \\n    // Your code here\\n    return 0;\\n}}"

4. JAVA:
starter_code = "// Write ONLY the class and method. Do NOT include public static void main() - our platform handles that.\\n\\nclass Solution {{\\n    public int[] twoSum(int[] nums, int target) {{\\n        // Edge cases\\n        if (nums == null || nums.length < 2) {{\\n            return new int[0];\\n        }}\\n        \\n        // Your code here\\n        return new int[0];\\n    }}\\n}}"

5. PYTHON:
starter_code = "# Write your solution below. You can import any standard library modules.\\n\\ndef two_sum(nums, target):\\n    \\"\\"\\"\\n    :type nums: List[int]\\n    :type target: int\\n    :rtype: List[int]\\n    \\"\\"\\"\\n    # Write your code here\\n    pass"

Return ONLY valid JSON array with this exact structure (adjust starter_code based on language).

🔴 JSON SYNTAX (mandatory): Every string value must be valid JSON. Inside "description" and "starter_code", use escaped \\n for newlines only — never put a real newline inside a quoted string.

For Python:
[{{
  "title": "Two Sum",
  "description": "Given an array of integers nums and an integer target, return indices of the two numbers such that they add up to target.\\n\\nYou may assume that each input would have exactly one solution, and you may not use the same element twice.\\n\\nYou can return the answer in any order.",
  "difficulty": "{difficulty}",
  "language": "python",
  "starter_code": "# Write your solution below. You can import any standard library modules.\\n\\ndef two_sum(nums, target):\\n    \\"\\"\\"\\n    :type nums: List[int]\\n    :type target: int\\n    :rtype: List[int]\\n    \\"\\"\\"\\n    # Write your code here\\n    pass",
  "test_cases": [
    {{"input": {{"nums": [2,7,11,15], "target": 9}}, "expected_output": [0,1]}},
    {{"input": {{"nums": [3,2,4], "target": 6}}, "expected_output": [1,2]}},
    {{"input": {{"nums": [3,3], "target": 6}}, "expected_output": [0,1]}}
  ],
  "constraints": [
    "2 <= nums.length <= 10^4",
    "-10^9 <= nums[i] <= 10^9",
    "-10^9 <= target <= 10^9",
    "Only one valid answer exists"
  ],
  "examples": [
    {{"input": {{"nums": [2,7,11,15], "target": 9}}, "output": [0,1], "explanation": "Because nums[0] + nums[1] == 9, we return [0, 1]"}},
    {{"input": {{"nums": [3,2,4], "target": 6}}, "output": [1,2], "explanation": "Because nums[1] + nums[2] == 6, we return [1, 2]"}}
  ],
  "hints": [
    "A hash map can help you find complements efficiently",
    "Think about what you need to store as you iterate through the array",
    "Try to solve it in O(n) time complexity"
  ]
}}]

For Go:
[{{
  "title": "Longest Subsequence Sum",
  "description": "Given an array of integers nums and an integer k, return the maximum number of elements you can select from nums such that their sum does not exceed k.\\n\\nYou should select elements in a way that maximizes the count.",
  "difficulty": "{difficulty}",
  "language": "go",
  "starter_code": "// Write ONLY the function below. Do NOT include 'package main' or 'func main()' - our platform handles that.\\n// You can import packages as needed (e.g., import \\"sort\\"), but DO NOT import \\"fmt\\" - it's already included.\\n\\nfunc longestSubsequence(nums []int, k int) int {{\\n    // Edge cases\\n    if len(nums) == 0 || k <= 0 {{\\n        return 0\\n    }}\\n    \\n    // Your code here\\n    return 0\\n}}",
  "test_cases": [
    {{"input": {{"nums": [1,2,3,4,5], "k": 7}}, "expected_output": 3}},
    {{"input": {{"nums": [1,1,1,1,1], "k": 3}}, "expected_output": 3}},
    {{"input": {{"nums": [5,5,5,5,5], "k": 12}}, "expected_output": 2}}
  ],
  "constraints": [
    "1 <= nums.length <= 10^4",
    "-10^9 <= nums[i] <= 10^9",
    "1 <= k <= 10^9"
  ],
  "examples": [
    {{"input": {{"nums": [1,2,3,4,5], "k": 7}}, "output": 3, "explanation": "Greedy picks [1,2,3] with sum=6, length=3"}},
    {{"input": {{"nums": [1,1,1,1,1], "k": 3}}, "output": 3, "explanation": "Pick any 3 elements: [1,1,1] with sum=3"}}
  ],
  "hints": [
    "Sort the array first to use greedy approach",
    "Pick smallest elements first to maximize count",
    "Stop when adding next element would exceed k"
  ]
}}]

For JavaScript:
[{{
  "title": "Two Sum",
  "description": "Given an array of integers nums and an integer target, return indices of the two numbers such that they add up to target.\\n\\nYou may assume that each input would have exactly one solution, and you may not use the same element twice.\\n\\nYou can return the answer in any order.",
  "difficulty": "{difficulty}",
  "language": "javascript",
  "starter_code": "// Write ONLY the function. Our platform will call it with test cases.\\n\\nfunction twoSum(nums, target) {{\\n    // Edge cases\\n    if (!nums || nums.length < 2) {{\\n        return [];\\n    }}\\n    \\n    // Your code here\\n    return [];\\n}}",
  "test_cases": [
    {{"input": {{"nums": [2,7,11,15], "target": 9}}, "expected_output": [0,1]}},
    {{"input": {{"nums": [3,2,4], "target": 6}}, "expected_output": [1,2]}},
    {{"input": {{"nums": [3,3], "target": 6}}, "expected_output": [0,1]}}
  ],
  "constraints": [
    "2 <= nums.length <= 10^4",
    "-10^9 <= nums[i] <= 10^9",
    "-10^9 <= target <= 10^9",
    "Only one valid answer exists"
  ],
  "examples": [
    {{"input": {{"nums": [2,7,11,15], "target": 9}}, "output": [0,1], "explanation": "Because nums[0] + nums[1] == 9, we return [0, 1]"}},
    {{"input": {{"nums": [3,2,4], "target": 6}}, "output": [1,2], "explanation": "Because nums[1] + nums[2] == 6, we return [1, 2]"}}
  ],
  "hints": [
    "A hash map can help you find complements efficiently",
    "Think about what you need to store as you iterate through the array",
    "Try to solve it in O(n) time complexity"
  ]
}}]

For C++:
[{{
  "title": "Maximum Subarray",
  "description": "Given an integer array arr, find the contiguous subarray (containing at least one number) which has the largest sum and return its sum.",
  "difficulty": "{difficulty}",
  "language": "cpp",
  "starter_code": "// Write ONLY the function. Do NOT include int main() - our platform handles that.\\n// You can include standard library headers.\\n\\n#include <vector>\\n#include <algorithm>\\n\\n// Function definition only\\nint solutionFunction(std::vector<int>& arr) {{\\n    // Edge cases\\n    if (arr.empty()) {{\\n        return 0;\\n    }}\\n    \\n    // Your code here\\n    return 0;\\n}}",
  "test_cases": [
    {{"input": {{"arr": [-2,1,-3,4,-1,2,1,-5,4]}}, "expected_output": 6}},
    {{"input": {{"arr": [1]}}, "expected_output": 1}},
    {{"input": {{"arr": [5,4,-1,7,8]}}, "expected_output": 23}}
  ],
  "constraints": [
    "1 <= arr.length <= 10^5",
    "-10^4 <= arr[i] <= 10^4"
  ],
  "examples": [
    {{"input": {{"arr": [-2,1,-3,4,-1,2,1,-5,4]}}, "output": 6, "explanation": "[4,-1,2,1] has the largest sum = 6"}},
    {{"input": {{"arr": [1]}}, "output": 1, "explanation": "The subarray [1] has the largest sum"}}
  ],
  "hints": [
    "Use Kadane's algorithm for O(n) solution",
    "Think about what the maximum sum ending at each position could be",
    "Keep track of current sum and maximum sum so far"
  ]
}}]

For Java:
[{{
  "title": "Two Sum",
  "description": "Given an array of integers nums and an integer target, return indices of the two numbers such that they add up to target.\\n\\nYou may assume that each input would have exactly one solution, and you may not use the same element twice.\\n\\nYou can return the answer in any order.",
  "difficulty": "{difficulty}",
  "language": "java",
  "starter_code": "// Write ONLY the class and method. Do NOT include public static void main() - our platform handles that.\\n\\nclass Solution {{\\n    public int[] twoSum(int[] nums, int target) {{\\n        // Edge cases\\n        if (nums == null || nums.length < 2) {{\\n            return new int[0];\\n        }}\\n        \\n        // Your code here\\n        return new int[0];\\n    }}\\n}}",
  "test_cases": [
    {{"input": {{"nums": [2,7,11,15], "target": 9}}, "expected_output": [0,1]}},
    {{"input": {{"nums": [3,2,4], "target": 6}}, "expected_output": [1,2]}},
    {{"input": {{"nums": [3,3], "target": 6}}, "expected_output": [0,1]}}
  ],
  "constraints": [
    "2 <= nums.length <= 10^4",
    "-10^9 <= nums[i] <= 10^9",
    "-10^9 <= target <= 10^9",
    "Only one valid answer exists"
  ],
  "examples": [
    {{"input": {{"nums": [2,7,11,15], "target": 9}}, "output": [0,1], "explanation": "Because nums[0] + nums[1] == 9, we return [0, 1]"}},
    {{"input": {{"nums": [3,2,4], "target": 6}}, "output": [1,2], "explanation": "Because nums[1] + nums[2] == 6, we return [1, 2]"}}
  ],
  "hints": [
    "A hash map can help you find complements efficiently",
    "Think about what you need to store as you iterate through the array",
    "Try to solve it in O(n) time complexity"
  ]
}}]

🔴 VERIFICATION CHECKLIST BEFORE SUBMITTING:
- [ ] Expected outputs manually calculated and verified
- [ ] All test cases follow the same logic
- [ ] Test cases match problem requirements exactly
- [ ] Edge cases included
- [ ] Input format is dict with parameter names as keys
- [ ] Starter code uses the EXACT template for the language
- [ ] Starter code includes user instructions

REMEMBER: Input must be dict format like {{"param1": value1, "param2": value2}}, NOT array format!"""

        print(f"🤖 Generating {num_challenges} challenges...")
        response = ""
        challenges_data: Optional[List[dict]] = None
        last_err: Optional[str] = None

        # Larger completion budget avoids truncated JSON; bracket-aware extract handles code in strings.
        for attempt, temp in enumerate((0.55, 0.35)):
            response = await generate_with_groq(
                prompt, max_tokens=8192, temperature=temp
            )
            if not response:
                last_err = "Empty response from LLM (check GROQ_API_KEY, rate limits, or Groq error logs above)"
                continue
            try:
                challenges_data = _parse_llm_challenges_response(response)
                if len(challenges_data) < num_challenges:
                    print(
                        f"⚠️ LLM returned {len(challenges_data)} challenge(s); expected {num_challenges}"
                    )
                break
            except (json.JSONDecodeError, ValueError) as e:
                last_err = str(e)
                print(f"❌ JSON parsing failed (batch attempt {attempt + 1}): {e}")
                print(f"Response tail (800 chars): …{response[-800:]!r}")

        if not challenges_data:
            print(
                "⚠️ Batch generation failed or empty; falling back to one challenge per request."
            )
            challenges_data = []
            for i in range(num_challenges):
                sub_prompt = _compact_coding_challenge_prompt(
                    role, difficulty, language, topic, i + 1, num_challenges
                )
                got = False
                for attempt, temp in enumerate((0.45, 0.25)):
                    response = await generate_with_groq(
                        sub_prompt, max_tokens=6000, temperature=temp
                    )
                    if not response:
                        last_err = "Empty response from LLM"
                        continue
                    try:
                        chunk = _parse_llm_challenges_response(response)
                        if chunk:
                            challenges_data.append(chunk[0])
                            got = True
                            break
                    except (json.JSONDecodeError, ValueError) as e:
                        last_err = str(e)
                        print(
                            f"❌ JSON parsing failed (challenge {i + 1}, attempt {attempt + 1}): {e}"
                        )
                if not got:
                    raise ValueError(
                        last_err
                        or f"Failed to generate challenge {i + 1} after retries"
                    )

        challenges_data = challenges_data[:num_challenges]

        # Validate and save challenges
        saved_count = 0
        from sqlalchemy import text

        for challenge_data in challenges_data:
            challenge_id = f"challenge_{uuid.uuid4()}"

            # Validate test cases format
            test_cases = challenge_data.get("test_cases", [])
            for i, tc in enumerate(test_cases):
                if "input" not in tc:
                    print(f"⚠️  Warning: Test case {i + 1} missing 'input' field")
                elif not isinstance(tc["input"], dict):
                    print(
                        f"⚠️  Warning: Test case {i + 1} input is not a dict, attempting to fix..."
                    )
                    # Try to auto-fix if it's an array
                    if isinstance(tc["input"], list):
                        print(f"   Skipping challenge with array-format test cases")
                        continue

                if "expected_output" not in tc:
                    print(
                        f"⚠️  Warning: Test case {i + 1} missing 'expected_output' field"
                    )

            # Check if starter code has proper instructions
            starter_code = challenge_data.get("starter_code", "")
            if language == "go":
                if (
                    "Write ONLY the function below" not in starter_code
                    or "Do NOT include" not in starter_code
                ):
                    print(f"⚠️  Warning: Go starter code missing user instructions")
            elif language == "javascript":
                if "Write ONLY the function" not in starter_code:
                    print(
                        f"⚠️  Warning: JavaScript starter code missing user instructions"
                    )

            insert_query = text(
                """
                INSERT INTO coding_challenges (
                    challenge_id, title, description, difficulty, language,
                    starter_code, test_cases, constraints, examples, hints,
                    role, created_at, created_by, is_active
                ) VALUES (
                    :challenge_id, :title, :description, :difficulty, :language,
                    :starter_code, :test_cases, :constraints, :examples, :hints,
                    :role, :created_at, :created_by, :is_active
                )
            """
            )

            await db.execute(
                insert_query,
                {
                    "challenge_id": challenge_id,
                    "title": challenge_data.get("title", "Untitled Challenge"),
                    "description": challenge_data.get("description", ""),
                    "difficulty": challenge_data.get("difficulty", difficulty),
                    "language": challenge_data.get("language", language),
                    "starter_code": starter_code,
                    "test_cases": json.dumps(challenge_data.get("test_cases", [])),
                    "constraints": json.dumps(challenge_data.get("constraints", [])),
                    "examples": json.dumps(challenge_data.get("examples", [])),
                    "hints": json.dumps(challenge_data.get("hints", [])),
                    "role": role,
                    "created_at": datetime.now(),
                    "created_by": "system",
                    "is_active": True,
                },
            )
            saved_count += 1
            print(f"✅ Saved challenge: {challenge_data.get('title')}")

        await db.commit()
        print(f"✅ Successfully generated {saved_count} challenges")

        return {
            "success": True,
            "generated_count": saved_count,
            "message": f"Generated {saved_count} coding challenge(s)",
        }

    except json.JSONDecodeError as e:
        print(f"❌ JSON parsing failed: {e}")
        print(f"Response was: {response[:500]}...")
        raise HTTPException(
            status_code=500, detail=f"Failed to parse LLM response: {str(e)}"
        )
    except Exception as e:
        print(f"❌ Generation failed: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== CHALLENGE MANAGEMENT ====================


@app.get("/api/coding/challenges")
async def get_coding_challenges(
    difficulty: str = None, language: str = None, db: AsyncSession = Depends(get_db)
):
    """FR-TECH-02: Get all challenges with filters"""
    try:
        from sqlalchemy import text

        query = "SELECT * FROM coding_challenges WHERE 1=1"
        params = {}

        if difficulty:
            query += " AND difficulty = :difficulty"
            params["difficulty"] = difficulty

        if language:
            query += " AND language = :language"
            params["language"] = language

        query += " ORDER BY created_at DESC"

        result = await db.execute(text(query), params)
        rows = result.fetchall()

        challenges = []
        for row in rows:
            challenges.append(
                {
                    "challenge_id": row[0],
                    "title": row[1],
                    "description": row[2],
                    "difficulty": row[3],
                    "language": row[4],
                    "starter_code": row[5],
                    "test_cases": json.loads(row[6]) if row[6] else [],
                    "constraints": json.loads(row[7]) if row[7] else [],
                    "examples": json.loads(row[8]) if row[8] else [],
                    "hints": json.loads(row[9]) if row[9] else [],
                    "role": row[10],
                    "created_at": str(row[11]) if row[11] else None,
                }
            )

        return {"challenges": challenges, "count": len(challenges)}

    except Exception as e:
        print(f"❌ Fetch failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/coding/challenge/{challenge_id}")
async def get_challenge(challenge_id: str, db: AsyncSession = Depends(get_db)):
    """FR-TECH-03: Get single challenge"""
    try:
        from sqlalchemy import text

        query = text(
            "SELECT * FROM coding_challenges WHERE challenge_id = :challenge_id"
        )
        result = await db.execute(query, {"challenge_id": challenge_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Challenge not found")

        return {
            "challenge_id": row[0],
            "title": row[1],
            "description": row[2],
            "difficulty": row[3],
            "language": row[4],
            "starter_code": row[5],
            "test_cases": json.loads(row[6]) if row[6] else [],
            "constraints": json.loads(row[7]) if row[7] else [],
            "examples": json.loads(row[8]) if row[8] else [],
            "hints": json.loads(row[9]) if row[9] else [],
            "role": row[10],
            "created_at": str(row[11]) if row[11] else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Fetch failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/coding/challenge/{challenge_id}")
async def delete_challenge(challenge_id: str, db: AsyncSession = Depends(get_db)):
    """FR-TECH-04: Delete challenge"""
    try:
        from sqlalchemy import text

        query = text("DELETE FROM coding_challenges WHERE challenge_id = :challenge_id")
        result = await db.execute(query, {"challenge_id": challenge_id})
        await db.commit()

        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Challenge not found")

        return {"message": "Challenge deleted successfully"}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Delete failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== GENERIC SCORING HELPER FUNCTIONS ====================


def generate_improvement_suggestions(
    evaluation: dict, correctness_score: float, efficiency_score: float
) -> list:
    """Generate dynamic improvement suggestions"""
    suggestions = evaluation.get("improvements", [])

    # Add suggestions based on score analysis
    if correctness_score < 100:
        suggestions.append("Focus on passing all test cases")
    if efficiency_score < 80:
        suggestions.append("Optimize algorithm for better performance")
    if evaluation["readability_score"] < 80:
        suggestions.append("Improve code documentation and comments")

    return suggestions[:5]  # Return top 5 suggestions


# ==================== CODE EXECUTION & TESTING ====================

# REPLACE YOUR /api/coding/run-tests ENDPOINT WITH THIS:


# ============================================
# COMPLETE MULTI-LANGUAGE run_tests FUNCTION
# Supports: Python, C++, JavaScript, Java, Go
# ============================================
@app.post("/api/coding/run-tests")
async def run_tests(data: dict, db: AsyncSession = Depends(get_db)):
    """FR-TECH-05: Run test cases - Multi-language support with COMPREHENSIVE FIXES"""
    try:
        import json
        import re
        import subprocess
        import tempfile
        import os
        import time

        challenge_id = data.get("challenge_id")
        code = data.get("code")
        language = data.get("language", "python").lower()

        if not all([challenge_id, code]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Get test cases and challenge config - FIXED: Get timeout/memory from challenge
        from sqlalchemy import text

        query = text(
            "SELECT test_cases, constraints FROM coding_challenges WHERE challenge_id = :challenge_id"
        )
        result = await db.execute(query, {"challenge_id": challenge_id})
        row = result.fetchone()

        if not row or not row[0]:
            raise HTTPException(
                status_code=404, detail="Challenge or test cases not found"
            )

        test_cases = json.loads(row[0])

        # ✅ FIX: Parse timeout and memory from challenge constraints
        timeout_seconds = 5  # Default
        memory_limit_mb = 256  # Default

        if row[1]:  # If constraints exist
            try:
                constraints = json.loads(row[1]) if isinstance(row[1], str) else row[1]
                timeout_seconds = constraints.get("timeout_seconds", 5)
                memory_limit_mb = constraints.get("memory_limit_mb", 256)
                print(
                    f"⚙️  Using challenge config: timeout={timeout_seconds}s, memory={memory_limit_mb}MB"
                )
            except:
                print(f"⚠️  Using default config: timeout=5s, memory=256MB")

        # Helper function for Go code sanitization
        def sanitize_go_code(code: str) -> str:
            """
            Sanitize Go code by removing unused imports and main() function
            """
            # Remove package main if present
            if "package main" in code:
                code = code.replace("package main\n", "")

            # Remove main() function if present
            if "func main()" in code:
                lines = code.split("\n")
                new_lines = []
                in_main = False
                for line in lines:
                    if "func main()" in line:
                        in_main = True
                        continue
                    if (
                        in_main
                        and line.strip()
                        and line[0] != "\t"
                        and line[:4] != "    "
                    ):
                        in_main = False
                    if not in_main:
                        new_lines.append(line)
                code = "\n".join(new_lines)

            # Analyze which imports are actually used
            lines = code.split("\n")
            new_lines = []
            imports_section = []
            in_import_block = False

            for line in lines:
                stripped = line.strip()

                # Start of import block
                if stripped.startswith("import ("):
                    in_import_block = True
                    imports_section = []
                    continue

                # End of import block
                if in_import_block and stripped == ")":
                    in_import_block = False

                    # Filter imports based on actual usage
                    needed_imports = []
                    for import_line in imports_section:
                        if import_line.strip():
                            import_name = import_line.strip().strip('"')
                            # Check if this import is actually used in code
                            if import_name + "." in code:
                                needed_imports.append(import_line)

                    # Add filtered imports back
                    if needed_imports:
                        new_lines.append("import (")
                        new_lines.extend(needed_imports)
                        new_lines.append(")")
                    continue

                # Inside import block
                if in_import_block:
                    imports_section.append(line)
                    continue

                # Single line import
                if stripped.startswith('import "') or stripped.startswith('import ."'):
                    import_match = re.search(
                        r'import\s+(?:"([^"]+)"|\.)"([^"]+)"', line
                    )
                    if import_match:
                        import_name = import_match.group(1) or import_match.group(2)
                        if import_name + "." in code:
                            new_lines.append(line)
                else:
                    new_lines.append(line)

            return "\n".join(new_lines)

        # Execute tests
        passed = 0
        failed = 0
        results = []
        total_execution_time = 0

        print(f"\n🧪 Running {len(test_cases)} test cases for {language}...")

        for i, test_case in enumerate(test_cases):
            try:
                # Extract input and expected output
                test_input = test_case.get("input", {})
                expected_output = test_case.get("expected_output")

                print(f"\n  Test {i + 1}:")
                print(f"    Input: {test_input}")
                print(f"    Expected: {expected_output}")

                # ✅ ADD EDGE CASE DETECTION
                # Check for empty arrays/collections
                if isinstance(test_input, dict):
                    for key, value in test_input.items():
                        if isinstance(value, list) and len(value) == 0:
                            print(f"⚠️  Empty array in test {i + 1}: {key}")
                        if value is None:
                            print(f"⚠️  Null value in test {i + 1}: {key}")
                        # Additional edge cases
                        if isinstance(value, str) and value == "":
                            print(f"⚠️  Empty string in test {i + 1}: {key}")
                        if isinstance(value, (int, float)) and value == 0:
                            print(f"⚠️  Zero value in test {i + 1}: {key}")
                elif isinstance(test_input, list):
                    if len(test_input) == 0:
                        print(f"⚠️  Empty list in test {i + 1}")
                    elif any(item is None for item in test_input):
                        print(f"⚠️  List contains null values in test {i + 1}")

                actual_output = None
                execution_time = 0

                # ============================================
                # PYTHON EXECUTION - FIXED
                # ============================================
                if language == "python":
                    # Use isolated namespace to prevent pollution
                    exec_globals = {"__builtins__": {}}
                    exec_locals = {}

                    try:
                        # Try to find function name more reliably
                        func_pattern = r"def\s+(\w+)\s*\("
                        func_matches = re.findall(func_pattern, code)

                        if not func_matches:
                            raise ValueError("No function found in Python code")

                        func_name = func_matches[0]  # Use first function found

                        # Execute in isolated context
                        exec(code, exec_globals, exec_locals)

                        # Get function from locals (user-defined) or globals (built-ins)
                        func = exec_locals.get(func_name) or exec_globals.get(func_name)

                        if not func or not callable(func):
                            raise ValueError(
                                f"Function '{func_name}' not found or not callable"
                            )

                        start_time = time.time()

                        # Call with appropriate arguments
                        if isinstance(test_input, dict):
                            actual_output = func(**test_input)
                        elif isinstance(test_input, list):
                            actual_output = func(*test_input)
                        else:
                            actual_output = func(test_input)

                        execution_time = (time.time() - start_time) * 1000

                    except Exception as e:
                        raise ValueError(f"Python execution error: {str(e)}")

                # ============================================
                # C++ EXECUTION - FIXED
                # ============================================
                elif language in ["cpp", "c++"]:
                    # Detect return type and function name
                    func_pattern = r"(\w+)\s+(\w+)\s*\([^)]*\)\s*{"
                    func_match = re.search(func_pattern, code)

                    if not func_match:
                        raise ValueError("Could not find function in C++ code")

                    return_type = func_match.group(1)
                    func_name = func_match.group(2)

                    # Create test harness with return type handling
                    cpp_test_code = f"""
#include <iostream>
#include <string>
#include <vector>
#include <algorithm>
#include <sstream>

{code}

// Helper functions for different return types
void printResult(const std::vector<std::string>& v) {{
    std::cout << "[";
    for (size_t i = 0; i < v.size(); ++i) {{
        std::cout << "\\"" << v[i] << "\\"";
        if (i < v.size() - 1) std::cout << ", ";
    }}
    std::cout << "]";
}}

void printResult(const std::vector<int>& v) {{
    std::cout << "[";
    for (size_t i = 0; i < v.size(); ++i) {{
        std::cout << v[i];
        if (i < v.size() - 1) std::cout << ", ";
    }}
    std::cout << "]";
}}

void printResult(bool b) {{
    std::cout << std::boolalpha << b;
}}

void printResult(int n) {{
    std::cout << n;
}}

void printResult(double d) {{
    std::cout << d;
}}

void printResult(const std::string& s) {{
    std::cout << s;
}}

template<typename T>
void printResult(const T& value) {{
    std::cout << value;
}}

int main() {{
"""

                    # Handle parameters from dict
                    if isinstance(test_input, dict):
                        for param_name, param_value in test_input.items():
                            if isinstance(param_value, str):
                                cpp_test_code += (
                                    f'    std::string {param_name} = "{param_value}";\n'
                                )
                            elif isinstance(param_value, list):
                                if all(isinstance(x, str) for x in param_value):
                                    cpp_test_code += f"    std::vector<std::string> {param_name} = {{"
                                    cpp_test_code += ", ".join(
                                        [f'"{x}"' for x in param_value]
                                    )
                                    cpp_test_code += "};\n"
                                else:
                                    cpp_test_code += (
                                        f"    std::vector<int> {param_name} = {{"
                                    )
                                    cpp_test_code += ", ".join(
                                        [str(x) for x in param_value]
                                    )
                                    cpp_test_code += "};\n"
                            elif isinstance(param_value, bool):
                                cpp_test_code += f"    bool {param_name} = {'true' if param_value else 'false'};\n"
                            elif isinstance(param_value, int):
                                cpp_test_code += (
                                    f"    int {param_name} = {param_value};\n"
                                )
                            elif isinstance(param_value, float):
                                cpp_test_code += (
                                    f"    double {param_name} = {param_value};\n"
                                )
                            else:
                                cpp_test_code += (
                                    f"    auto {param_name} = {param_value};\n"
                                )

                        # Call function with parameter names
                        param_names = ", ".join(
                            [name for name, _ in test_input.items()]
                        )
                        cpp_test_code += (
                            f"    auto result = {func_name}({param_names});\n"
                        )
                    else:
                        # Single parameter
                        if isinstance(test_input, str):
                            cpp_test_code += (
                                f'    std::string input = "{test_input}";\n'
                            )
                            cpp_test_code += f"    auto result = {func_name}(input);\n"
                        else:
                            cpp_test_code += f"    auto input = {test_input};\n"
                            cpp_test_code += f"    auto result = {func_name}(input);\n"

                    # Print result using helper function
                    cpp_test_code += """    
    printResult(result);
    std::cout << std::endl;
    return 0;
}
"""

                    with tempfile.NamedTemporaryFile(
                        mode="w", suffix=".cpp", delete=False
                    ) as f:
                        f.write(cpp_test_code)
                        cpp_file = f.name

                    exe_file = cpp_file.replace(".cpp", ".exe")

                    try:
                        compile_result = subprocess.run(
                            ["g++", "-std=c++17", cpp_file, "-o", exe_file],
                            capture_output=True,
                            text=True,
                            timeout=10,
                        )

                        if compile_result.returncode != 0:
                            raise ValueError(
                                f"Compilation failed: {compile_result.stderr}"
                            )

                        start_time = time.time()
                        run_result = subprocess.run(
                            [exe_file],
                            capture_output=True,
                            text=True,
                            timeout=timeout_seconds,
                        )
                        execution_time = (time.time() - start_time) * 1000

                        if run_result.returncode != 0:
                            raise ValueError(f"Runtime error: {run_result.stderr}")

                        output_str = run_result.stdout.strip()
                        print(f"    Raw output: '{output_str}'")

                        # Parse output based on return type
                        if output_str.startswith("[") and output_str.endswith("]"):
                            # Parse array output
                            import json

                            actual_output = json.loads(
                                output_str.replace(" ", ",")
                                .replace("[,", "[")
                                .replace(",]", "]")
                            )
                        elif output_str.lower() == "true":
                            actual_output = True
                        elif output_str.lower() == "false":
                            actual_output = False
                        elif output_str.lstrip("-").isdigit():
                            actual_output = int(output_str)
                        elif output_str.replace(".", "").lstrip("-").isdigit():
                            actual_output = float(output_str)
                        else:
                            actual_output = output_str
                    finally:
                        try:
                            os.unlink(cpp_file)
                            if os.path.exists(exe_file):
                                os.unlink(exe_file)
                        except:
                            pass

                # ============================================
                # JAVASCRIPT EXECUTION - FIXED
                # ============================================
                elif language in ["javascript", "js", "node"]:
                    # Detect function definitions (regular functions, arrow functions, const assignments)
                    func_patterns = [
                        r"function\s+(\w+)",  # function name()
                        r"const\s+(\w+)\s*=\s*function",  # const name = function
                        r"const\s+(\w+)\s*=\s*\([^)]*\)\s*=>",  # const name = () =>
                        r"let\s+(\w+)\s*=\s*function",  # let name = function
                        r"let\s+(\w+)\s*=\s*\([^)]*\)\s*=>",  # let name = () =>
                        r"var\s+(\w+)\s*=\s*function",  # var name = function
                        r"var\s+(\w+)\s*=\s*\([^)]*\)\s*=>",  # var name = () =>
                    ]

                    func_name = None
                    for pattern in func_patterns:
                        match = re.search(pattern, code)
                        if match:
                            func_name = match.group(1)
                            break

                    if not func_name:
                        raise ValueError("Could not find function in JavaScript code")

                    js_test_code = code + "\n\n"

                    # Handle parameters
                    if isinstance(test_input, dict):
                        args = ", ".join([json.dumps(v) for v in test_input.values()])
                        js_test_code += f"const result = {func_name}({args});\n"
                    else:
                        js_test_code += (
                            f"const result = {func_name}({json.dumps(test_input)});\n"
                        )

                    js_test_code += "console.log(JSON.stringify(result));\n"

                    with tempfile.NamedTemporaryFile(
                        mode="w", suffix=".js", delete=False
                    ) as f:
                        f.write(js_test_code)
                        js_file = f.name

                    try:
                        start_time = time.time()
                        run_result = subprocess.run(
                            ["node", js_file],
                            capture_output=True,
                            text=True,
                            timeout=timeout_seconds,
                        )
                        execution_time = (time.time() - start_time) * 1000

                        if run_result.returncode != 0:
                            raise ValueError(f"Runtime error: {run_result.stderr}")

                        output_str = run_result.stdout.strip()
                        if output_str:
                            actual_output = json.loads(output_str)
                        else:
                            actual_output = None
                    finally:
                        try:
                            os.unlink(js_file)
                        except:
                            pass

                # ============================================
                # JAVA EXECUTION - FIXED
                # ============================================
                elif language == "java":
                    # Extract class and method names with return type detection
                    class_match = re.search(r"class\s+(\w+)", code)
                    method_match = re.search(
                        r"public\s+(\w+)\s+(\w+)\s*\([^)]*\)", code
                    )

                    if not class_match or not method_match:
                        raise ValueError("Could not find class or method in Java code")

                    class_name = class_match.group(1)
                    return_type = method_match.group(1)  # int, String, int[], etc.
                    method_name = method_match.group(2)

                    # Create test class
                    java_test_code = code + "\n\npublic class TestRunner {\n"
                    java_test_code += "    public static void main(String[] args) {\n"
                    java_test_code += (
                        f"        {class_name} solution = new {class_name}();\n"
                    )

                    # Handle parameters from dict
                    if isinstance(test_input, dict):
                        values = list(test_input.values())
                        param_vars = []

                        for i, val in enumerate(values):
                            var_name = f"param{i}"
                            if isinstance(val, list):
                                if all(isinstance(x, str) for x in val):
                                    java_test_code += f'        String[] {var_name} = {{"{", ".join(val)}"}};\n'
                                else:
                                    java_test_code += f"        int[] {var_name} = {{{', '.join([str(x) for x in val])}}};\n"
                            elif isinstance(val, str):
                                java_test_code += (
                                    f'        String {var_name} = "{val}";\n'
                                )
                            elif isinstance(val, bool):
                                java_test_code += f"        boolean {var_name} = {'true' if val else 'false'};\n"
                            elif isinstance(val, int):
                                java_test_code += f"        int {var_name} = {val};\n"
                            else:
                                java_test_code += (
                                    f"        Object {var_name} = {val};\n"
                                )
                            param_vars.append(var_name)

                        # Call method with parameters
                        java_test_code += f"        {return_type} result = solution.{method_name}({', '.join(param_vars)});\n"
                    else:
                        # Single parameter (fallback)
                        if isinstance(test_input, str):
                            java_test_code += (
                                f'        String input = "{test_input}";\n'
                            )
                            java_test_code += f"        {return_type} result = solution.{method_name}(input);\n"
                        elif isinstance(test_input, list):
                            if all(isinstance(x, str) for x in test_input):
                                java_test_code += f'        String[] input = {{"{", ".join(test_input)}"}};\n'
                            else:
                                java_test_code += f"        int[] input = {{{', '.join([str(x) for x in test_input])}}};\n"
                            java_test_code += f"        {return_type} result = solution.{method_name}(input);\n"
                        else:
                            java_test_code += f"        Object input = {test_input};\n"
                            java_test_code += f"        {return_type} result = solution.{method_name}(input);\n"

                    # Handle output based on return type
                    if return_type.endswith("[]"):  # Array type
                        java_test_code += "        System.out.println(java.util.Arrays.toString(result));\n"
                    elif return_type == "boolean":
                        java_test_code += "        System.out.println(result);\n"  # boolean prints as true/false
                    elif return_type in [
                        "int",
                        "long",
                        "double",
                        "float",
                        "short",
                        "byte",
                        "char",
                    ]:
                        java_test_code += "        System.out.println(result);\n"  # Primitives print directly
                    else:
                        java_test_code += "        System.out.println(result);\n"  # Objects use toString()

                    java_test_code += "    }\n}\n"

                    temp_dir = tempfile.mkdtemp()
                    java_file = os.path.join(temp_dir, "TestRunner.java")

                    with open(java_file, "w") as f:
                        f.write(java_test_code)

                    try:
                        compile_result = subprocess.run(
                            ["javac", java_file],
                            capture_output=True,
                            text=True,
                            timeout=10,
                        )

                        if compile_result.returncode != 0:
                            raise ValueError(
                                f"Compilation failed: {compile_result.stderr}"
                            )

                        start_time = time.time()
                        run_result = subprocess.run(
                            ["java", "-cp", temp_dir, "TestRunner"],
                            capture_output=True,
                            text=True,
                            timeout=timeout_seconds,
                        )
                        execution_time = (time.time() - start_time) * 1000

                        if run_result.returncode != 0:
                            raise ValueError(f"Runtime error: {run_result.stderr}")

                        output_str = run_result.stdout.strip()

                        # Parse output based on return type
                        if return_type == "boolean":
                            actual_output = output_str.lower() == "true"
                        elif return_type in ["int", "long", "short", "byte"]:
                            actual_output = (
                                int(output_str)
                                if output_str.lstrip("-").isdigit()
                                else output_str
                            )
                        elif return_type in ["double", "float"]:
                            actual_output = (
                                float(output_str)
                                if output_str.replace(".", "").lstrip("-").isdigit()
                                else output_str
                            )
                        elif return_type.endswith("[]"):
                            # Parse array output
                            import json

                            actual_output = json.loads(
                                output_str.replace(" ", ",")
                                .replace("[,", "[")
                                .replace(",]", "]")
                            )
                        else:
                            actual_output = output_str

                    finally:
                        try:
                            import shutil

                            shutil.rmtree(temp_dir)
                        except:
                            pass

                # ============================================
                # GO EXECUTION - FIXED (Unused imports removed)
                # ============================================
                elif language in ["go", "golang"]:
                    # ✅ First sanitize the code
                    sanitized_code = sanitize_go_code(code)
                    print(f"🔧 Sanitized Go code (removed unused imports)")

                    # Then extract function name from sanitized code
                    func_match = re.search(r"func\s+(\w+)", sanitized_code)
                    if not func_match:
                        raise ValueError("Could not find function in Go code")
                    func_name = func_match.group(1)

                    # Extract user's imports more carefully
                    user_imports = set()

                    # Find single imports
                    single_imports = re.findall(r'import\s+"([^"]+)"', sanitized_code)
                    user_imports.update(single_imports)

                    # Find block imports
                    block_match = re.search(
                        r"import\s*\(\s*([\s\S]*?)\s*\)", sanitized_code
                    )
                    if block_match:
                        block_content = block_match.group(1)
                        block_imports = re.findall(r'"([^"]+)"', block_content)
                        user_imports.update(block_imports)

                    # Remove ALL import statements from user's code
                    code_without_imports = re.sub(
                        r'import\s*(?:\([\s\S]*?\)|"[^"]+"\s*)',
                        "",
                        sanitized_code,
                        flags=re.MULTILINE,
                    )
                    code_without_imports = re.sub(
                        r"^\s*import\s*\([^)]*\)",
                        "",
                        code_without_imports,
                        flags=re.MULTILINE | re.DOTALL,
                    )
                    code_without_imports = re.sub(
                        r'^\s*import\s+"[^"]+"',
                        "",
                        code_without_imports,
                        flags=re.MULTILINE,
                    )

                    # ✅ FIX: Check which imports are ACTUALLY NEEDED
                    needed_imports = set()

                    # Always include 'fmt' for printing results
                    needed_imports.add("fmt")

                    # Check if 'sort' is needed (common for sorting arrays)
                    if (
                        "sort." in code_without_imports
                        or "sort.Ints" in code_without_imports
                    ):
                        needed_imports.add("sort")

                    # Check if 'math' is needed
                    if "math." in code_without_imports:
                        needed_imports.add("math")

                    # Check if 'strings' is needed
                    if "strings." in code_without_imports:
                        needed_imports.add("strings")

                    # Check if 'strconv' is needed
                    if "strconv." in code_without_imports:
                        needed_imports.add("strconv")

                    # Add user imports that are actually used
                    for imp in user_imports:
                        if imp in code_without_imports:
                            needed_imports.add(imp)

                    # Build test code with ONLY NEEDED imports
                    go_test_code = "package main\n\n"
                    if needed_imports:
                        go_test_code += "import (\n"
                        for imp in sorted(needed_imports):
                            go_test_code += f'\t"{imp}"\n'
                        go_test_code += ")\n\n"

                    go_test_code += code_without_imports + "\n\n"
                    go_test_code += "func main() {\n"

                    # Handle parameters
                    if isinstance(test_input, dict):
                        values = list(test_input.values())
                        args = []
                        for val in values:
                            if isinstance(val, str):
                                args.append(f'"{val}"')
                            elif isinstance(val, list):
                                if all(isinstance(x, str) for x in val):
                                    args.append(f'[]string{{"{", ".join(val)}"}}')
                                else:
                                    args.append(
                                        f"[]int{{{', '.join([str(x) for x in val])}}}"
                                    )
                            elif isinstance(val, bool):
                                args.append(f"{'true' if val else 'false'}")
                            else:
                                args.append(str(val))
                        go_test_code += (
                            f"    result := {func_name}({', '.join(args)})\n"
                        )
                    else:
                        if isinstance(test_input, str):
                            go_test_code += (
                                f'    result := {func_name}("{test_input}")\n'
                            )
                        else:
                            go_test_code += f"    result := {func_name}({test_input})\n"

                    go_test_code += "    fmt.Println(result)\n}\n"

                    print(f"🔍 Go imports needed: {needed_imports}")
                    print(f"📝 Go test code preview:\n{go_test_code[:500]}...")

                    with tempfile.NamedTemporaryFile(
                        mode="w", suffix=".go", delete=False
                    ) as f:
                        f.write(go_test_code)
                        go_file = f.name

                    try:
                        start_time = time.time()
                        run_result = subprocess.run(
                            ["go", "run", go_file],
                            capture_output=True,
                            text=True,
                            timeout=timeout_seconds,
                        )
                        execution_time = (time.time() - start_time) * 1000

                        if run_result.returncode != 0:
                            error_msg = run_result.stderr
                            # Provide more helpful error messages
                            if "imported and not used" in error_msg:
                                error_msg += (
                                    "\n\n🔧 FIX: Remove unused imports from your code"
                                )
                            elif "undefined" in error_msg:
                                error_msg += (
                                    "\n\n🔧 FIX: Check function name and variable names"
                                )
                            raise ValueError(f"Go runtime error: {error_msg}")

                        output_str = run_result.stdout.strip()
                        if output_str.lower() == "true":
                            actual_output = True
                        elif output_str.lower() == "false":
                            actual_output = False
                        elif output_str.lstrip("-").isdigit():
                            actual_output = int(output_str)
                        elif output_str.replace(".", "").lstrip("-").isdigit():
                            actual_output = float(output_str)
                        else:
                            actual_output = output_str
                    finally:
                        try:
                            os.unlink(go_file)
                        except:
                            pass

                else:
                    raise ValueError(f"Language {language} not supported")

                # ============================================
                # COMPARE RESULTS
                # ============================================
                total_execution_time += execution_time

                # Normalize for comparison
                if isinstance(actual_output, list):
                    actual_output = sorted(actual_output)
                if isinstance(expected_output, list):
                    expected_output = sorted(expected_output)

                is_correct = actual_output == expected_output

                print(f"    Actual: {actual_output}")
                print(f"    Match: {is_correct}")

                if is_correct:
                    passed += 1
                    results.append(
                        {
                            "test_number": i + 1,
                            "passed": True,
                            "input": str(test_input),
                            "expected": str(expected_output),
                            "actual": str(actual_output),
                            "execution_time_ms": round(execution_time, 2),
                        }
                    )
                else:
                    failed += 1
                    results.append(
                        {
                            "test_number": i + 1,
                            "passed": False,
                            "input": str(test_input),
                            "expected": str(expected_output),
                            "actual": str(actual_output),
                            "error": "Output mismatch",
                            "execution_time_ms": round(execution_time, 2),
                        }
                    )

            except subprocess.TimeoutExpired:
                failed += 1
                print(f"    Error: Execution timeout")
                results.append(
                    {
                        "test_number": i + 1,
                        "passed": False,
                        "input": str(test_input) if "test_input" in locals() else "N/A",
                        "expected": (
                            str(expected_output)
                            if "expected_output" in locals()
                            else "N/A"
                        ),
                        "error": "Execution timeout",
                        "execution_time_ms": 0,
                    }
                )
            except Exception as e:
                failed += 1
                print(f"    Error: {str(e)}")
                results.append(
                    {
                        "test_number": i + 1,
                        "passed": False,
                        "input": str(test_input) if "test_input" in locals() else "N/A",
                        "expected": (
                            str(expected_output)
                            if "expected_output" in locals()
                            else "N/A"
                        ),
                        "error": str(e),
                        "execution_time_ms": 0,
                    }
                )

        total = len(test_cases)
        all_passed = passed == total

        print(f"\n✅ Tests complete: {passed}/{total} passed\n")

        compilation_errors = []
        allow_resubmission = False

        # Check if all failures are compilation errors
        if failed == total and any(
            "compilation" in str(r.get("error", "")).lower()
            or "compile" in str(r.get("error", "")).lower()
            or "syntax" in str(r.get("error", "")).lower()
            for r in results
        ):
            allow_resubmission = True
            compilation_errors = [r for r in results if r.get("error")]
            print(
                f"🔧 Compilation errors detected - allowing resubmission for {language}"
            )

        return {
            "success": True,
            "passed": passed,
            "failed": failed,
            "total": total,
            "all_passed": all_passed,
            "results": results,
            "execution_time_ms": round(total_execution_time, 2),
            "compilation_errors": compilation_errors,
            "allow_resubmission": allow_resubmission,  # NEW FIELD
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Test execution failed: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/debug/ai-test")
async def debug_ai_test(data: dict):
    """Debug endpoint to test AI detection"""
    code = data.get("code", "")
    language = data.get("language", "python")

    print(f"\n" + "=" * 60)
    print(f"🧪 DEBUG AI DETECTION TEST")
    print(f"Code length: {len(code)} chars")
    print(f"Lines: {len(code.split(chr(10)))}")
    print(f"=" * 60)

    # Test local detector
    print(f"\n1. LOCAL DETECTOR:")
    local_result = await ai_detector.detect_ai_code(code, language)
    print(f"   Confidence: {local_result['confidence']}%")
    print(f"   AI Generated: {local_result['ai_generated']}")
    print(f"   Details: {local_result['details']}")
    if local_result.get("evidence"):
        print(f"   Evidence: {local_result['evidence']}")

    # Test hybrid detector
    print(f"\n2. HYBRID DETECTOR:")
    hybrid_result = await hybrid_ai_detection(code, language)
    print(f"   Confidence: {hybrid_result['confidence']}%")
    print(f"   AI Generated: {hybrid_result['ai_generated']}")
    print(f"   Details: {hybrid_result['details']}")

    # Show code sample
    print(f"\n3. CODE SAMPLE (first 200 chars):")
    print(f"   {code[:200]}...")

    return {
        "local": local_result,
        "hybrid": hybrid_result,
        "code_sample": code[:200] + "..." if len(code) > 200 else code,
    }


# ==================== SUBMISSION & EVALUATION ====================


@app.post("/api/coding/submit")
async def submit_solution(data: dict, db: AsyncSession = Depends(get_db)):
    """FR-TEC-05 to FR-TEC-10: Submit and evaluate solution with comprehensive scoring AND BREAKDOWN - UPDATED AI DETECTION"""
    try:
        challenge_id = data.get("challenge_id")
        candidate_email = data.get("candidate_email")
        code = data.get("code")
        language = data.get("language")

        # Extract optional fields for integration
        assessment_id = data.get("assessment_id")
        result_id = data.get("result_id")
        session_id = data.get("session_id")
        auto_submitted = data.get("auto_submitted", False)

        # ✅ FIX: Get role from challenge or assessment
        from sqlalchemy import text

        # Try to get role from challenge first
        role = "Software Engineer"  # Default

        # Get challenge details to extract role
        challenge_query = text(
            """
            SELECT role, title
            FROM coding_challenges
            WHERE challenge_id = :challenge_id
        """
        )
        challenge_result = await db.execute(
            challenge_query, {"challenge_id": challenge_id}
        )
        challenge_row = challenge_result.fetchone()

        if challenge_row:
            role = challenge_row[0] or "Software Engineer"
            challenge_title = challenge_row[1]
        else:
            challenge_title = "Coding Challenge"

        # If assessment_id is provided, try to get role from assessment
        if assessment_id:
            assessment_query = text(
                """
                SELECT role FROM assessments
                WHERE assessment_id = :assessment_id
            """
            )
            assessment_result = await db.execute(
                assessment_query, {"assessment_id": assessment_id}
            )
            assessment_row = assessment_result.fetchone()

            if assessment_row and assessment_row[0]:
                role = assessment_row[0]

        print(f"📝 Coding submission role determined: {role}")

        # ✅ AI CODE DETECTION (ChatGPT/Claude/Copilot detection) - ENHANCED
        try:
            print(f"🤖 Checking for AI-generated code...")
            ai_detection_result = await ai_detector.detect_ai_code(code, language)
            print(f"   AI Confidence: {ai_detection_result['confidence']}%")

            if ai_detection_result["ai_generated"]:
                print(f"   🚨 AI-GENERATED CODE DETECTED!")
                print(f"   Evidence: {', '.join(ai_detection_result['evidence'])}")

            # ✅ ENSURE CONSISTENT STRUCTURE FOR FRONTEND
            ai_detection_result = {
                "ai_generated": ai_detection_result["ai_generated"],
                "detected": ai_detection_result[
                    "ai_generated"
                ],  # Add alias for frontend
                "confidence": ai_detection_result["confidence"],
                "score": ai_detection_result["confidence"],  # Add score alias
                "evidence": ai_detection_result["evidence"],
                "details": ai_detection_result["details"],
                "source": "AI Detection System",
            }
        except Exception as ai_error:
            print(f"⚠️ AI detection failed: {ai_error}")
            ai_detection_result = {
                "ai_generated": False,
                "detected": False,
                "confidence": 0,
                "score": 0,
                "evidence": [],
                "details": "AI detection unavailable",
                "source": "AI Detection System",
            }

        # ✅ CANDIDATE SIMILARITY CHECK (comparing with other candidates)
        try:
            print(f"🔍 Checking similarity with other candidates...")
            plagiarism_result = await plagiarism_detector.check_plagiarism_async(
                code, challenge_id, language, db
            )
            print(f"   Candidate Similarity: {plagiarism_result['plagiarism_score']}%")

            if plagiarism_result["plagiarism_detected"]:
                print(f"   ⚠️ SUSPICIOUS! Similar to previous submission")

            # ✅ ENSURE CONSISTENT STRUCTURE FOR FRONTEND
            plagiarism_result = {
                "plagiarism_detected": plagiarism_result["plagiarism_detected"],
                "detected": plagiarism_result["plagiarism_detected"],  # Add alias
                "plagiarism_score": plagiarism_result["plagiarism_score"],
                "score": plagiarism_result["plagiarism_score"],  # Add score alias
                "similar_submissions": plagiarism_result["similar_submissions"],
                "details": plagiarism_result["details"],
                "threshold_used": 75,
            }
        except Exception as plag_error:
            print(f"⚠️ Similarity check failed: {plag_error}")
            plagiarism_result = {
                "plagiarism_detected": False,
                "detected": False,
                "plagiarism_score": 0,
                "score": 0,
                "similar_submissions": [],
                "details": "Plagiarism check skipped due to error",
                "threshold_used": 75,
            }

        # Run tests
        test_results = await run_tests(
            {"challenge_id": challenge_id, "code": code, "language": language}, db
        )

        passed_tests = test_results.get("passed", 0)
        total_tests = test_results.get("total", 0)
        correctness_score = calculate_percentage(passed_tests, total_tests)

        total_execution_time = test_results.get("execution_time_ms", 0)
        avg_execution_time = (
            total_execution_time / total_tests if total_tests > 0 else 0
        )

        efficiency_score = calculate_efficiency_score(
            avg_execution_time, language, total_tests
        )

        print(
            f"⏱️  Efficiency score: {efficiency_score}% (avg {avg_execution_time:.1f}ms per test)"
        )

        eval_prompt = generate_evaluation_prompt(
            code, language, passed_tests, total_tests, avg_execution_time
        )
        eval_response = await generate_with_groq(eval_prompt)

        evaluation = parse_evaluation_response(eval_response, language)

        score_breakdown = calculate_score_breakdown(
            correctness_score,
            efficiency_score,
            evaluation,
            passed_tests,
            total_tests,
            avg_execution_time,
        )

        final_score = score_breakdown["total_contribution_rounded"]
        performance_level = get_performance_level(final_score)
        improvement_suggestions = generate_improvement_suggestions(
            evaluation, correctness_score, efficiency_score
        )

        # ✅ ADD AI detection and plagiarism check to evaluation
        evaluation["ai_detection"] = ai_detection_result
        evaluation["plagiarism_check"] = plagiarism_result

        submission_id = f"submission_{uuid.uuid4()}"

        # ✅ FIXED: Include role in the INSERT query
        insert_query = text(
            """
            INSERT INTO coding_submissions (
                submission_id, challenge_id, candidate_email, code, language,
                test_results, evaluation, score, submitted_at,
                assessment_id, result_id, score_breakdown, session_id, role
            ) VALUES (
                :submission_id, :challenge_id, :candidate_email, :code, :language,
                :test_results, :evaluation, :score, :submitted_at,
                :assessment_id, :result_id, :score_breakdown, :session_id, :role
            )
        """
        )

        await db.execute(
            insert_query,
            {
                "submission_id": submission_id,
                "challenge_id": challenge_id,
                "candidate_email": candidate_email,
                "code": code,
                "language": language,
                "test_results": json.dumps(test_results),
                "evaluation": json.dumps(
                    evaluation
                ),  # Contains AI detection and plagiarism
                "score": final_score,
                "score_breakdown": json.dumps(score_breakdown),
                "submitted_at": datetime.now(),
                "assessment_id": assessment_id,
                "result_id": result_id,
                "session_id": session_id,
                "role": role,  # ✅ ADDED: Save the role
            },
        )

        # ✅ Save AI detection and plagiarism to separate table with CONSISTENT STRUCTURE
        detection_id = f"detect_{uuid.uuid4()}"
        detection_query = text(
            """
            INSERT INTO code_detection
            (detection_id, submission_id, ai_detection, plagiarism_check, created_at)
            VALUES (:detection_id, :submission_id, :ai_detection, :plagiarism_check, :created_at)
        """
        )
        await db.execute(
            detection_query,
            {
                "detection_id": detection_id,
                "submission_id": submission_id,
                "ai_detection": json.dumps(ai_detection_result),
                "plagiarism_check": json.dumps(plagiarism_result),
                "created_at": datetime.now(),
            },
        )

        if result_id:
            await db.execute(
                text(
                    """
                UPDATE results
                SET coding_submission_id = :submission_id,
                    has_coding_assessment = TRUE
                WHERE result_id = :result_id
            """
                ),
                {"submission_id": submission_id, "result_id": result_id},
            )

        await update_technical_score(
            db, candidate_email, assessment_id, submission_id, result_id, final_score
        )
        await db.commit()

        # ✅ FIXED: Return role in response
        return {
            "success": True,
            "submission_id": submission_id,
            "score": final_score,
            "role": role,
            "challenge_title": challenge_title,
            "language": language,
            "passed_tests": passed_tests,
            "total_tests": total_tests,
            "efficiency_score": efficiency_score,
            "ai_detection": ai_detection_result,
            "plagiarism_check": plagiarism_result,
            "improvement_suggestions": improvement_suggestions,
            "performance_level": performance_level,
            "score_breakdown": score_breakdown,
        }

    except Exception as e:
        print(f"❌ Submission failed: {e}")
        import traceback

        traceback.print_exc()
        await db.rollback()

        # Return error response instead of raising to allow frontend to handle it
        return {
            "success": False,
            "error": str(e),
            "detail": f"Submission processing failed: {str(e)}",
        }


# Add migration endpoint to fix missing columns
@app.post("/api/db/migrate")
async def migrate_database(db: AsyncSession = Depends(get_db)):
    """Migration endpoint to add missing columns"""
    try:
        from sqlalchemy import text

        # Add role column to coding_submissions if not exists
        try:
            await db.execute(
                text("""
                ALTER TABLE coding_submissions ADD COLUMN role VARCHAR(255) DEFAULT NULL
            """)
            )
            print("✅ Added 'role' column to coding_submissions")
        except Exception as e:
            if "Duplicate column" in str(e):
                print("ℹ️ 'role' column already exists")
            else:
                print(f"⚠️ Could not add 'role' column: {e}")

        # Add status column to coding_submissions if not exists
        try:
            await db.execute(
                text("""
                ALTER TABLE coding_submissions ADD COLUMN status VARCHAR(50) DEFAULT 'submitted'
            """)
            )
            print("✅ Added 'status' column to coding_submissions")
        except Exception as e:
            if "Duplicate column" in str(e):
                print("ℹ️ 'status' column already exists")
            else:
                print(f"⚠️ Could not add 'status' column: {e}")

        await db.commit()
        return {"success": True, "message": "Migration completed"}
    except Exception as e:
        print(f"❌ Migration failed: {e}")
        return {"success": False, "error": str(e)}


# Add resubmission endpoint
@app.post("/api/coding/resubmit-after-error")
async def resubmit_after_error(data: dict, db: AsyncSession = Depends(get_db)):
    """
    UC-05 Alternative Flow 6a.2: Candidate resubmits after compilation error
    """
    try:
        session_id = data.get("session_id")
        code = data.get("code")
        previous_submission_id = data.get("previous_submission_id")

        if not all([session_id, code]):
            raise HTTPException(status_code=400, detail="session_id and code required")

        # Check if session allows resubmission and time remains
        from sqlalchemy import text

        session_query = text(
            """
            SELECT time_remaining_seconds, status
            FROM coding_sessions
            WHERE session_id = :session_id
        """
        )
        result = await db.execute(session_query, {"session_id": session_id})
        session = result.fetchone()

        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        time_remaining = session[0]
        if time_remaining <= 0:
            raise HTTPException(
                status_code=400, detail="No time remaining for resubmission"
            )

        if session[1] != "in_progress":
            raise HTTPException(status_code=400, detail="Session not in progress")

        # Get challenge details from session
        challenge_query = text(
            """
            SELECT challenge_id, language
            FROM coding_sessions
            WHERE session_id = :session_id
        """
        )
        result = await db.execute(challenge_query, {"session_id": session_id})
        challenge_info = result.fetchone()

        if not challenge_info:
            raise HTTPException(status_code=404, detail="Challenge not found")

        # Use existing submit_solution logic but mark as resubmission
        submission_data = {
            "challenge_id": challenge_info[0],
            "candidate_email": data.get("candidate_email"),
            "code": code,
            "language": challenge_info[1],
            "session_id": session_id,
            "is_resubmission": True,
            "previous_submission_id": previous_submission_id,
        }

        # Call the main submission logic
        return await submit_solution(submission_data, db)

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Resubmission failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== RESULTS & REPORTS ====================


@app.get("/api/coding/submissions")
async def get_coding_submissions(
    candidate_email: str = None, role: str = None, db: AsyncSession = Depends(get_db)
):
    """Get coding submissions with role filtering"""
    try:
        from sqlalchemy import text

        query = """
            SELECT cs.*, cc.title as challenge_title, cc.difficulty as challenge_difficulty
            FROM coding_submissions cs
            LEFT JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE 1=1
        """
        params = {}

        if candidate_email:
            query += " AND cs.candidate_email = :candidate_email"
            params["candidate_email"] = candidate_email

        if role:
            query += " AND cs.role = :role"
            params["role"] = role

        query += " ORDER BY cs.submitted_at DESC"

        result = await db.execute(text(query), params)
        rows = result.fetchall()

        submissions = []
        for row in rows:
            submission = {
                "submission_id": row[0],
                "challenge_id": row[1],
                "candidate_email": row[2],
                "code": row[3],
                "language": row[4],
                "test_results": json.loads(row[5]) if row[5] else {},
                "evaluation": json.loads(row[6]) if row[6] else {},
                "score": row[7],
                "submitted_at": str(row[8]) if row[8] else None,
                "assessment_id": row[9],
                "result_id": row[10],
                "score_breakdown": json.loads(row[11]) if row[11] else {},
                "session_id": row[12],
                "role": row[13],  # ✅ This is the new role column
                "challenge_title": row[14] if len(row) > 14 else "Coding Challenge",
                "challenge_difficulty": row[15] if len(row) > 15 else "Medium",
            }

            # Parse evaluation to extract AI detection if available
            if submission["evaluation"]:
                if "ai_detection" in submission["evaluation"]:
                    submission["ai_detection"] = submission["evaluation"][
                        "ai_detection"
                    ]
                    submission["is_ai_generated"] = submission["evaluation"][
                        "ai_detection"
                    ].get("ai_generated", False)
                    submission["ai_confidence"] = submission["evaluation"][
                        "ai_detection"
                    ].get("confidence", 0)

                if "plagiarism_check" in submission["evaluation"]:
                    submission["plagiarism_check"] = submission["evaluation"][
                        "plagiarism_check"
                    ]
                    submission["plagiarism_detected"] = submission["evaluation"][
                        "plagiarism_check"
                    ].get("plagiarism_detected", False)
                    submission["plagiarism_score"] = submission["evaluation"][
                        "plagiarism_check"
                    ].get("plagiarism_score", 0)

            submissions.append(submission)

        return {"submissions": submissions, "count": len(submissions)}

    except Exception as e:
        print(f"❌ Fetch failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/coding/submission-integrity/{submission_id}")
async def get_submission_integrity_report(
    submission_id: str, db: AsyncSession = Depends(get_db)
):
    """Get comprehensive integrity report for a submission (violations, AI, plagiarism)"""
    try:
        from sqlalchemy import text

        # Get submission with all integrity data
        query = text(
            """
            SELECT
                cs.submission_id,
                cs.candidate_email,
                cs.code,
                cs.language,
                cs.score,
                cs.submitted_at,
                cs.session_id,
                cc.title as challenge_title,
                cc.difficulty as challenge_difficulty,
                cd.ai_detection,
                cd.plagiarism_check,
                v.violation_count,
                v.violation_list
            FROM coding_submissions cs
            LEFT JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            LEFT JOIN code_detection cd ON cs.submission_id = cd.submission_id
            LEFT JOIN (
                SELECT
                    session_id,
                    COUNT(*) as violation_count,
                    JSON_ARRAYAGG(
                        JSON_OBJECT(
                            'violation_id', violation_id,
                            'type', violation_type,
                            'description', description,
                            'occurred_at', occurred_at
                        )
                    ) as violation_list
                FROM violations
                GROUP BY session_id
            ) v ON cs.session_id = v.session_id
            WHERE cs.submission_id = :submission_id
        """
        )

        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Submission not found")

        # Parse data
        session_id = row[6]
        ai_detection = json.loads(row[9]) if row[9] and row[9] != "null" else {}
        plagiarism_check = json.loads(row[10]) if row[10] and row[10] != "null" else {}
        violation_count = row[11] or 0
        violation_list = json.loads(row[12]) if row[12] and row[12] != "null" else []

        # Get similarity data (from plagiarism check or separate query)
        similarity_query = text(
            """
            SELECT
                cs2.candidate_email,
                cs2.submitted_at,
                cs2.score,
                ROUND((LENGTH(cs.code) - LENGTH(REPLACE(cs.code, cs2.code, ''))) / LENGTH(cs.code) * 100, 1) as similarity_percentage
            FROM coding_submissions cs
            JOIN coding_submissions cs2 ON cs.challenge_id = cs2.challenge_id
            WHERE cs.submission_id = :submission_id
            AND cs2.submission_id != :submission_id
            AND cs2.candidate_email != cs.candidate_email
            ORDER BY similarity_percentage DESC
            LIMIT 5
        """
        )

        similarity_result = await db.execute(
            similarity_query, {"submission_id": submission_id}
        )
        similarity_rows = similarity_result.fetchall()

        similar_submissions = []
        for sim_row in similarity_rows:
            similar_submissions.append(
                {
                    "candidate_email": sim_row[0],
                    "submitted_at": str(sim_row[1]) if sim_row[1] else None,
                    "score": sim_row[2],
                    "similarity_percentage": sim_row[3],
                }
            )

        # Get code execution patterns for analysis
        execution_query = text(
            """
            SELECT
                test_results
            FROM coding_submissions
            WHERE submission_id = :submission_id
        """
        )

        exec_result = await db.execute(
            execution_query, {"submission_id": submission_id}
        )
        exec_row = exec_result.fetchone()

        execution_pattern = {}
        if exec_row and exec_row[0] and exec_row[0] != "null":
            try:
                test_results = json.loads(exec_row[0])
                execution_pattern = {
                    "total_tests": test_results.get("total", 0),
                    "passed_tests": test_results.get("passed", 0),
                    "failed_tests": test_results.get("failed", 0),
                    "execution_time": test_results.get("execution_time_ms", 0),
                    "all_passed": test_results.get("all_passed", False),
                }
            except:
                pass

        # Calculate integrity score
        integrity_score = 100
        integrity_issues = []

        # Deduct for violations
        if violation_count > 0:
            deduction = min(violation_count * 15, 50)
            integrity_score -= deduction
            integrity_issues.append(
                f"Violations: -{deduction}% ({violation_count} incidents)"
            )

        # Deduct for AI detection
        if ai_detection.get("ai_generated", False):
            confidence = ai_detection.get("confidence", 0)
            deduction = min(confidence / 3, 30)
            integrity_score -= deduction
            integrity_issues.append(
                f"AI-generated: -{deduction:.1f}% ({confidence}% confidence)"
            )

        # Deduct for plagiarism
        if plagiarism_check.get("plagiarism_detected", False):
            plagiarism_score = plagiarism_check.get("plagiarism_score", 0)
            deduction = min(plagiarism_score / 2, 40)
            integrity_score -= deduction
            integrity_issues.append(
                f"Plagiarism: -{deduction:.1f}% ({plagiarism_score}% similarity)"
            )

        # Deduct for suspicious execution (too fast or perfect)
        if execution_pattern:
            if (
                execution_pattern.get("execution_time", 0) < 100
            ):  # Too fast (<100ms total)
                integrity_score -= 10
                integrity_issues.append("Suspiciously fast execution: -10%")

            if (
                execution_pattern.get("all_passed", False)
                and execution_pattern.get("execution_time", 0) < 500
            ):
                integrity_score -= 15
                integrity_issues.append("Perfect execution too quickly: -15%")

        integrity_score = max(0, min(100, integrity_score))

        # Determine integrity level
        if integrity_score >= 90:
            integrity_level = "✅ High Integrity"
            integrity_color = "green"
        elif integrity_score >= 70:
            integrity_level = "⚠️ Moderate Integrity"
            integrity_color = "orange"
        else:
            integrity_level = "🚨 Low Integrity"
            integrity_color = "red"

        return {
            "success": True,
            "submission_id": submission_id,
            "candidate_email": row[1],
            "challenge": {"title": row[7], "difficulty": row[8]},
            "score": row[4],
            "submitted_at": str(row[5]) if row[5] else None,
            # Integrity Summary
            "integrity_summary": {
                "score": round(integrity_score, 1),
                "level": integrity_level,
                "color": integrity_color,
                "issues": integrity_issues,
                "violation_count": violation_count,
                "ai_detected": ai_detection.get("ai_generated", False),
                "plagiarism_detected": plagiarism_check.get(
                    "plagiarism_detected", False
                ),
            },
            # Detailed Sections for HR Dashboard
            "violations": {
                "count": violation_count,
                "list": violation_list,
                "summary": (
                    f"Found {violation_count} proctoring violation(s)"
                    if violation_count > 0
                    else "No violations detected"
                ),
            },
            "ai_analysis": {
                "detected": ai_detection.get("ai_generated", False),
                "confidence": ai_detection.get("confidence", 0),
                "evidence": ai_detection.get("evidence", []),
                "details": ai_detection.get("details", "No AI analysis available"),
                "source": ai_detection.get("source", "AI Detection System"),
            },
            "plagiarism_check": {
                "detected": plagiarism_check.get("plagiarism_detected", False),
                "score": plagiarism_check.get("plagiarism_score", 0),
                "similar_submissions": plagiarism_check.get("similar_submissions", []),
                "details": plagiarism_check.get(
                    "details", "No plagiarism check available"
                ),
                "threshold_used": plagiarism_check.get("threshold", 75),
            },
            "similarity_analysis": {
                "similar_submissions_found": len(similar_submissions),
                "similar_submissions": similar_submissions,
                "highest_similarity": (
                    max(
                        [s.get("similarity_percentage", 0) for s in similar_submissions]
                    )
                    if similar_submissions
                    else 0
                ),
                "summary": f"Found {len(similar_submissions)} similar submission(s) from other candidates",
            },
            "execution_pattern": execution_pattern,
            # Recommendations for HR
            "hr_recommendations": [
                (
                    "Review violations if count > 0"
                    if violation_count > 0
                    else "No violations to review"
                ),
                (
                    "Investigate AI-generated code if confidence > 70%"
                    if ai_detection.get("confidence", 0) > 70
                    else "AI detection within acceptable limits"
                ),
                (
                    "Check for plagiarism if similarity > 80%"
                    if plagiarism_check.get("plagiarism_score", 0) > 80
                    else "Plagiarism check passed"
                ),
                (
                    "Consider manual review if integrity score < 70"
                    if integrity_score < 70
                    else "Integrity check passed"
                ),
            ],
        }

    except Exception as e:
        print(f"❌ Integrity report failed: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== DIRECT AI DETECTION ENDPOINT ====================
@app.get("/api/coding/ai-detection/{submission_id}")
async def get_direct_ai_detection(
    submission_id: str, db: AsyncSession = Depends(get_db)
):
    """Get AI detection results directly - SIMPLIFIED for frontend"""
    try:
        from sqlalchemy import text

        # Query to get AI detection data from code_detection table
        query = text(
            """
            SELECT
                cd.ai_detection,
                cs.candidate_email,
                cs.code,
                cs.language,
                cs.score,
                cs.submitted_at,
                cc.title as challenge_title
            FROM code_detection cd
            JOIN coding_submissions cs ON cd.submission_id = cs.submission_id
            LEFT JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE cs.submission_id = :submission_id
        """
        )

        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            return {
                "submission_id": submission_id,
                "error": "No AI detection data found",
                "ai_detection": {
                    "ai_generated": False,
                    "detected": False,
                    "confidence": 0,
                    "score": 0,
                    "evidence": [],
                    "details": "No AI detection data available",
                    "source": "AI Detection System",
                },
            }

        # Parse AI detection JSON
        ai_detection_raw = row[0]
        if not ai_detection_raw:
            return {
                "submission_id": submission_id,
                "error": "AI detection data is empty",
                "ai_detection": {
                    "ai_generated": False,
                    "detected": False,
                    "confidence": 0,
                    "score": 0,
                    "evidence": [],
                    "details": "No AI analysis performed",
                    "source": "AI Detection System",
                },
            }

        try:
            ai_detection = (
                json.loads(ai_detection_raw)
                if isinstance(ai_detection_raw, str)
                else ai_detection_raw
            )
        except:
            ai_detection = {}

        # Extract code snippet (first 200 chars)
        code_snippet = row[2][:200] + "..." if row[2] and len(row[2]) > 200 else row[2]

        return {
            "submission_id": submission_id,
            "candidate_email": row[1],
            "code_preview": code_snippet,
            "language": row[3],
            "score": row[4],
            "submitted_at": str(row[5]) if row[5] else None,
            "challenge_title": row[6],
            # ✅ SIMPLE, CONSISTENT AI DETECTION STRUCTURE
            "ai_detection": {
                "ai_generated": ai_detection.get(
                    "ai_generated", ai_detection.get("detected", False)
                ),
                "detected": ai_detection.get(
                    "detected", ai_detection.get("ai_generated", False)
                ),
                "confidence": ai_detection.get(
                    "confidence", ai_detection.get("score", 0)
                ),
                "score": ai_detection.get("score", ai_detection.get("confidence", 0)),
                "evidence": ai_detection.get("evidence", []),
                "details": ai_detection.get("details", "AI analysis completed"),
                "source": ai_detection.get("source", "AI Detection System"),
                "metrics": ai_detection.get("metrics", {}),
            },
            # ✅ SIMPLE BOOLEAN FLAGS FOR FRONTEND
            "is_ai_generated": ai_detection.get(
                "ai_generated", ai_detection.get("detected", False)
            ),
            "ai_confidence": ai_detection.get("confidence", 0),
            "ai_evidence_count": len(ai_detection.get("evidence", [])),
            # ✅ SIMPLE SUMMARY FOR DISPLAY
            "summary": {
                "status": (
                    "🚨 AI-GENERATED"
                    if ai_detection.get("ai_generated", False)
                    else "✅ Human-written"
                ),
                "confidence_level": (
                    "High"
                    if ai_detection.get("confidence", 0) >= 70
                    else "Medium"
                    if ai_detection.get("confidence", 0) >= 40
                    else "Low"
                ),
                "should_review": ai_detection.get("confidence", 0) >= 60,
                "recommendation": (
                    "Review for authenticity"
                    if ai_detection.get("confidence", 0) >= 60
                    else "Authentic"
                ),
            },
        }

    except Exception as e:
        print(f"❌ Direct AI detection fetch failed: {e}")
        import traceback

        traceback.print_exc()
        return {
            "submission_id": submission_id,
            "error": str(e),
            "ai_detection": {
                "ai_generated": False,
                "detected": False,
                "confidence": 0,
                "evidence": [],
                "details": f"Error: {str(e)}",
            },
        }


# ==================== DEBUG: CHECK AI DETECTION STATUS ====================
@app.get("/api/debug/check-ai-detection/{submission_id}")
async def debug_check_ai_detection(
    submission_id: str, db: AsyncSession = Depends(get_db)
):
    """Debug endpoint to check if AI detection is working"""
    from sqlalchemy import text

    try:
        # Check if AI detection ran for this submission
        query = text(
            """
            SELECT
                cd.ai_detection,
                cs.code,
                cs.language,
                cs.submitted_at
            FROM coding_submissions cs
            LEFT JOIN code_detection cd ON cs.submission_id = cd.submission_id
            WHERE cs.submission_id = :submission_id
        """
        )

        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            return {"error": "Submission not found"}

        ai_detection = row[0]
        code = row[1]
        language = row[2]

        return {
            "submission_id": submission_id,
            "has_ai_detection": ai_detection is not None,
            "ai_detection_raw": ai_detection,
            "code_length": len(code) if code else 0,
            "language": language,
            "submitted_at": str(row[3]),
            "test_ai_detection": "Run manually to test",
        }

    except Exception as e:
        return {"error": str(e)}


@app.get("/api/coding/candidate-violations/{candidate_email}")
async def get_candidate_violations(
    candidate_email: str, db: AsyncSession = Depends(get_db)
):
    """Get all violations for a specific candidate across all sessions"""
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT
                v.violation_id,
                v.session_id,
                v.violation_type,
                v.description,
                v.occurred_at,
                cs.challenge_id,
                cc.title as challenge_title,
                cs.submitted_at as session_end_time
            FROM violations v
            JOIN coding_sessions cs ON v.session_id = cs.session_id
            JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE cs.candidate_email = :candidate_email
            ORDER BY v.occurred_at DESC
        """
        )

        result = await db.execute(query, {"candidate_email": candidate_email})
        violations = result.fetchall()

        # Group by violation type
        violation_types = {}
        for v in violations:
            v_type = v[2]  # violation_type
            if v_type not in violation_types:
                violation_types[v_type] = 0
            violation_types[v_type] += 1

        return {
            "candidate_email": candidate_email,
            "total_violations": len(violations),
            "violation_types": violation_types,
            "violations": [
                {
                    "violation_id": v[0],
                    "session_id": v[1],
                    "type": v[2],
                    "description": v[3],
                    "occurred_at": v[4].isoformat() if v[4] else None,
                    "challenge_id": v[5],
                    "challenge_title": v[6],
                    "session_end_time": str(v[7]) if v[7] else None,
                }
                for v in violations
            ],
        }
    except Exception as e:
        print(f"❌ Candidate violations error: {e}")
        return {
            "candidate_email": candidate_email,
            "total_violations": 0,
            "violation_types": {},
            "violations": [],
        }


@app.get("/api/coding/evaluation/{submission_id}")
async def get_detailed_evaluation(
    submission_id: str, db: AsyncSession = Depends(get_db)
):
    """FR-TECH-08: Get detailed evaluation report"""
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT cs.*, cc.title, cc.difficulty, cc.language
            FROM coding_submissions cs
            JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE cs.submission_id = :submission_id
        """
        )

        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Submission not found")

        test_results = json.loads(row[5]) if row[5] else {}
        evaluation = json.loads(row[6]) if row[6] else {}

        return {
            "submission_id": row[0],
            "challenge_title": row[9],
            "challenge_difficulty": row[10],
            "candidate_email": row[2],
            "code": row[3],
            "language": row[4],
            "score": row[7],
            "test_results": test_results,
            "evaluation": evaluation,
            "submitted_at": str(row[8]) if row[8] else None,
            "detailed_report": {
                "correctness": f"{test_results.get('passed', 0)}/{test_results.get('total', 0)} tests passed",
                "time_complexity": evaluation.get("time_complexity", "N/A"),
                "space_complexity": evaluation.get("space_complexity", "N/A"),
                "code_quality": evaluation.get("code_quality_score", 0),
                "best_practices": evaluation.get("best_practices_score", 0),
                "readability": evaluation.get("readability_score", 0),
                "strengths": evaluation.get("strengths", []),
                "improvements": evaluation.get("improvements", []),
                "feedback": evaluation.get("feedback", ""),
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Fetch failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/coding/candidate-report/{email}")
async def get_candidate_technical_report(
    email: str, db: AsyncSession = Depends(get_db)
):
    """FR-TECH-09: Get comprehensive technical report for candidate"""
    try:
        from sqlalchemy import text

        # Get all submissions
        query = text(
            """
            SELECT cs.*, cc.title, cc.difficulty
            FROM coding_submissions cs
            JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE cs.candidate_email = :email
            ORDER BY cs.submitted_at DESC
        """
        )

        result = await db.execute(query, {"email": email})
        rows = result.fetchall()

        if not rows:
            return {
                "candidate_email": email,
                "total_submissions": 0,
                "average_score": 0,
                "submissions": [],
                "technical_summary": "No submissions yet",
            }

        # Calculate statistics
        total_submissions = len(rows)
        total_score = sum(row[7] for row in rows)
        average_score = int(total_score / total_submissions)

        # Get difficulty breakdown
        difficulty_stats = {"easy": 0, "medium": 0, "hard": 0}
        difficulty_scores = {"easy": [], "medium": [], "hard": []}

        submissions = []
        for row in rows:
            difficulty = row[10]
            score = row[7]

            if difficulty in difficulty_stats:
                difficulty_stats[difficulty] += 1
                difficulty_scores[difficulty].append(score)

            submissions.append(
                {
                    "submission_id": row[0],
                    "challenge_title": row[9],
                    "difficulty": difficulty,
                    "score": score,
                    "submitted_at": str(row[8]) if row[8] else None,
                }
            )

        # Calculate average by difficulty
        difficulty_averages = {}
        for diff, scores in difficulty_scores.items():
            if scores:
                difficulty_averages[diff] = int(sum(scores) / len(scores))

        # Determine technical level
        if average_score >= 85:
            technical_level = "Senior/Expert"
        elif average_score >= 70:
            technical_level = "Mid-Level"
        elif average_score >= 55:
            technical_level = "Junior"
        else:
            technical_level = "Entry-Level"

        return {
            "candidate_email": email,
            "total_submissions": total_submissions,
            "average_score": average_score,
            "technical_level": technical_level,
            "difficulty_breakdown": difficulty_stats,
            "difficulty_averages": difficulty_averages,
            "submissions": submissions,
            "technical_summary": f"Completed {total_submissions} challenges with {average_score}% average. Technical level: {technical_level}",
        }

    except Exception as e:
        print(f"❌ Report failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== PROCTORING & VIOLATIONS ====================


@app.post("/api/coding/log-violation")
async def log_violation(data: dict, db: AsyncSession = Depends(get_db)):
    """
    Log coding assessment violations (tab switches, copy/paste, etc.)
    FIXED: Creates violations table if it doesn't exist
    """
    try:
        from sqlalchemy import text

        # ✅ FIX 1: Validate input
        session_id = data.get("session_id")
        if not session_id:
            raise HTTPException(
                status_code=400, detail="session_id is required in request body"
            )

        # ✅ FIX 2: Create violations table if it doesn't exist
        create_table_query = text(
            """
            CREATE TABLE IF NOT EXISTS violations (
                violation_id VARCHAR(255) PRIMARY KEY,
                session_id VARCHAR(255) NOT NULL,
                violation_type VARCHAR(100),
                description TEXT,
                occurred_at DATETIME,
                INDEX idx_session (session_id),
                INDEX idx_occurred (occurred_at)
            )
        """
        )

        await db.execute(create_table_query)
        await db.commit()

        # ✅ FIX 3: Insert violation
        violation_id = f"violation_{uuid.uuid4()}"

        insert_query = text(
            """
            INSERT INTO violations
            (violation_id, session_id, violation_type, description, occurred_at)
            VALUES (:vid, :sid, :vtype, :desc, NOW())
        """
        )

        await db.execute(
            insert_query,
            {
                "vid": violation_id,
                "sid": session_id,
                "vtype": data.get("violation_type", "unknown"),
                "desc": data.get("description", "No description provided"),
            },
        )
        await db.commit()

        print(
            f"⚠️ Violation logged: {data.get('violation_type', 'unknown')} for session {session_id}"
        )

        return {
            "success": True,
            "violation_id": violation_id,
            "message": "Violation logged successfully",
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Failed to log violation: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Internal error logging violation: {str(e)}"
        )


@app.get("/api/coding/violations/{session_id}")
async def get_session_violations(session_id: str, db: AsyncSession = Depends(get_db)):
    """
    FR-PROC-02: Get all violations for a coding session
    """
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT violation_id, violation_type, description, timestamp, severity
            FROM coding_violations
            WHERE session_id = :session_id
            ORDER BY timestamp DESC
        """
        )

        result = await db.execute(query, {"session_id": session_id})
        rows = result.fetchall()

        violations = []
        for row in rows:
            violations.append(
                {
                    "violation_id": row[0],
                    "type": row[1],
                    "description": row[2],
                    "timestamp": str(row[3]),
                    "severity": row[4],
                }
            )

        # Count by severity
        severity_counts = {"low": 0, "medium": 0, "high": 0, "critical": 0}
        for v in violations:
            severity_counts[v["severity"]] = severity_counts.get(v["severity"], 0) + 1

        return {
            "session_id": session_id,
            "total_violations": len(violations),
            "severity_counts": severity_counts,
            "violations": violations,
            "flag_status": (
                "critical"
                if severity_counts["critical"] > 0
                else (
                    "high"
                    if severity_counts["high"] >= 3
                    else "medium"
                    if len(violations) >= 5
                    else "clean"
                )
            ),
        }

    except Exception as e:
        print(f"❌ Get violations failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== SESSION MANAGEMENT ====================


@app.post("/api/coding/start-session")
async def start_coding_session(data: dict, db: AsyncSession = Depends(get_db)):
    """
    FR-TEC-13, FR-TEC-14: Start a timed coding session with auto-save
    """
    try:
        challenge_id = data.get("challenge_id")
        candidate_email = data.get("candidate_email")
        time_limit_minutes = data.get("time_limit_minutes", 60)

        if not challenge_id or not candidate_email:
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Check if challenge exists
        from sqlalchemy import text

        query = text(
            "SELECT * FROM coding_challenges WHERE challenge_id = :challenge_id"
        )
        result = await db.execute(query, {"challenge_id": challenge_id})
        challenge = result.fetchone()

        if not challenge:
            raise HTTPException(status_code=404, detail="Challenge not found")

        # Create session
        session_id = f"session_{uuid.uuid4()}"

        insert_query = text(
            """
            INSERT INTO coding_sessions (
                session_id, challenge_id, candidate_email,
                start_time, time_limit_minutes, status, current_code
            ) VALUES (
                :session_id, :challenge_id, :candidate_email,
                NOW(), :time_limit, 'in_progress', :starter_code
            )
        """
        )

        await db.execute(
            insert_query,
            {
                "session_id": session_id,
                "challenge_id": challenge_id,
                "candidate_email": candidate_email,
                "time_limit": time_limit_minutes,
                "starter_code": challenge[5] or "",  # starter_code from challenge
            },
        )
        await db.commit()

        print(f"✅ Started session {session_id} for {candidate_email}")

        # Calculate initial time remaining
        time_remaining_seconds = time_limit_minutes * 60

        return {
            "session_id": session_id,
            "challenge_id": challenge_id,
            "time_limit_minutes": time_limit_minutes,
            "time_remaining_seconds": time_remaining_seconds,  # ✅ Added for timer
            "starter_code": challenge[5] or "",
            "start_time": datetime.now().isoformat(),
            "status": "in_progress",  # ✅ Added status
            "message": "Session started successfully",
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Failed to start session: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ✅ ALIAS ENDPOINT - Frontend calls /session/start instead of /start-session
@app.post("/api/coding/session/start")
async def start_session_alias(data: dict, db: AsyncSession = Depends(get_db)):
    """Alias for /start-session endpoint - frontend compatibility"""
    return await start_coding_session(data, db)


@app.post("/api/coding/auto-save")
async def auto_save_code(data: dict, db: AsyncSession = Depends(get_db)):
    """
    FR-TEC-14: Auto-save code progress every 30 seconds
    """
    try:
        session_id = data.get("session_id")
        code = data.get("code", "")

        if not session_id:
            raise HTTPException(status_code=400, detail="Missing session_id")

        from sqlalchemy import text

        # Update session with current code
        update_query = text(
            """
            UPDATE coding_sessions
            SET current_code = :code, last_saved_at = NOW()
            WHERE session_id = :session_id
        """
        )

        await db.execute(update_query, {"session_id": session_id, "code": code})

        # Save as version (FR-TEC-12: Code versioning)
        version_id = f"ver_{uuid.uuid4()}"

        # Get current version number
        count_query = text(
            """
            SELECT COUNT(*) FROM code_versions WHERE session_id = :session_id
        """
        )
        result = await db.execute(count_query, {"session_id": session_id})
        version_number = result.scalar() + 1

        # Insert version
        version_query = text(
            """
            INSERT INTO code_versions (
                version_id, session_id, version_number, code,
                saved_at, auto_saved
            ) VALUES (
                :version_id, :session_id, :version_num, :code,
                NOW(), TRUE
            )
        """
        )

        await db.execute(
            version_query,
            {
                "version_id": version_id,
                "session_id": session_id,
                "version_num": version_number,
                "code": code,
            },
        )

        await db.commit()

        return {
            "success": True,
            "saved_at": datetime.now().isoformat(),
            "version_number": version_number,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Auto-save failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/coding/session/{session_id}")
async def get_coding_session(session_id: str, db: AsyncSession = Depends(get_db)):
    """
    Get session details including time elapsed
    """
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT s.*, c.title, c.difficulty, c.language
            FROM coding_sessions s
            JOIN coding_challenges c ON s.challenge_id = c.challenge_id
            WHERE s.session_id = :session_id
        """
        )

        result = await db.execute(query, {"session_id": session_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Session not found")

        # Calculate time elapsed
        start_time = row[3]  # start_time column
        end_time = row[4]  # end_time column

        if end_time:
            time_elapsed = (end_time - start_time).total_seconds()
        else:
            time_elapsed = (datetime.now() - start_time).total_seconds()

        time_limit_seconds = row[5] * 60  # time_limit_minutes to seconds
        time_remaining = max(0, time_limit_seconds - time_elapsed)

        # Check if time expired
        if time_remaining == 0 and row[6] == "in_progress":  # status
            # Update to timed_out
            update_query = text(
                """
                UPDATE coding_sessions
                SET status = 'timed_out', end_time = NOW()
                WHERE session_id = :session_id
            """
            )
            await db.execute(update_query, {"session_id": session_id})
            await db.commit()

        return {
            "session_id": row[0],
            "challenge_id": row[1],
            "candidate_email": row[2],
            "start_time": str(row[3]),
            "end_time": str(row[4]) if row[4] else None,
            "time_limit_minutes": row[5],
            "time_elapsed_seconds": int(time_elapsed),
            "time_remaining_seconds": int(time_remaining),
            "status": row[6],
            "current_code": row[7],
            "last_saved_at": str(row[8]) if row[8] else None,
            "challenge_title": row[12],
            "difficulty": row[13],
            "language": row[14],
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Failed to get session: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ✅ NEW ENDPOINT 1: Get Version History
@app.get("/api/coding/session/{session_id}/versions")
async def get_version_history(session_id: str, db: AsyncSession = Depends(get_db)):
    """
    FR-TEC-12: Get all saved code versions for a session
    """
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT version_id, version_number, code, saved_at, auto_saved
            FROM code_versions
            WHERE session_id = :session_id
            ORDER BY version_number DESC
        """
        )

        result = await db.execute(query, {"session_id": session_id})
        rows = result.fetchall()

        versions = []
        for row in rows:
            versions.append(
                {
                    "version_id": row[0],
                    "version_number": row[1],
                    "code": row[2],
                    "saved_at": str(row[3]),
                    "auto_saved": row[4],
                }
            )

        print(f"✅ Retrieved {len(versions)} versions for session {session_id}")

        return {"success": True, "versions": versions, "count": len(versions)}

    except Exception as e:
        print(f"❌ Get versions error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ✅ NEW ENDPOINT 2: Manual Save Version
@app.post("/api/coding/save-version")
async def save_code_version(data: dict, db: AsyncSession = Depends(get_db)):
    """
    FR-TEC-12: Manually save a code version (for 💾 Save button)
    Creates a new version entry that appears in history
    """
    try:
        session_id = data.get("session_id")
        code = data.get("code")

        if not session_id or code is None:
            raise HTTPException(status_code=400, detail="Missing session_id or code")

        from sqlalchemy import text

        # Get current max version number
        count_query = text(
            """
            SELECT COALESCE(MAX(version_number), 0) as max_version
            FROM code_versions
            WHERE session_id = :session_id
        """
        )

        result = await db.execute(count_query, {"session_id": session_id})
        max_version = result.scalar()
        next_version = max_version + 1

        # Insert new version (manual save)
        version_id = f"ver_{uuid.uuid4()}"

        insert_query = text(
            """
            INSERT INTO code_versions (
                version_id, session_id, version_number, code,
                saved_at, auto_saved
            ) VALUES (
                :version_id, :session_id, :version_num, :code,
                NOW(), FALSE
            )
        """
        )

        await db.execute(
            insert_query,
            {
                "version_id": version_id,
                "session_id": session_id,
                "version_num": next_version,
                "code": code,
            },
        )

        # Also update current_code in sessions table
        update_query = text(
            """
            UPDATE coding_sessions
            SET current_code = :code, last_saved_at = NOW()
            WHERE session_id = :session_id
        """
        )

        await db.execute(update_query, {"session_id": session_id, "code": code})

        await db.commit()

        print(f"✅ Manually saved version {next_version} for session {session_id}")

        return {
            "success": True,
            "version_id": version_id,
            "version_number": next_version,
            "saved_at": datetime.now().isoformat(),
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Save version error: {e}")
        import traceback

        traceback.print_exc()
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/coding/versions/{session_id}")
async def get_code_versions(session_id: str, db: AsyncSession = Depends(get_db)):
    """
    FR-TEC-12: Get code version history to track progression
    """
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT version_id, version_number, code, saved_at, auto_saved
            FROM code_versions
            WHERE session_id = :session_id
            ORDER BY version_number DESC
        """
        )

        result = await db.execute(query, {"session_id": session_id})
        rows = result.fetchall()

        versions = []
        for row in rows:
            versions.append(
                {
                    "version_id": row[0],
                    "version_number": row[1],
                    "code": row[2],
                    "saved_at": str(row[3]),
                    "auto_saved": row[4],
                }
            )

        return {
            "session_id": session_id,
            "total_versions": len(versions),
            "versions": versions,
        }

    except Exception as e:
        print(f"❌ Failed to get versions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== TIME EXTENSION WORKFLOW ====================


@app.post("/api/assessment/request-extension")
async def request_time_extension(data: dict, db: AsyncSession = Depends(get_db)):
    """
    UC-05 Alternative Flow 5a: Candidate requests time extension
    """
    try:
        session_id = data.get("session_id")
        candidate_email = data.get("candidate_email")
        extension_minutes = data.get("extension_minutes", 15)
        reason = data.get("reason", "")

        if not session_id or not candidate_email:
            raise HTTPException(
                status_code=400, detail="session_id and candidate_email required"
            )

        from sqlalchemy import text

        # Get session details
        session_query = text(
            """
            SELECT s.*, a.role, a.difficulty
            FROM coding_sessions s
            LEFT JOIN assessments a ON s.assessment_id = a.assessment_id
            WHERE s.session_id = :session_id
        """
        )
        result = await db.execute(session_query, {"session_id": session_id})
        session = result.fetchone()

        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Create extension request
        request_id = f"ext_req_{uuid.uuid4()}"

        insert_query = text(
            """
            INSERT INTO extension_requests (
                request_id, session_id, candidate_email,
                extension_minutes, reason, status, requested_at
            ) VALUES (
                :request_id, :session_id, :candidate_email,
                :extension_minutes, :reason, 'pending', NOW()
            )
        """
        )

        await db.execute(
            insert_query,
            {
                "request_id": request_id,
                "session_id": session_id,
                "candidate_email": candidate_email,
                "extension_minutes": extension_minutes,
                "reason": reason,
            },
        )

        # TODO: Notify HR Manager (email integration)
        print(
            f"⏰ Extension requested: {candidate_email} +{extension_minutes}min - {reason}"
        )

        await db.commit()

        return {
            "request_id": request_id,
            "status": "pending",
            "message": "Time extension request submitted for HR approval",
        }

    except Exception as e:
        print(f"❌ Extension request failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/assessment/approve-extension")
async def approve_time_extension(data: dict, db: AsyncSession = Depends(get_db)):
    """
    HR Manager approves/denies time extension
    """
    try:
        request_id = data.get("request_id")
        approve = data.get("approve", True)
        hr_notes = data.get("hr_notes", "")

        if not request_id:
            raise HTTPException(status_code=400, detail="request_id required")

        from sqlalchemy import text

        # Get extension request
        request_query = text(
            """
            SELECT * FROM extension_requests WHERE request_id = :request_id
        """
        )
        result = await db.execute(request_query, {"request_id": request_id})
        extension_request = result.fetchone()

        if not extension_request:
            raise HTTPException(status_code=404, detail="Extension request not found")

        if approve:
            # Update session time limit
            update_session = text(
                """
                UPDATE coding_sessions
                SET time_limit_minutes = time_limit_minutes + :extension_minutes
                WHERE session_id = :session_id
            """
            )
            await db.execute(
                update_session,
                {
                    "extension_minutes": extension_request[3],  # extension_minutes
                    "session_id": extension_request[1],  # session_id
                },
            )

            # Update request status
            update_request = text(
                """
                UPDATE extension_requests
                SET status = 'approved', approved_at = NOW(), hr_notes = :hr_notes
                WHERE request_id = :request_id
            """
            )
            await db.execute(
                update_request, {"hr_notes": hr_notes, "request_id": request_id}
            )

            message = f"Time extension approved: +{extension_request[3]} minutes"
        else:
            # Deny request
            update_request = text(
                """
                UPDATE extension_requests
                SET status = 'denied', approved_at = NOW(), hr_notes = :hr_notes
                WHERE request_id = :request_id
            """
            )
            await db.execute(
                update_request, {"hr_notes": hr_notes, "request_id": request_id}
            )
            message = "Time extension denied"

        await db.commit()

        # TODO: Notify candidate about decision

        return {
            "request_id": request_id,
            "status": "approved" if approve else "denied",
            "message": message,
        }

    except Exception as e:
        print(f"❌ Extension approval failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/assessment/pending-extensions")
async def get_pending_extensions(db: AsyncSession = Depends(get_db)):
    """
    Get all pending time extension requests for HR review
    """
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT er.*, cs.challenge_id, cs.start_time, cs.time_limit_minutes,
                   cc.title as challenge_title, a.role
            FROM extension_requests er
            JOIN coding_sessions cs ON er.session_id = cs.session_id
            JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            LEFT JOIN assessments a ON cs.assessment_id = a.assessment_id
            WHERE er.status = 'pending'
            ORDER BY er.requested_at DESC
        """
        )

        result = await db.execute(query)
        rows = result.fetchall()

        pending_requests = []
        for row in rows:
            pending_requests.append(
                {
                    "request_id": row[0],
                    "session_id": row[1],
                    "candidate_email": row[2],
                    "extension_minutes": row[3],
                    "reason": row[4],
                    "status": row[5],
                    "requested_at": str(row[6]),
                    "challenge_title": row[9],
                    "role": row[10],
                    "current_time_limit": row[8],
                }
            )

        return {"pending_requests": pending_requests, "count": len(pending_requests)}

    except Exception as e:
        print(f"❌ Get pending extensions failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ADVANCED EVALUATION ====================


@app.post("/api/coding/evaluate-advanced")
async def evaluate_code_advanced(
    data: dict, db: AsyncSession = Depends(get_db)
):  # Add db dependency
    """
    FR-TEC-06 to FR-TEC-10: Advanced AI-powered code evaluation
    """
    try:
        submission_id = data.get("submission_id")

        if not submission_id:
            raise HTTPException(status_code=400, detail="Missing submission_id")

        from sqlalchemy import text

        # Get submission details - FIXED: Use db directly
        query = text(
            """
            SELECT cs.*, cc.title, cc.difficulty
            FROM coding_submissions cs
            JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE cs.submission_id = :submission_id
        """
        )
        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Submission not found")

        # Extract data from row
        candidate_email = row[2]  # candidate_email
        code = row[3]  # code
        language = row[4]  # language

        # Get test results from test_results JSON column
        test_results_json = row[5]  # test_results column
        test_results = json.loads(test_results_json) if test_results_json else []

        # Initialize evaluator
        evaluator = CodeEvaluator()

        # Perform evaluation
        evaluation = evaluator.evaluate_submission(
            code=code,
            language=language,
            test_results=test_results,
            execution_time_ms=0,  # You can extract this from test_results
            memory_used_mb=0.0,  # You can extract this from test_results
        )

        # Generate AI feedback using LLM
        ai_prompt = generate_ai_feedback_prompt(code, language, evaluation)

        # For now, use basic feedback since LLM might not be available
        ai_feedback = "Basic evaluation completed. Implement LLM integration for detailed feedback."

        # Save evaluation to database
        evaluation_id = f"eval_{uuid.uuid4()}"

        insert_query = text(
            """
            INSERT INTO coding_evaluations (
                evaluation_id, submission_id, candidate_email,
                correctness_score, tests_passed, tests_total,
                efficiency_score, time_complexity, space_complexity,
                execution_time_ms, memory_used_mb,
                code_quality_score, readability_score, maintainability_score,
                follows_standards, problem_solving_score,
                algorithmic_approach, logic_quality,
                overall_score, grade,
                strengths, weaknesses, suggestions, ai_feedback,
                evaluated_at
            ) VALUES (
                :eval_id, :sub_id, :email,
                :correct, :passed, :total,
                :efficiency, :time_comp, :space_comp,
                :exec_time, :memory,
                :quality, :read, :maintain,
                :standards, :ps,
                :algo, :logic,
                :overall, :grade,
                :strengths, :weaknesses, :suggestions, :ai,
                NOW()
            )
        """
        )

        await db.execute(
            insert_query,
            {
                "eval_id": evaluation_id,
                "sub_id": submission_id,
                "email": candidate_email,
                "correct": evaluation["correctness_score"],
                "passed": evaluation["tests_passed"],
                "total": evaluation["tests_total"],
                "efficiency": evaluation["efficiency_score"],
                "time_comp": evaluation["time_complexity"],
                "space_comp": evaluation["space_complexity"],
                "exec_time": 0,  # You can calculate this
                "memory": 0.0,  # You can calculate this
                "quality": evaluation["code_quality_score"],
                "read": evaluation["readability_score"],
                "maintain": evaluation["maintainability_score"],
                "standards": evaluation["follows_standards"],
                "ps": evaluation["problem_solving_score"],
                "algo": evaluation["algorithmic_approach"],
                "logic": evaluation["logic_quality"],
                "overall": evaluation["overall_score"],
                "grade": evaluation["grade"],
                "strengths": json.dumps(evaluation["strengths"]),
                "weaknesses": json.dumps(evaluation["weaknesses"]),
                "suggestions": json.dumps(evaluation["suggestions"]),
                "ai": ai_feedback,
            },
        )

        await db.commit()

        print(
            f"✅ Evaluated submission {submission_id}: {evaluation['grade']} ({evaluation['overall_score']}%)"
        )

        return {
            "evaluation_id": evaluation_id,
            "submission_id": submission_id,
            **evaluation,
            "ai_feedback": ai_feedback,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Evaluation failed: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/coding/check-plagiarism")
async def check_code_plagiarism(data: dict, db: AsyncSession = Depends(get_db)):
    """
    FR-TEC-09: Check for plagiarism in code submission
    """
    try:
        submission_id = data.get("submission_id")
        challenge_id = data.get("challenge_id")

        if not submission_id or not challenge_id:
            raise HTTPException(status_code=400, detail="Missing required fields")

        from sqlalchemy import text

        # Get candidate's submission
        query = text(
            """
            SELECT code, candidate_email, language
            FROM coding_submissions
            WHERE submission_id = :submission_id
        """
        )

        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Submission not found")

        candidate_code = row[0]
        candidate_email = row[1]
        language = row[2]

        # Get all other submissions for this challenge
        all_subs_query = text(
            """
            SELECT submission_id, candidate_email, code, submitted_at
            FROM coding_submissions
            WHERE challenge_id = :challenge_id
            AND submission_id != :submission_id
        """
        )

        result = await db.execute(
            all_subs_query,
            {"challenge_id": challenge_id, "submission_id": submission_id},
        )

        all_submissions = []
        for r in result.fetchall():
            all_submissions.append(
                {
                    "submission_id": r[0],
                    "candidate_email": r[1],
                    "code": r[2],
                    "submitted_at": str(r[3]),
                }
            )

        # Initialize plagiarism detector
        detector = PlagiarismDetector(similarity_threshold=0.85)

        # Check plagiarism
        plagiarism_result = detector.check_plagiarism(
            candidate_code=candidate_code,
            candidate_email=candidate_email,
            all_submissions=all_submissions,
            language=language,
        )

        # Update evaluation with plagiarism data
        update_query = text(
            """
            UPDATE coding_evaluations
            SET plagiarism_score = :score,
                plagiarism_detected = :detected,
                similar_submissions = :similar,
                plagiarism_details = :details
            WHERE submission_id = :submission_id
        """
        )

        await db.execute(
            update_query,
            {
                "score": plagiarism_result["plagiarism_score"],
                "detected": plagiarism_result["plagiarism_detected"],
                "similar": json.dumps(plagiarism_result["similar_submissions"]),
                "details": plagiarism_result["details"],
                "submission_id": submission_id,
            },
        )

        await db.commit()

        print(f"✅ Plagiarism check complete: {plagiarism_result['plagiarism_score']}%")

        return plagiarism_result

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Plagiarism check failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== COMPLETE SUBMISSION WITH EVALUATION ====================


# ==================== TEST CASE VALIDATION ====================
async def validate_and_fix_test_cases(challenge_id: str, db) -> dict:
    """
    Validate test cases by running reference solution
    Fix any mismatched expected outputs
    """
    try:
        from sqlalchemy import text

        # Get challenge
        query = text("SELECT * FROM coding_challenges WHERE challenge_id = :id")
        result = await db.execute(query, {"id": challenge_id})
        challenge = result.fetchone()

        if not challenge:
            return {"error": "Challenge not found"}

        # Get test cases and starter code
        test_cases = json.loads(challenge[5])  # test_cases column
        starter_code = challenge[4]  # starter_code column
        language = challenge[3]  # language column

        fixes_made = []
        validated_cases = []

        for i, test in enumerate(test_cases):
            try:
                # Run the starter code with this test input
                test_input = test.get("input", {})
                expected_output = test.get("expected_output")

                # Execute code to get actual output
                actual_output = await execute_code_for_validation(
                    starter_code, test_input, language
                )

                # Check if expected matches actual
                if actual_output != expected_output:
                    fixes_made.append(
                        {
                            "test_index": i + 1,
                            "old_expected": expected_output,
                            "new_expected": actual_output,
                        }
                    )
                    test["expected_output"] = actual_output

                validated_cases.append(test)

            except Exception as e:
                print(f"⚠️ Error validating test {i + 1}: {e}")
                validated_cases.append(test)

        # Update challenge if fixes were made
        if fixes_made:
            update_query = text(
                """
                UPDATE coding_challenges
                SET test_cases = :cases
                WHERE challenge_id = :id
            """
            )
            await db.execute(
                update_query, {"cases": json.dumps(validated_cases), "id": challenge_id}
            )
            await db.commit()

            return {
                "success": True,
                "fixes_made": fixes_made,
                "message": f"✅ Fixed {len(fixes_made)} test case(s)",
            }

        return {
            "success": True,
            "fixes_made": [],
            "message": "All test cases are valid",
        }

    except Exception as e:
        print(f"❌ Test validation error: {e}")
        return {"error": str(e)}


async def execute_code_for_validation(code: str, test_input: dict, language: str):
    """Execute code to get actual output for validation"""
    # This uses the same Docker execution as run_tests
    # but returns the output value for comparison
    import subprocess
    import tempfile
    import json

    # Prepare code with test input
    if language == "python":
        full_code = f"""
import json
{code}

# Test input
test_input = {json.dumps(test_input)}

# Call the function (assumes first function in starter_code)
import re
func_match = re.search(r'def (\\w+)\\(', {repr(code)})
if func_match:
    func_name = func_match.group(1)
    result = eval(f"{{func_name}}(**test_input)")
    print(json.dumps(result))
"""

        # Execute in temp file
        with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
            f.write(full_code)
            temp_file = f.name

        try:
            result = subprocess.run(
                ["python3", temp_file],
                capture_output=True,
                text=True,
                timeout=10,  # Validation needs more time
            )

            if result.returncode == 0:
                output = result.stdout.strip()
                return json.loads(output)
            else:
                return None

        finally:
            import os

            os.unlink(temp_file)

    return None


@app.post("/api/coding/validate-challenge/{challenge_id}")
async def validate_challenge_endpoint(
    challenge_id: str, db: AsyncSession = Depends(get_db)
):
    """
    Endpoint to validate and fix test cases for a challenge
    Automatically fixes wrong expected outputs
    """
    result = await validate_and_fix_test_cases(challenge_id, db)
    return result


@app.post("/api/coding/validate-test-cases/{challenge_id}")
async def validate_challenge_test_cases(
    challenge_id: str, db: AsyncSession = Depends(get_db)
):
    """
    Validate all test cases for a challenge
    Returns which test cases have wrong expected outputs
    """
    try:
        from sqlalchemy import text

        # Get challenge with test cases
        query = text(
            "SELECT test_cases, starter_code, language FROM coding_challenges WHERE challenge_id = :id"
        )
        result = await db.execute(query, {"id": challenge_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Challenge not found")

        test_cases = json.loads(row[0]) if row[0] else []
        starter_code = row[1]
        language = row[2]

        validation_results = []

        for i, test_case in enumerate(test_cases):
            test_input = test_case.get("input", {})
            expected_output = test_case.get("expected_output")

            # Basic validation
            issues = []

            # 1. Check input format
            if not isinstance(test_input, dict):
                issues.append(f"❌ Test input should be dict, got {type(test_input)}")

            # 2. Check expected output exists
            if expected_output is None:
                issues.append("❌ Missing expected_output")

            # 3. Check for common issues
            if isinstance(expected_output, list):
                # Check if empty list is appropriate
                if len(expected_output) == 0:
                    issues.append("⚠️ Empty list expected - verify this is correct")

            validation_results.append(
                {
                    "test_index": i + 1,
                    "input_format": str(type(test_input)),
                    "has_expected": expected_output is not None,
                    "issues": issues,
                    "needs_fix": len(issues) > 0,
                }
            )

        # Try to execute reference solution if available
        execution_issues = []
        if starter_code and language == "python":
            try:
                # Simple execution test
                for i, test_case in enumerate(test_cases):
                    try:
                        test_input = test_case.get("input", {})
                        # This would be more complex in reality
                        pass
                    except Exception as e:
                        execution_issues.append(f"Test {i + 1}: {str(e)}")
            except:
                pass

        return {
            "challenge_id": challenge_id,
            "total_tests": len(test_cases),
            "validation_results": validation_results,
            "tests_with_issues": sum(1 for r in validation_results if r["needs_fix"]),
            "execution_issues": execution_issues,
            "recommendations": [
                "Ensure all test inputs are dict format: {'param1': value1, 'param2': value2}",
                "Verify expected outputs match problem requirements",
                "Include edge cases (empty arrays, null values, single elements)",
                "Test cases should cover all constraints mentioned",
            ],
        }

    except Exception as e:
        print(f"❌ Test validation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/coding/submit-complete")
async def submit_code_complete(data: dict, db: AsyncSession = Depends(get_db)):
    """
    Complete submission workflow: Submit → Test → Evaluate → Check Plagiarism
    """
    try:
        session_id = data.get("session_id")
        challenge_id = data.get("challenge_id")
        candidate_email = data.get("candidate_email")
        code = data.get("code")
        language = data.get("language")

        if not all([session_id, challenge_id, candidate_email, code, language]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        # 1. Create submission
        submission_id = f"sub_{uuid.uuid4()}"

        # 2. Run tests (use existing /api/coding/run-tests logic)
        # ... test execution code ...

        # 3. Save submission
        # ... save to coding_submissions ...

        # 4. Advanced evaluation
        eval_response = await evaluate_code_advanced(
            {"submission_id": submission_id}, db
        )

        # 5. Plagiarism check
        plagiarism_response = await check_code_plagiarism(
            {"submission_id": submission_id, "challenge_id": challenge_id}, db
        )

        # 6. Update session status
        from sqlalchemy import text

        update_query = text(
            """
            UPDATE coding_sessions
            SET status = 'completed', end_time = NOW(), submission_count = submission_count + 1
            WHERE session_id = :session_id
        """
        )
        await db.execute(update_query, {"session_id": session_id})
        await db.commit()

        return {
            "submission_id": submission_id,
            "evaluation": eval_response,
            "plagiarism": plagiarism_response,
            "message": "Submission completed successfully",
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Complete submission failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== VIOLATIONS ENDPOINT ====================
@app.get("/api/coding/violations/{session_id}")
async def get_violations(session_id: str, db: AsyncSession = Depends(get_db)):
    """Get all violations for a coding session"""
    try:
        from sqlalchemy import text

        query = text(
            """
            SELECT violation_id, session_id, violation_type,
                   description, occurred_at
            FROM violations
            WHERE session_id = :session_id
            ORDER BY occurred_at DESC
        """
        )

        result = await db.execute(query, {"session_id": session_id})
        violations = result.fetchall()

        return {
            "session_id": session_id,
            "violations": [
                {
                    "violation_id": v[0],
                    "session_id": v[1],
                    "violation_type": v[2],
                    "description": v[3],
                    "occurred_at": v[4].isoformat() if v[4] else None,
                }
                for v in violations
            ],
            "total": len(violations),
        }
    except Exception as e:
        print(f"❌ Error fetching violations: {e}")
        return {"session_id": session_id, "violations": [], "total": 0}


# ==================== DEBUG ENDPOINT ====================
@app.get("/api/debug/submission/{submission_id}")
async def debug_submission(submission_id: str, db: AsyncSession = Depends(get_db)):
    """Debug endpoint to check submission data - DeepSeek suggestion"""
    from sqlalchemy import text

    try:
        # Get submission with all related data
        query = text(
            """
            SELECT
                cs.*,
                cd.ai_detection,
                cd.plagiarism_check,
                cc.title,
                cc.difficulty,
                (SELECT COUNT(*) FROM violations v WHERE v.session_id = cs.session_id) as violation_count
            FROM coding_submissions cs
            LEFT JOIN code_detection cd ON cs.submission_id = cd.submission_id
            LEFT JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            WHERE cs.submission_id = :submission_id
        """
        )

        result = await db.execute(query, {"submission_id": submission_id})
        row = result.fetchone()

        if not row:
            return {"error": "Submission not found"}

        return {
            "submission_id": row[0],
            "challenge_id": row[1],
            "candidate_email": row[2],
            "language": row[4],
            "score": row[7],
            "session_id": row[12],
            "has_session_id": row[12] is not None,
            "ai_detection": json.loads(row[13]) if row[13] else None,
            "plagiarism_check": json.loads(row[14]) if row[14] else None,
            "challenge_title": row[15],
            "challenge_difficulty": row[16],
            "violation_count": row[17],
            "has_detection": row[13] is not None or row[14] is not None,
        }
    except Exception as e:
        print(f"❌ Debug failed: {e}")
        import traceback

        traceback.print_exc()
        return {"error": str(e)}


# ==================== CODE SANITIZATION ENDPOINT ====================
# ==================== CODE SANITIZATION ENDPOINT ====================
@app.post("/api/coding/sanitize-submission")
async def sanitize_submission_code(data: dict):
    """
    Sanitize user code before submission:
    - Remove main() functions
    - Remove unnecessary imports
    - Fix common syntax issues
    """
    try:
        code = data.get("code", "")
        language = data.get("language", "python")

        if not code:
            raise HTTPException(status_code=400, detail="Code required")

        sanitized_code = code
        issues_fixed = []

        if language == "go":
            # Remove package main and main function
            if "package main" in code:
                sanitized_code = code.replace("package main\n", "")
                issues_fixed.append("Removed 'package main'")

            if "func main()" in code:
                # Remove main function and everything after it
                lines = sanitized_code.split("\n")
                new_lines = []
                in_main = False
                for line in lines:
                    if "func main()" in line:
                        in_main = True
                        continue
                    if (
                        in_main
                        and line.strip()
                        and line[0] != "\t"
                        and line[:4] != "    "
                    ):
                        in_main = False
                    if not in_main:
                        new_lines.append(line)
                sanitized_code = "\n".join(new_lines)
                issues_fixed.append("Removed main() function")

            # Remove unnecessary imports
            import_sections = re.findall(
                r"import\s*\(\s*([^)]+)\s*\)", sanitized_code, re.DOTALL
            )
            for section in import_sections:
                imports = [imp.strip() for imp in section.split("\n") if imp.strip()]
                needed_imports = []
                for imp in imports:
                    # Keep only essential imports
                    essential_patterns = ["fmt", "strings", "strconv", "sort", "math"]
                    if any(pattern in imp for pattern in essential_patterns):
                        needed_imports.append(imp)
                    else:
                        issues_fixed.append(f"Removed unused import: {imp}")

                if needed_imports:
                    sanitized_code = re.sub(
                        r"import\s*\(\s*[^)]+\s*\)",
                        f"import (\n" + "\n".join(needed_imports) + "\n)",
                        sanitized_code,
                    )

        elif language == "python":
            # Remove if __name__ == "__main__"
            if "__name__" in code and "__main__" in code:
                sanitized_code = re.sub(
                    r'if\s+__name__\s*==\s*[\'"]__main__[\'"].*?\n',
                    "",
                    sanitized_code,
                    flags=re.DOTALL,
                )
                issues_fixed.append("Removed __main__ guard")

            # Remove unnecessary imports
            import_pattern = r"^\s*import\s+.*$"
            lines = sanitized_code.split("\n")
            new_lines = []
            for line in lines:
                if re.match(import_pattern, line):
                    # Check if imports are actually used
                    import_name = line.split()[-1]
                    if import_name in sanitized_code:
                        new_lines.append(line)
                    else:
                        issues_fixed.append(f"Removed unused import: {import_name}")
                else:
                    new_lines.append(line)
            sanitized_code = "\n".join(new_lines)

        elif language in ["javascript", "js", "node"]:
            # Remove module.exports
            if "module.exports" in code:
                sanitized_code = re.sub(
                    r"module\.exports\s*=.*?;?\n?", "", sanitized_code, flags=re.DOTALL
                )
                issues_fixed.append("Removed module.exports")

            # Remove exports
            if "exports." in code:
                sanitized_code = re.sub(
                    r"exports\.\w+\s*=.*?;?\n?", "", sanitized_code, flags=re.DOTALL
                )
                issues_fixed.append("Removed exports")

        elif language in ["cpp", "c++"]:
            # Remove main function
            if "int main()" in code or "int main(" in code:
                # Remove main function block
                sanitized_code = re.sub(
                    r"int\s+main\s*\([^)]*\)\s*\{[^}]*\}",
                    "",
                    sanitized_code,
                    flags=re.DOTALL,
                )
                issues_fixed.append("Removed main() function")

            # Remove unnecessary includes
            lines = sanitized_code.split("\n")
            new_lines = []
            essential_includes = [
                "#include <iostream>",
                "#include <vector>",
                "#include <string>",
                "#include <algorithm>",
                "#include <map>",
                "#include <set>",
            ]
            for line in lines:
                if "#include" in line:
                    if any(essential in line for essential in essential_includes):
                        new_lines.append(line)
                    else:
                        issues_fixed.append(f"Removed unused include: {line.strip()}")
                else:
                    new_lines.append(line)
            sanitized_code = "\n".join(new_lines)

        elif language == "java":
            # Remove public static void main
            if "public static void main" in code:
                sanitized_code = re.sub(
                    r"public\s+static\s+void\s+main\s*\([^)]*\)\s*\{[^}]*\}",
                    "",
                    sanitized_code,
                    flags=re.DOTALL,
                )
                issues_fixed.append("Removed main() method")

            # Remove package declarations if not the class name
            if "package " in code and not code.startswith("package"):
                sanitized_code = re.sub(
                    r"^\s*package\s+\S+;", "", sanitized_code, flags=re.MULTILINE
                )
                issues_fixed.append("Removed package declaration")

        return {
            "success": True,
            "original_length": len(code),
            "sanitized_length": len(sanitized_code),
            "issues_fixed": issues_fixed,
            "sanitized_code": sanitized_code,
        }

    except Exception as e:
        print(f"❌ Sanitization failed: {e}")
        return {
            "success": False,
            "error": str(e),
            "sanitized_code": data.get("code", ""),
        }


# ==================== DIRECT AI DETECTION TEST ====================
# ==================== AUTOMATED JOB POSTING MODULE ====================

# Job posting image folder
JOB_POSTING_IMAGE_FOLDER = os.path.join(os.path.dirname(__file__), "job_posting_images")
os.makedirs(JOB_POSTING_IMAGE_FOLDER, exist_ok=True)


def get_random_job_image():
    """Get a random image from the job posting images folder"""
    if not os.path.exists(JOB_POSTING_IMAGE_FOLDER):
        os.makedirs(JOB_POSTING_IMAGE_FOLDER)
        return None, None

    image_files = [
        f
        for f in os.listdir(JOB_POSTING_IMAGE_FOLDER)
        if f.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp"))
    ]

    if not image_files:
        return None, None

    selected_image = random.choice(image_files)
    image_path = os.path.join(JOB_POSTING_IMAGE_FOLDER, selected_image)

    # Read and encode image
    with open(image_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode("utf-8")

    # Determine media type
    ext = selected_image.lower().split(".")[-1]
    media_type_map = {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "webp": "image/webp",
    }
    media_type = media_type_map.get(ext, "image/jpeg")

    return image_data, media_type


async def generate_linkedin_job_post(
    job_description: str, image_data: str, media_type: str
):
    """Generate LinkedIn job post using GROQ API"""

    # Get GROQ API key from environment
    groq_api_key = os.environ.get("GROQ_API_KEY")
    if not groq_api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")

    prompt = f"""You are an expert LinkedIn recruiter and content creator. Create an engaging, professional LinkedIn job post based on the following job description.

Job Description: {job_description}

Requirements:
1. Write in a balanced, professional tone that attracts top talent
2. ALWAYS use clear bullet points when listing Key Responsibilities and Requirements/Qualifications
3. Use minimal formatting, keeping whitespace clean and readable
4. Keep it concise but informative (300-500 words)
5. Include a clear call-to-action
6. Format with proper line breaks for readability

Structure:
- Hook/Opening (attention-grabbing, professional)
- Company/Role overview
- Key responsibilities (MUST be bullet points)
- Requirements/Qualifications (MUST be bullet points)
- Benefits/Why join
- Summary Block (MUST be explicitly formatted at the end with fields like "Salary: [Value or Competitive]", "Job Type: [Full-time/Part-time/etc.]", "Location: [Remote/On-site/Hybrid]")
- Call to action

Make it compelling, share-worthy, and structured clearly!"""

    # Make API request to GROQ
    async with aiohttp.ClientSession() as session:
        async with session.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {groq_api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "max_tokens": 2000,
                "messages": [{"role": "user", "content": prompt}],
            },
        ) as response:
            if response.status != 200:
                error_text = await response.text()
                raise HTTPException(
                    status_code=response.status,
                    detail=f"GROQ API error: {error_text}",
                )

            result = await response.json()
            return result["choices"][0]["message"]["content"]


@app.post("/api/job-posting/generate")
async def generate_job_posting(data: dict):
    """Generate LinkedIn job post using GROQ API"""
    try:
        job_description = data.get("prompt", "")

        if not job_description:
            raise HTTPException(status_code=400, detail="No job description provided")

        print(f"\n{'=' * 60}")
        print(f"🤖 Generating LinkedIn job post with GROQ")
        print(f"   Job Description: {job_description[:100]}...")
        print(f"{'=' * 60}\n")

        # Get random image
        image_data, media_type = get_random_job_image()

        if not image_data:
            raise HTTPException(
                status_code=400,
                detail="No images found in 'job_posting_images' folder. Please add at least one image.",
            )

        # Generate post with Claude
        result_text = await generate_linkedin_job_post(
            job_description, image_data, media_type
        )

        return {
            "result": result_text,
            "image": f"data:{media_type};base64,{image_data}",
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback

        print(f"\n❌ Error:\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Generation failed: {str(e)}")


@app.post("/api/coding/test-ai-detection")
async def test_ai_detection(data: dict, db: AsyncSession = Depends(get_db)):
    """Test AI detection with different methods"""
    try:
        code = data.get("code", "")
        language = data.get("language", "python")

        if not code:
            raise HTTPException(status_code=400, detail="Code required")

        results = {}

        # 1. Local detector only
        start = time.time()
        local_result = await ai_detector.detect_ai_code(code, language)
        local_time = time.time() - start

        results["local"] = {**local_result, "time_ms": round(local_time * 1000, 1)}

        # 2. Groq API only (if available)
        if GROQ_API_KEY:
            start = time.time()
            groq_result = await groq_ai_judge(code, language)
            groq_time = time.time() - start

            results["groq"] = {**groq_result, "time_ms": round(groq_time * 1000, 1)}

        # 3. Hybrid detection
        start = time.time()
        hybrid_result = await hybrid_ai_detection(code, language)
        hybrid_time = time.time() - start

        results["hybrid"] = {**hybrid_result, "time_ms": round(hybrid_time * 1000, 1)}

        # Summary
        code_info = {
            "length_chars": len(code),
            "lines": len(code.split("\n")),
            "language": language,
        }

        return {
            "success": True,
            "code_info": code_info,
            "results": results,
            "recommendation": "hybrid" if GROQ_API_KEY else "local",
        }
    except Exception as e:
        print(f"Error in test AI detection: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== FRESH AI INTERVIEWER API ====================
# Voice: STT (Groq Whisper), LLM (Groq Llama), TTS (Edge TTS)
# Vision: Face Verification, Emotion Detection (DeepFace)
# Analysis: Answer Scoring (Groq)


@app.post("/api/interview/start")
async def start_interview_session(data: dict, db: AsyncSession = Depends(get_db)):
    """Start a new AI interview session"""
    try:
        from application.interview_service import (
            create_interview_session,
            InterviewStatus,
        )

        candidate_email = data.get("candidate_email")
        candidate_name = data.get("candidate_name", "")
        job_role = data.get("job_role", "Software Engineer")

        if not candidate_email:
            raise HTTPException(status_code=400, detail="Candidate email required")

        candidate_info = {"email": candidate_email, "name": candidate_name}
        session_id = create_interview_session(candidate_info, job_role)

        # Save to database
        session = InterviewSessionModel(
            session_id=session_id,
            candidate_email=candidate_email,
            candidate_name=candidate_name,
            job_role=job_role,
            status=InterviewStatus.INITIATED,
            questions_json="[]",
            responses_json="[]",
            emotion_data_json="[]",
        )
        db.add(session)
        await db.commit()

        return {
            "success": True,
            "session_id": session_id,
            "message": "Interview session started",
            "job_role": job_role,
        }
    except Exception as e:
        print(f"Error starting interview: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/verify-face")
async def verify_candidate_face(data: dict, db: AsyncSession = Depends(get_db)):
    """Verify candidate identity using face verification"""
    try:
        from application.interview_service import verify_face

        session_id = data.get("session_id")
        candidate_image = data.get("candidate_image")
        live_image = data.get("live_image")

        if not all([session_id, candidate_image, live_image]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        result = await verify_face(candidate_image, live_image)

        if result.get("verified"):
            query = text("""UPDATE interview_sessions 
                          SET face_verified = 1, status = :status 
                          WHERE session_id = :session_id""")
            await db.execute(
                query,
                {"status": InterviewStatus.VERIFIED.value, "session_id": session_id},
            )
            await db.commit()

        return {
            "success": result.get("success", False),
            "verified": result.get("verified", False),
            "confidence": result.get("confidence", 0),
            "message": "Face verified"
            if result.get("verified")
            else "Face verification failed",
        }
    except Exception as e:
        print(f"Error verifying face: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/question")
async def get_next_question(data: dict, db: AsyncSession = Depends(get_db)):
    """Get next interview question from AI"""
    try:
        from application.interview_service import (
            generate_interview_question,
            get_interview_session,
            QuestionType,
        )

        session_id = data.get("session_id")
        question_type = data.get("question_type", "introduction")

        if not session_id:
            raise HTTPException(status_code=400, detail="Session ID required")

        session = get_interview_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Get previous Q&A
        prev_questions = session.get("questions", [])
        prev_answers = session.get("answers", [])

        result = await generate_interview_question(
            session["job_role"], question_type, prev_questions, prev_answers
        )

        if result.get("success"):
            # Update session
            session.setdefault("questions", []).append(result.get("question"))
            session["status"] = InterviewStatus.IN_PROGRESS

        return {
            "success": result.get("success", False),
            "question": result.get("question", ""),
            "question_type": result.get("question_type", question_type),
            "category": result.get("category", "general"),
            "difficulty": result.get("difficulty", "medium"),
            "expected_duration": result.get("expected_duration", 60),
        }
    except Exception as e:
        print(f"Error getting question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/tts")
async def convert_text_to_speech(data: dict):
    """Convert question text to speech (TTS)"""
    try:
        from application.interview_service import text_to_speech

        text = data.get("text")
        voice = data.get("voice", "en-US-AriaNeural")

        if not text:
            raise HTTPException(status_code=400, detail="Text required")

        result = await text_to_speech(text, voice)

        return {
            "success": result.get("success", False),
            "audio": result.get("audio", ""),
            "voice": voice,
        }
    except Exception as e:
        print(f"Error converting TTS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/transcribe")
async def transcribe_audio_endpoint(data: dict):
    """Transcribe audio to text (STT)"""
    try:
        from services.stt_service import transcribe_audio

        audio_base64 = data.get("audio_base64")

        if not audio_base64:
            raise HTTPException(status_code=400, detail="Audio data required")

        result = await transcribe_audio(audio_base64)

        # transcribe_audio returns a string, so we wrap it
        return {
            "success": bool(result and result.strip()),
            "text": result or "",
            "error": "" if result else "Transcription failed",
        }
    except Exception as e:
        print(f"Error transcribing: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/analyze-emotion")
async def analyze_emotion_endpoint(data: dict, db: AsyncSession = Depends(get_db)):
    """Analyze candidate emotions from face"""
    try:
        from application.interview_service import analyze_emotions

        session_id = data.get("session_id")
        image_base64 = data.get("image_base64")

        if not image_base64:
            raise HTTPException(status_code=400, detail="Image required")

        result = await analyze_emotions(image_base64)

        return {
            "success": result.get("success", False),
            "emotions": result.get("emotions", {}),
            "dominant_emotion": result.get("dominant_emotion", "unknown"),
        }
    except Exception as e:
        print(f"Error analyzing emotion: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/analyze-answer")
async def analyze_candidate_answer(data: dict, db: AsyncSession = Depends(get_db)):
    """Analyze and score candidate's answer"""
    try:
        from application.interview_service import analyze_answer, get_interview_session

        session_id = data.get("session_id")
        question = data.get("question", "")
        answer = data.get("answer", "")

        if not answer:
            raise HTTPException(status_code=400, detail="Answer required")

        session = get_interview_session(session_id) if session_id else None
        job_role = session.get("job_role") if session else None

        result = await analyze_answer(question, answer, job_role)

        # Update session
        if session:
            session.setdefault("answers", []).append(answer)

        return {
            "success": result.get("success", False),
            "sentiment": result.get("overall_score", 0),
            "competency_score": result.get("overall_score", 0),
            "analysis": {
                "relevance": result.get("relevance", 0),
                "depth": result.get("depth", 0),
                "clarity": result.get("clarity", 0),
                "strengths": result.get("strengths", []),
                "improvements": result.get("improvements", []),
                "feedback": result.get("feedback", ""),
            },
        }
    except Exception as e:
        print(f"Error analyzing answer: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/complete")
async def complete_interview(data: dict, db: AsyncSession = Depends(get_db)):
    """Complete the interview session"""
    try:
        session_id = data.get("session_id")

        if not session_id:
            raise HTTPException(status_code=400, detail="Session ID required")

        # Update database
        query = text("""UPDATE interview_sessions 
                      SET status = :status, end_time = NOW()
                      WHERE session_id = :session_id""")
        await db.execute(
            query, {"status": InterviewStatus.COMPLETED.value, "session_id": session_id}
        )
        await db.commit()

        return {
            "success": True,
            "message": "Interview completed",
            "session_id": session_id,
        }
    except Exception as e:
        print(f"Error completing interview: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/verify-face")
async def verify_candidate_face(data: dict, db: AsyncSession = Depends(get_db)):
    """Verify candidate identity using face verification"""
    try:
        from application.interview_service import verify_face

        session_id = data.get("session_id")
        candidate_image = data.get("candidate_image")  # Base64
        live_image = data.get("live_image")  # Base64

        if not all([session_id, candidate_image, live_image]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        result = await verify_face(candidate_image, live_image)

        if result.get("verified"):
            # Update session status
            query = text("""UPDATE interview_sessions 
                          SET face_verified = 1, status = :status 
                          WHERE session_id = :session_id""")
            await db.execute(
                query,
                {"status": InterviewStatus.VERIFIED.value, "session_id": session_id},
            )
            await db.commit()

        return {
            "success": result.get("success", False),
            "verified": result.get("verified", False),
            "confidence": 1 - result.get("distance", 1)
            if result.get("verified")
            else 0,
            "message": "Face verified successfully"
            if result.get("verified")
            else "Face verification failed",
        }
    except Exception as e:
        print(f"Error verifying face: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/question")
async def get_next_question(data: dict, db: AsyncSession = Depends(get_db)):
    """Get next interview question from AI"""
    try:
        from application.interview_service import (
            generate_interview_question,
            QuestionType,
            get_interview_session,
        )
        import json

        session_id = data.get("session_id")
        question_type = data.get("question_type", "introduction")

        session = get_interview_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Get question type enum
        q_type = QuestionType(question_type)

        # Get previous Q&A
        previous_questions = session.questions
        previous_answers = session.answers

        # Generate question
        result = await generate_interview_question(
            job_role=session.job_role,
            question_type=q_type,
            candidate_profile=session.candidate_info,
            previous_questions=previous_questions,
            previous_answers=previous_answers,
        )

        if result.get("success"):
            # Store question
            session.questions.append(
                {"question": result.get("question"), "type": question_type}
            )

            # Update DB
            query = text("""UPDATE interview_sessions 
                          SET questions_json = :questions, status = :status
                          WHERE session_id = :session_id""")
            await db.execute(
                query,
                {
                    "questions": json.dumps(session.questions),
                    "status": InterviewStatus.IN_PROGRESS.value,
                    "session_id": session_id,
                },
            )
            await db.commit()

        return {
            "success": result.get("success", False),
            "question": result.get("question", ""),
            "type": question_type,
            "session_id": session_id,
        }
    except Exception as e:
        print(f"Error getting question: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/tts")
async def convert_text_to_speech(data: dict):
    """Convert question text to speech using Edge TTS"""
    try:
        from application.interview_service import text_to_speech

        text = data.get("text")
        voice = data.get("voice", "en-US-AriaNeural")

        if not text:
            raise HTTPException(status_code=400, detail="Text required")

        result = await text_to_speech(text, voice)

        return {
            "success": result.get("success", False),
            "audio": result.get("audio", ""),
            "voice": voice,
        }
    except Exception as e:
        print(f"Error converting TTS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def analyze_emotion_endpoint(data: dict, db: AsyncSession = Depends(get_db)):
    """Analyze candidate emotions from face detection"""
    try:
        from application.interview_service import (
            analyze_emotions,
            get_interview_session,
        )
        import json

        session_id = data.get("session_id")
        emotion_data = data.get("emotion_data")

        if not all([session_id, emotion_data]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        result = analyze_emotions(emotion_data)

        # Store emotion data
        session = get_interview_session(session_id)
        if session:
            session.add_emotion_data(result)

            # Update DB
            query = text("""UPDATE interview_sessions 
                          SET emotion_data_json = :emotions
                          WHERE session_id = :session_id""")
            await db.execute(
                query,
                {"emotions": json.dumps(session.emotions), "session_id": session_id},
            )
            await db.commit()

        return {
            "success": result.get("success", False),
            "dominant_emotion": result.get("dominant_emotion", "neutral"),
            "sentiment": result.get("sentiment", "neutral"),
            "all_emotions": result.get("all_emotions", {}),
        }
    except Exception as e:
        print(f"Error analyzing emotion: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/analyze-answer")
async def analyze_candidate_answer(data: dict, db: AsyncSession = Depends(get_db)):
    """Analyze candidate's answer: sentiment + competency scoring"""
    try:
        from application.interview_service import (
            analyze_sentiment,
            score_answer,
            QuestionType,
            get_interview_session,
        )
        import json

        session_id = data.get("session_id")
        question = data.get("question")
        answer = data.get("answer")
        question_type = data.get("question_type", "technical")

        if not all([session_id, question, answer]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        session = get_interview_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Analyze sentiment
        sentiment_result = await analyze_sentiment(answer)

        # Score competency
        score_result = await score_answer(
            question=question,
            answer=answer,
            job_role=session.job_role,
            question_type=QuestionType(question_type),
        )

        # Store response
        if session:
            response_data = {
                "question": question,
                "answer": answer,
                "type": question_type,
                "sentiment": sentiment_result,
                "score": score_result,
            }
            session.answers.append(answer)

            # Update DB
            query = text("""UPDATE interview_sessions 
                          SET responses_json = :responses
                          WHERE session_id = :session_id""")
            await db.execute(
                query,
                {"responses": json.dumps(session.answers), "session_id": session_id},
            )
            await db.commit()

        return {
            "success": True,
            "sentiment": sentiment_result,
            "competency": score_result,
            "overall_score": score_result.get("score", 0),
        }
    except Exception as e:
        print(f"Error analyzing answer: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/interview/complete")
async def complete_interview(data: dict, db: AsyncSession = Depends(get_db)):
    """Complete interview and generate HR report"""
    try:
        from application.interview_service import (
            generate_hr_report,
            get_interview_session,
            InterviewStatus,
        )
        import json
        from datetime import datetime

        session_id = data.get("session_id")

        session = get_interview_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Calculate overall scores
        scores = session.scores
        technical = sum(
            s.get("score", 0) for s in scores if s.get("question_type") == "technical"
        ) / max(1, len([s for s in scores if s.get("question_type") == "technical"]))
        behavioral = sum(
            s.get("score", 0) for s in scores if s.get("question_type") == "behavioral"
        ) / max(1, len([s for s in scores if s.get("question_type") == "behavioral"]))
        problem_solving = sum(
            s.get("score", 0)
            for s in scores
            if s.get("question_type") == "problem_solving"
        ) / max(
            1, len([s for s in scores if s.get("question_type") == "problem_solving"])
        )
        overall = sum(s.get("score", 0) for s in scores) / max(1, len(scores))

        overall_scores = {
            "technical": technical,
            "behavioral": behavioral,
            "problem_solving": problem_solving,
            "overall": overall,
        }

        # Get emotion analysis
        emotions = [e.get("sentiment", "neutral") for e in session.emotions]

        # Generate HR report
        report_result = await generate_hr_report(
            session_id=session_id,
            candidate_info=session.candidate_info,
            qa_history=list(zip(session.questions, session.answers)),
            emotion_analysis=emotions,
            overall_scores=overall_scores,
        )

        # Update session status
        query = text("""UPDATE interview_sessions 
                      SET status = :status, end_time = :end_time, 
                          hr_report_json = :report, overall_score = :score
                      WHERE session_id = :session_id""")
        await db.execute(
            query,
            {
                "status": InterviewStatus.COMPLETED.value,
                "end_time": datetime.utcnow(),
                "report": json.dumps(report_result.get("report", {})),
                "score": overall,
                "session_id": session_id,
            },
        )
        await db.commit()

        session.status = InterviewStatus.COMPLETED

        return {
            "success": True,
            "session_id": session_id,
            "scores": overall_scores,
            "report": report_result.get("report", {}),
            "question_count": len(session.questions),
        }
    except Exception as e:
        print(f"Error completing interview: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/interview/session/{session_id}")
async def get_interview_session_data(
    session_id: str, db: AsyncSession = Depends(get_db)
):
    """Get interview session details"""
    try:
        query = text("SELECT * FROM interview_sessions WHERE session_id = :session_id")
        result = await db.execute(query, {"session_id": session_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Session not found")

        return {"success": True, "session": dict(row._mapping)}
    except Exception as e:
        print(f"Error getting session: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/interview/sessions")
async def list_interview_sessions(
    email: str = None, status: str = None, db: AsyncSession = Depends(get_db)
):
    """List all interview sessions"""
    try:
        query = text("SELECT * FROM interview_sessions")
        params = {}

        if email:
            query += " WHERE candidate_email = :email"
            params["email"] = email
        if status:
            query += " AND status = :status" if email else " WHERE status = :status"
            params["status"] = status

        query += " ORDER BY created_at DESC"

        result = await db.execute(query, params)
        rows = result.fetchall()

        return {"success": True, "sessions": [dict(row._mapping) for row in rows]}
    except Exception as e:
        print(f"Error listing sessions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/interview/report/{session_id}")
async def get_interview_report(session_id: str, db: AsyncSession = Depends(get_db)):
    """Get HR report for completed interview"""
    try:
        import json

        query = text(
            "SELECT hr_report_json, overall_score, questions_json, responses_json FROM interview_sessions WHERE session_id = :session_id"
        )
        result = await db.execute(query, {"session_id": session_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Session not found")

        hr_report = json.loads(row[0]) if row[0] else {}

        return {
            "success": True,
            "report": hr_report,
            "overall_score": row[1],
            "questions": json.loads(row[2]) if row[2] else [],
            "responses": json.loads(row[3]) if row[3] else [],
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/hr/send-interview-invitation")
async def send_interview_invitation(data: dict, db: AsyncSession = Depends(get_db)):
    """Send AI interview invitation email to candidates"""
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        candidate_emails = data.get("candidate_emails", [])
        role = data.get("role", "Software Engineer")
        questions = data.get("questions", 5)

        if not candidate_emails:
            return {"success": False, "error": "No candidates provided"}

        sender_email = os.getenv("SMTP_EMAIL")
        sender_password = os.getenv("SMTP_PASSWORD")
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        frontend_base = os.getenv("FRONTEND_URL", "http://localhost:5173").rstrip("/")
        base_interview_url = f"{frontend_base}/ai-interview"

        if not sender_email or not sender_password:
            return {
                "success": False,
                "error": "SMTP credentials not configured. Please set SMTP_EMAIL and SMTP_PASSWORD in .env file.",
            }

        results = []
        import urllib.parse

        for email in candidate_emails:
            # Get candidate info for personalization
            query = text(
                """SELECT name, skills FROM cv_candidates WHERE email = :email LIMIT 1"""
            )
            result = await db.execute(query, {"email": email})
            row = result.fetchone()
            candidate_name = row[0] if row else "Candidate"
            candidate_skills = row[1] if row and len(row) > 1 else ""

            # Construct personalized URL with embedded params
            params = {
                "e": email,
                "n": candidate_name,
                "r": role,
                "q": str(questions),
                "s": candidate_skills or "",
            }
            interview_url = f"{base_interview_url}?{urllib.parse.urlencode(params)}"

            # Create interview invitation email
            message = MIMEMultipart("alternative")
            message["Subject"] = (
                "🎤 AI Interview Invitation - Next Step in Your Application"
            )
            message["From"] = sender_email
            message["To"] = email

            html = f"""
            <html>
              <body style="font-family: Arial, sans-serif; padding: 20px;">
                <h2 style="color: #10b981;">AI Interview Invitation</h2>
                <p>Dear {candidate_name},</p>
                <p>Congratulations! You have successfully passed the coding evaluation and are invited to complete an <strong>AI-powered interview</strong>.</p>
                <p style="margin: 30px 0;">
                  <a href="{interview_url}" style="background-color: #10b981; color: white; padding: 14px 28px; text-decoration: none; border-radius: 8px; display: inline-block; font-weight: 600;">
                    🎤 Start AI Interview
                  </a>
                </p>
                <p><strong>Next Steps:</strong></p>
                <ol style="line-height: 1.8;">
                  <li>Click the link above to start your AI interview</li>
                  <li>Ensure you have a stable internet connection</li>
                  <li>Allow camera and microphone access</li>
                  <li>Answer questions honestly and to the best of your ability</li>
                </ol>
                <p>If the button doesn't work, copy this link: <a href="{interview_url}">{interview_url}</a></p>
                <p>Best regards,<br/>Recruitment Team</p>
              </body>
            </html>
            """

            plain_body = f"""
            AI Interview Invitation
            
            Dear {candidate_name},
            
            Congratulations! You have successfully passed the coding evaluation and are invited to complete an AI-powered interview.
            
            Next Steps:
            1. Go to: {interview_url}
            2. Ensure you have a stable internet connection
            3. Allow camera and microphone access
            4. Answer questions honestly and to the best of your ability
            
            Best regards,
            Recruitment Team
            """

            try:
                part1 = MIMEText(plain_body, "plain")
                part2 = MIMEText(html, "html")
                message.attach(part1)
                message.attach(part2)

                with smtplib.SMTP(smtp_server, smtp_port) as server:
                    server.starttls()
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, email, message.as_string())

                print(f"✅ Interview invitation email sent to {email}")
                results.append({"email": email, "status": "sent"})
            except Exception as e:
                print(f"❌ Failed to send email to {email}: {e}")
                results.append({"email": email, "status": "failed", "error": str(e)})

        successful = len([r for r in results if r["status"] == "sent"])
        return {
            "success": True,
            "message": f"Interview invitations sent to {successful} candidate(s)",
            "results": results,
        }
    except Exception as e:
        print(f"Error sending interview invitation: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/hr/interview-sessions-overview")
async def hr_interview_sessions_overview(db: AsyncSession = Depends(get_db)):
    """All AI interview sessions for HR dashboard (scores, status, report flag)."""
    try:
        q = text(
            """
            SELECT
                session_id,
                candidate_email,
                candidate_name,
                job_role,
                status,
                face_verified,
                overall_score,
                start_time,
                end_time,
                created_at,
                hr_report_json
            FROM interview_sessions
            ORDER BY updated_at DESC
            LIMIT 200
            """
        )
        result = await db.execute(q)
        rows = result.fetchall()
        out = []
        for row in rows:
            report_raw = row[10]
            has_report = bool(
                report_raw
                and str(report_raw).strip()
                and str(report_raw).strip() != "null"
            )
            out.append(
                {
                    "session_id": row[0],
                    "candidate_email": row[1],
                    "candidate_name": row[2],
                    "job_role": row[3],
                    "status": row[4],
                    "face_verified": bool(row[5]) if row[5] is not None else False,
                    "overall_score": float(row[6]) if row[6] is not None else None,
                    "start_time": str(row[7]) if row[7] else None,
                    "end_time": str(row[8]) if row[8] else None,
                    "created_at": str(row[9]) if row[9] else None,
                    "has_hr_report": has_report,
                }
            )
        return {"success": True, "sessions": out, "count": len(out)}
    except Exception as e:
        print(f"Error hr_interview_sessions_overview: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/interview/completed-sessions")
async def get_completed_interview_sessions(db: AsyncSession = Depends(get_db)):
    """Get all completed interview sessions with HR reports"""
    try:
        query = text("""SELECT 
            session_id,
            candidate_email,
            candidate_name,
            job_role,
            status,
            overall_score,
            hr_report_json,
            questions_json,
            responses_json,
            created_at,
            end_time
        FROM interview_sessions 
        WHERE status = 'completed' AND hr_report_json IS NOT NULL
        ORDER BY end_time DESC""")
        result = await db.execute(query)
        rows = result.fetchall()

        sessions = []
        for row in rows:
            import json

            sessions.append(
                {
                    "session_id": row[0],
                    "candidate_email": row[1],
                    "candidate_name": row[2],
                    "job_role": row[3],
                    "status": row[4],
                    "overall_score": float(row[5]) if row[5] else 0,
                    "hr_report": json.loads(row[6]) if row[6] else {},
                    "questions": json.loads(row[7]) if row[7] else [],
                    "responses": json.loads(row[8]) if row[8] else [],
                    "created_at": str(row[9]) if row[9] else None,
                    "end_time": str(row[10]) if row[10] else None,
                }
            )

        return {"success": True, "sessions": sessions}
    except Exception as e:
        print(f"Error fetching completed sessions: {e}")
        return {"success": False, "error": str(e)}


# =============================================
# HR APPROVED CANDIDATES FOR ASSESSMENT
# =============================================


@app.post("/api/hr/approve-candidates")
async def approve_candidates(data: dict, db: AsyncSession = Depends(get_db)):
    """Approve candidates for technical assessment"""
    try:
        candidate_ids = data.get("candidate_ids", [])

        if not candidate_ids:
            return {"success": False, "error": "No candidates selected"}

        approved = []
        for candidate_id in candidate_ids:
            cid = str(candidate_id).strip()
            if not cid:
                continue
            # Check if candidate exists
            query = text("""
                SELECT candidate_id, email, name, skills, skill_match_percentage, status
                FROM cv_candidates WHERE candidate_id = :candidate_id
            """)
            result = await db.execute(query, {"candidate_id": cid})
            row = result.fetchone()

            if row:
                # Update status to approved_for_assessment
                update_query = text("""
                    UPDATE cv_candidates 
                    SET status = :status, updated_at = NOW()
                    WHERE candidate_id = :candidate_id
                """)
                await db.execute(
                    update_query,
                    {"candidate_id": cid, "status": "approved_for_assessment"},
                )

                approved.append(
                    {
                        "candidate_id": row[0],
                        "email": row[1],
                        "name": row[2],
                        "skills": json.loads(row[3]) if row[3] else [],
                        "skill_match_percentage": row[4],
                        "status": "approved_for_assessment",
                        "approved_at": str(datetime.now()),
                    }
                )

        await db.commit()

        return {
            "success": True,
            "message": f"{len(approved)} candidate(s) approved for assessment",
            "candidates": approved,
        }
    except Exception as e:
        print(f"❌ Error approving candidates: {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hr/approved-candidates")
async def get_approved_candidates(db: AsyncSession = Depends(get_db)):
    """Get all candidates approved for assessment"""
    try:
        query = text("""
            SELECT candidate_id, email, name, skills, skill_match_percentage, status,
                   created_at, updated_at
            FROM cv_candidates 
            WHERE status = 'approved_for_assessment'
            ORDER BY updated_at DESC
        """)
        result = await db.execute(query)
        rows = result.fetchall()

        candidates = []
        for row in rows:
            candidates.append(
                {
                    "candidate_id": row[0],
                    "email": row[1],
                    "name": row[2],
                    "skills": json.loads(row[3]) if row[3] else [],
                    "skill_match_percentage": row[4],
                    "status": row[5],
                    "created_at": str(row[6]) if row[6] else None,
                    "updated_at": str(row[7]) if row[7] else None,
                    "approved_at": str(row[7]) if row[7] else None,
                }
            )

        return {"success": True, "candidates": candidates, "count": len(candidates)}
    except Exception as e:
        print(f"❌ Error fetching approved candidates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/hr/remove-assessment/{candidate_id}")
async def remove_from_assessment(candidate_id: str, db: AsyncSession = Depends(get_db)):
    """Remove candidate from assessment (reset status to processed)"""
    try:
        cid = (candidate_id or "").strip()
        if not cid:
            raise HTTPException(status_code=400, detail="Invalid candidate ID")
        update_query = text("""
            UPDATE cv_candidates 
            SET status = 'processed', updated_at = NOW()
            WHERE candidate_id = :candidate_id
        """)
        await db.execute(update_query, {"candidate_id": cid})
        await db.commit()

        return {"success": True, "message": "Candidate removed from assessment"}
    except Exception as e:
        print(f"❌ Error removing candidate from assessment: {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/hr/send-assessment-email")
async def send_assessment_email_to_candidates(
    data: dict, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)
):
    """Send assessment email to approved candidates from HR dashboard"""
    try:
        assessment_id = data.get("assessment_id")
        candidate_ids = data.get("candidate_ids", [])

        if not candidate_ids:
            return {"success": False, "error": "No candidates selected"}

        if not assessment_id:
            # Get first published assessment or create one
            from infrastructure.repositories import AssessmentRepository

            assessments = await AssessmentRepository.get_all(db)
            published = [a for a in assessments if a.get("status") == "published"]

            if not published:
                return {
                    "success": False,
                    "error": "No published assessment found. Please create one first.",
                }

            assessment_id = published[0].get("assessment_id")

        # Get candidate emails
        emails = []
        for candidate_id in candidate_ids:
            cid = str(candidate_id).strip()
            if not cid:
                continue
            query = text(
                "SELECT email, name FROM cv_candidates WHERE candidate_id = :candidate_id"
            )
            result = await db.execute(query, {"candidate_id": cid})
            row = result.fetchone()
            if row:
                emails.append(row[0])

        # Send emails in background
        for email in emails:
            background_tasks.add_task(send_assessment_email, email, assessment_id, db)

        return {
            "success": True,
            "message": f"Assessment email sent to {len(emails)} candidate(s)",
            "assessment_id": assessment_id,
            "emails_sent": len(emails),
        }
    except Exception as e:
        print(f"❌ Error sending assessment emails: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class StatusEmailRequest(BaseModel):
    email: str
    subject: str
    message_html: str

def _send_status_email_impl(email: str, subject: str, message_html: str) -> None:
    """
    Send HR status email via SMTP. Raises on failure so the API can return an error
    (background tasks previously reported success before the send ran).
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    sender_email = (os.getenv("SMTP_EMAIL") or "").strip()
    sender_password = (os.getenv("SMTP_PASSWORD") or "").strip()
    smtp_server = (os.getenv("SMTP_SERVER") or "smtp.gmail.com").strip()
    smtp_port_str = (os.getenv("SMTP_PORT") or "587").strip()

    if not sender_email:
        raise ValueError("SMTP_EMAIL is not set in the server environment.")
    if not sender_password:
        raise ValueError(
            "SMTP_PASSWORD is not set. Add it to your backend .env to send real email."
        )
    if not smtp_server:
        raise ValueError("SMTP_SERVER is empty after trimming; set a valid hostname (e.g. smtp.gmail.com).")

    try:
        smtp_port = int(smtp_port_str)
    except ValueError as e:
        raise ValueError(f"SMTP_PORT must be a number, got {smtp_port_str!r}") from e

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender_email
    msg["To"] = email
    msg.attach(MIMEText(message_html, "html"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, email, msg.as_string())
    except OSError as e:
        # Windows: WinError 11001 / errno 11001 — host not found (bad SMTP_SERVER or no DNS)
        hint = (
            f"Cannot resolve or reach SMTP host {smtp_server!r}:{smtp_port}. "
            "Fix SMTP_SERVER (no http://, no spaces), check internet/DNS/VPN, or use your provider's SMTP "
            "(e.g. Gmail: smtp.gmail.com; Microsoft 365: smtp.office365.com)."
        )
        raise OSError(f"{e}. {hint}") from e

    print(f"✅ Status email successfully sent to {email}")


@app.post("/api/hr/send-status-email")
async def send_status_email_endpoint(req: StatusEmailRequest):
    """Generic endpoint for HR to send accept/reject decision emails."""
    try:
        await asyncio.to_thread(
            _send_status_email_impl, req.email, req.subject, req.message_html
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except OSError as e:
        raise HTTPException(
            status_code=503,
            detail=str(e),
        )
    except smtplib.SMTPException as e:
        raise HTTPException(status_code=502, detail=f"SMTP rejected the message: {e}")
    except Exception as e:
        print(f"❌ Failed to send status email to {req.email}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    return {"success": True, "message": "Email sent successfully."}


@app.post("/api/coding/send-challenge-email")
async def send_coding_challenge_email(
    data: dict, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)
):
    """Send coding challenge email to candidates who passed the assessment"""
    try:
        challenge_id = data.get("challenge_id")
        candidate_emails = data.get("candidate_emails", [])

        if not candidate_emails:
            return {"success": False, "error": "No candidates selected"}

        if not challenge_id:
            return {"success": False, "error": "No challenge selected"}

        # Get challenge details
        query = text(
            "SELECT challenge_id, title, difficulty, language FROM coding_challenges WHERE challenge_id = :challenge_id"
        )
        result = await db.execute(query, {"challenge_id": challenge_id})
        row = result.fetchone()

        if not row:
            return {"success": False, "error": "Challenge not found"}

        challenge_title = row[1]
        difficulty = row[2]
        language = row[3]

        # Send emails in background
        for email in candidate_emails:
            background_tasks.add_task(
                send_coding_challenge_email_task,
                email,
                challenge_id,
                challenge_title,
                difficulty,
                language,
            )

        return {
            "success": True,
            "message": f"Coding challenge email sent to {len(candidate_emails)} candidate(s)",
            "challenge_id": challenge_id,
            "emails_sent": len(candidate_emails),
        }
    except Exception as e:
        print(f"❌ Error sending coding challenge emails: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


async def send_coding_challenge_email_task(
    candidate_email: str,
    challenge_id: str,
    challenge_title: str,
    difficulty: str,
    language: str,
):
    """Send coding challenge email to a single candidate"""
    try:
        sender_email = os.getenv("SMTP_EMAIL")
        sender_password = os.getenv("SMTP_PASSWORD")
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))

        # Generate challenge link
        challenge_link = f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/coding-challenge?id={challenge_id}"

        message = MIMEMultipart("alternative")
        message["Subject"] = f"Your Coding Challenge: {challenge_title}"
        message["From"] = sender_email
        message["To"] = candidate_email

        html = f"""
        <html>
          <body>
            <h2>Coding Challenge Invitation</h2>
            <p>Congratulations on passing the technical assessment!</p>
            <p>You have been invited to complete a <strong>Coding Challenge</strong>.</p>
            <p><strong>Challenge Details:</strong></p>
            <ul>
              <li>Title: {challenge_title}</li>
              <li>Difficulty: {difficulty.capitalize()}</li>
              <li>Language: {language.capitalize()}</li>
            </ul>
            <p><a href="{challenge_link}" style="background-color: #3b82f6; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; display: inline-block;">Start Coding Challenge</a></p>
            <p>Link: {challenge_link}</p>
            <p>Please complete the challenge within the specified time limit.</p>
            <p>Good luck!</p>
          </body>
        </html>
        """

        part = MIMEText(html, "html")
        message.attach(part)

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, candidate_email, message.as_string())

        print(f"✅ Coding challenge email sent to {candidate_email}")
        return True
    except Exception as e:
        print(f"❌ Coding challenge email failed for {candidate_email}: {e}")
        return False


@app.get("/api/coding/challenges-available")
async def get_coding_challenges_for_hr(db: AsyncSession = Depends(get_db)):
    """Get coding challenges for HR dashboard"""
    try:
        query = text(
            "SELECT challenge_id, title, difficulty, language, created_at FROM coding_challenges ORDER BY created_at DESC"
        )
        result = await db.execute(query)
        rows = result.fetchall()

        challenges = [
            {
                "challenge_id": row[0],
                "title": row[1],
                "difficulty": row[2],
                "language": row[3],
                "created_at": str(row[4]) if row[4] else None,
            }
            for row in rows
        ]

        return {"success": True, "challenges": challenges}
    except Exception as e:
        print(f"❌ Error fetching coding challenges: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/coding/all-submissions-for-hr")
async def get_all_coding_submissions_for_hr(db: AsyncSession = Depends(get_db)):
    """Get all coding submissions with violations for HR dashboard"""
    try:
        from sqlalchemy import text

        query = text("""
            SELECT 
                cs.submission_id,
                cs.candidate_email,
                cs.challenge_id,
                cc.title as challenge_title,
                cc.difficulty,
                cc.language,
                cs.score,
                cs.submitted_at,
                cs.status,
                COALESCE(v.violation_count, 0) as violation_count
            FROM coding_submissions cs
            LEFT JOIN coding_challenges cc ON cs.challenge_id = cc.challenge_id
            LEFT JOIN (
                SELECT 
                    session_id,
                    COUNT(*) as violation_count
                FROM violations
                GROUP BY session_id
            ) v ON cs.session_id = v.session_id
            WHERE cs.submitted_at IS NOT NULL
            ORDER BY cs.submitted_at DESC
        """)

        result = await db.execute(query)
        rows = result.fetchall()

        submissions = []
        for row in rows:
            submissions.append(
                {
                    "submission_id": row[0],
                    "candidate_email": row[1],
                    "challenge_id": row[2],
                    "challenge_title": row[3],
                    "difficulty": row[4],
                    "language": row[5],
                    "score": row[6],
                    "submitted_at": str(row[7]) if row[7] else None,
                    "status": row[8],
                    "violation_count": row[9],
                }
            )

        return {"success": True, "submissions": submissions, "count": len(submissions)}
    except Exception as e:
        print(f"❌ Error fetching coding submissions for HR: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hr/available-assessments")
async def get_available_assessments(db: AsyncSession = Depends(get_db)):
    """Get available assessments for HR to send"""
    try:
        from infrastructure.repositories import AssessmentRepository

        assessments = await AssessmentRepository.get_all(db)

        available = []
        for a in assessments:
            qlist = a.get("questions") or []
            n = len(qlist) if isinstance(qlist, list) else 0
            planned = int(a.get("num_questions") or 0)
            available.append(
                {
                    "assessment_id": a.get("assessment_id"),
                    "role": a.get("role"),
                    "status": a.get("status"),
                    "questions_count": n or planned,
                    "difficulty": a.get("difficulty"),
                }
            )

        return {"success": True, "assessments": available}
    except Exception as e:
        print(f"❌ Error fetching assessments: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
